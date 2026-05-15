"""
Router Cost Report — rendicontazione per job vs quotazione, Over/Under analysis.
"""
from fastapi import APIRouter, Depends, HTTPException, Request, Form, Response
from fastapi.responses import HTMLResponse
from typing import Optional
from datetime import date, datetime, time
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from app.database import get_db
from app.models import (
    Job, JobCostLine, Expense, Invoice, InvoiceStatus, JobResourceAssignment,
    Booking, BookingAssignment, BookingExecutionStatus, BookingOvertimeStatus, BookingStatus,
    Resource, WorkingHoursPolicy,
    BillingBatch, BillingBatchStatus, JCLBillingStatus,
    JobDeliverable,
    SupplierInvoice, SupplierInvoiceStatus,
)
from app.services.booking_cost import compute_assignment_breakdown, BookingBreakdown
from app.services.working_hours import get_holidays
from app.services.pdf_export import generate_client_cost_report_pdf
from app.services.cost_line_sync import is_time_based_unit
from app.context import current_tenant_id

router = APIRouter(prefix="/cost-report", tags=["cost_report"])

# Costante condivisa con cost_line_sync.HOURS_PER_DAY: conversione
# day-unit ↔ ore. Tenuta locale qui per evitare import circolari.
HOURS_PER_DAY = 8.0

# v3.5.0-alpha.66.15.2 — tenant scope (R1)


def _tpl():
    from app.main import templates
    return templates


@router.get("/", response_class=HTMLResponse)
async def cost_report_page(request: Request, db: Session = Depends(get_db)):
    # v3.5.0-alpha.16: passiamo solo i jobs per il modal "Assegna risorsa".
    # La lista dei cost report ora viene caricata via /cost-report/api/list
    # con ricerca + filtri (pattern allineato a /quotes).
    # v3.5.0-alpha.66.15.2 — tenant scope (R1)
    jobs = db.query(Job).filter(
        Job.tenant_id == current_tenant_id(),
    ).options(joinedload(Job.client), joinedload(Job.quote)).all()
    return _tpl().TemplateResponse("pages/cost_report.html", {"request": request, "jobs": jobs})


@router.get("/api/list")
async def list_cost_reports(
    db: Session = Depends(get_db),
):
    """Lista riassuntiva di tutti i cost report (1 per job).

    Per ogni job ritorna codice/titolo/cliente/stato + KPI rapidi (totale
    quotato, maturato, stimato, over/under), pensati per la lista filtrabile
    in `/cost-report` (sostituisce il dropdown pre-alpha.16).
    """
    # v3.5.0-alpha.66.15.2 — tenant scope (R1)
    jobs = (
        db.query(Job)
        .options(
            joinedload(Job.client),
            joinedload(Job.quote),
            joinedload(Job.cost_lines),
        )
        .filter(
            Job.tenant_id == current_tenant_id(),
            Job.client_id.isnot(None),
        )
        .order_by(Job.created_at.desc())
        .all()
    )
    # v3.5.0-alpha.60: pre-fetch slice per i 3 totali per-job in singola query
    from app.services.billing_slice_guard import billed_locked_bulk
    all_jcl_ids = []
    for j in jobs:
        for l in j.cost_lines:
            all_jcl_ids.append(l.id)
    billed_map = billed_locked_bulk(db, all_jcl_ids)
    out = []
    for j in jobs:
        total_quoted = sum(l.total_quoted for l in j.cost_lines)
        total_accrued = sum(l.total_accrued for l in j.cost_lines)
        total_expected = sum(l.total_expected for l in j.cost_lines)
        # v3.5.0-alpha.66.21 — α.67 cost-side risorsa
        total_cost_accrued = sum((l.total_cost_accrued or 0.0) for l in j.cost_lines)
        real_margin = round(total_accrued - total_cost_accrued, 2)
        # v3.5.0-alpha.55: convenzione segno positivo = OVER (sforamento).
        over_under_now = round(total_accrued - total_quoted, 2)
        over_under_forecast = round(total_expected - total_quoted, 2)
        # v3.5.0-alpha.60: 3 colonne aggregate per job.
        billed_locked = round(sum(billed_map.get(l.id, 0.0) for l in j.cost_lines), 2)
        accrued_post_period = round(max(0.0, total_accrued - billed_locked), 2)
        forecast_future = round(max(0.0, total_expected - total_accrued), 2)
        out.append({
            "id": j.id,
            "code": j.code,
            "title": j.title,
            "status": j.status,
            "client_id": j.client_id,
            "client": j.client.name if j.client else None,
            "quote_id": j.quote_id,
            "quote_number": j.quote.number if j.quote else None,
            "quote_title": j.quote.title if j.quote else None,
            "start_date": str(j.start_date) if j.start_date else None,
            "end_date": str(j.end_date) if j.end_date else None,
            "total_quoted": round(total_quoted, 2),
            "total_accrued": round(total_accrued, 2),
            "total_expected": round(total_expected, 2),
            # v3.5.0-alpha.66.21 — α.67 cost-side
            "total_cost_accrued": round(total_cost_accrued, 2),
            "real_margin": real_margin,
            "over_under_now": over_under_now,
            "over_under_forecast": over_under_forecast,
            # Alias back-compat (= forecast). Da non usare in nuovi consumer.
            "over_under": over_under_forecast,
            # v3.5.0-alpha.60: 3 colonne basate sulle JCLBilledSlice.
            # billed_locked + accrued_post_period = total_accrued.
            # billed_locked + accrued_post_period + forecast_future = total_expected.
            "billed_locked": billed_locked,
            "accrued_post_period": accrued_post_period,
            "forecast_future": forecast_future,
            "lines_count": len(j.cost_lines),
        })
    return out


def _resource_policy_for_cost(resource: Resource, db: Session) -> Optional[WorkingHoursPolicy]:
    """Helper duplicato volutamente: la versione in planning router non è
    importabile direttamente. Logica triviale, mantieniamo separato."""
    if resource.working_hours_policy_id and resource.working_hours_policy:
        return resource.working_hours_policy
    return db.query(WorkingHoursPolicy).filter(
        WorkingHoursPolicy.is_default == True  # noqa: E712
    ).first()


def _bookings_hours_cost(job_id: int, db: Session, *, client_view: bool = False) -> dict:
    """v3.4.33 — ore + costo derivati dai BOOKING del job (non più Timesheet).

    Se `client_view=True` (v3.5.0-alpha.66.11) esclude i booking attribuiti a
    JobDeliverable la cui JCL è non-time-based (pc/min/TB/shot/version/allow):
    quelle ore sono "produzione interna del file" e diventano hardcost interno
    (vedi `deliverable_hardcost_internal` nel response). Il cliente NON le vede
    come monte ore fatturabili. Default False = vista interna (tutto).

    Ritorna:
      - total_hours: ore totali pianificate (escluse cancellate e pool not_done)
      - total_cost: costo equivalente = weighted_factor × rate giornaliero/8 (per ora)
      - breakdown_total: BookingBreakdown aggregato
      - by_resource: lista breakdown per-risorsa con rate e costo
    """
    # v3.5.0-alpha.66.11: pre-calcolo l'insieme dei deliverable_id da escludere
    # (deliverable la cui JCL associata è non-time-based)
    excluded_dlv_ids: set[int] = set()
    if client_view:
        from app.services.cost_line_sync import is_time_based_unit
        deliverables = (
            db.query(JobDeliverable, JobCostLine)
            .outerjoin(JobCostLine, JobDeliverable.job_cost_line_id == JobCostLine.id)
            .filter(
                JobDeliverable.job_id == job_id,
                JobDeliverable.deleted_at.is_(None),
            )
            .all()
        )
        for d, jcl in deliverables:
            # Deliverable senza JCL: assumiamo non-billable (esclude dal cliente)
            if jcl is None or not is_time_based_unit(jcl.unit):
                excluded_dlv_ids.add(d.id)

    bookings = (
        db.query(Booking).options(
            joinedload(Booking.assignments).joinedload(BookingAssignment.resource),
        )
        .filter(
            Booking.job_id == job_id,
            Booking.status != BookingStatus.cancelled,
        )
        .all()
    )
    if client_view and excluded_dlv_ids:
        bookings = [b for b in bookings if b.job_deliverable_id not in excluded_dlv_ids]
    totals = BookingBreakdown()
    by_resource: dict[int, dict] = {}
    holidays_cache: dict = {}

    def _hols(policy, y0, y1):
        key = (id(policy), y0, y1)
        if key not in holidays_cache:
            holidays_cache[key] = get_holidays(policy, y0, y1)
        return holidays_cache[key]

    for b in bookings:
        for a in b.assignments:
            if not a.resource:
                continue
            policy = _resource_policy_for_cost(a.resource, db)
            if not policy:
                continue
            hols = _hols(policy, a.start_datetime.year, a.end_datetime.year)
            br = compute_assignment_breakdown(a, policy, hols, b)
            totals.add(br)
            rmap = by_resource.setdefault(a.resource.id, {
                "resource_id": a.resource.id,
                "resource_name": a.resource.name,
                "daily_rate": a.resource.daily_rate or 0,
                "hourly_rate": a.resource.hourly_rate or 0,
                "breakdown": BookingBreakdown(),
            })
            rmap["breakdown"].add(br)

    total_cost = 0.0
    by_res_list = []
    for r in by_resource.values():
        b = r["breakdown"]
        # rate orario: se hourly_rate noto, usa quello; altrimenti daily/8
        rate_h = r["hourly_rate"] or (r["daily_rate"] / 8 if r["daily_rate"] else 0)
        cost = b.weighted_factor * rate_h
        total_cost += cost
        by_res_list.append({
            "resource_id": r["resource_id"],
            "resource_name": r["resource_name"],
            "daily_rate": r["daily_rate"],
            "hourly_rate": r["hourly_rate"],
            "rate_used_per_hour": round(rate_h, 2),
            "breakdown": b.as_dict(),
            "cost_estimated": round(cost, 2),
        })
    return {
        "total_hours": round(totals.total_hours, 2),
        "total_cost": round(total_cost, 2),
        "breakdown_total": totals.as_dict(),
        "by_resource": by_res_list,
    }


# v3.5.0-alpha.48 — Helper Billing flow (Step 3 Cost Report → Billing).

def _billing_batches_for_job(db: Session, job: Job) -> list[dict]:
    """Lista BillingBatch del progetto del job. Usata dal widget Fatturazione
    nel cost report per mostrare i batch già trasmessi/approvati/fatturati.

    Mostra TUTTI gli stati (incluso cancelled) per audit completo. Il client
    UI nasconde i cancelled di default."""
    if not job.project_id:
        return []
    batches = (
        db.query(BillingBatch)
        .options(joinedload(BillingBatch.invoice))
        .filter(BillingBatch.project_id == job.project_id)
        .order_by(BillingBatch.transmitted_at.desc())
        .all()
    )
    return [
        {
            "id": b.id,
            "code": b.code,
            "status": b.status.value,
            "period_start": str(b.period_start) if b.period_start else None,
            "period_end": str(b.period_end) if b.period_end else None,
            "total_proposed": round(b.total_proposed or 0, 2),
            "total_approved": round(b.total_approved or 0, 2),
            "total_lost": round(b.total_lost or 0, 2),
            "transmitted_at": b.transmitted_at.isoformat() if b.transmitted_at else None,
            "approved_at": b.approved_at.isoformat() if b.approved_at else None,
            "invoice_id": b.invoice_id,
            "invoice_number": b.invoice.number if b.invoice else None,
        }
        for b in batches
    ]


def _billing_summary_for_job(job: Job) -> dict:
    """Aggregati JCL per stato fatturazione. Usato dalla card Fatturazione
    nell'header del cost report per mostrare a colpo d'occhio quanto è
    not_billed / in_batch / billed / paid / lost."""
    summary = {
        "not_billed": 0.0,
        "in_batch": 0.0,
        "billed": 0.0,
        "paid": 0.0,
        "lost": 0.0,
    }
    for l in job.cost_lines:
        st = l.billing_status.value if l.billing_status else "not_billed"
        # billed/paid uso billed_amount (post-modifica manager); altri usano accrued
        amt = (l.billed_amount if st in ("billed", "paid") and l.billed_amount is not None
               else l.total_accrued)
        if st in summary:
            summary[st] += amt or 0
    return {k: round(v, 2) for k, v in summary.items()}


@router.get("/api/job/{job_id}")
async def job_cost_report(job_id: int, db: Session = Depends(get_db)):
    """Report completo: quotazione vs reale (accrued/expected), Over/Under.

    v3.4.33: la fonte ore lavorate è ora **Booking** (non più Timesheet).
    Il campo legacy `hours_cost` (da Timesheet) è ancora esposto per back-compat
    ma rinominato `hours_cost_legacy_timesheet` per esplicitarne la deprecazione.
    Nuovo campo `bookings_hours_cost` è quello canonico.
    """
    job = db.query(Job).options(
        joinedload(Job.client),
        joinedload(Job.quote),
        joinedload(Job.cost_lines).joinedload(JobCostLine.price_item),
        joinedload(Job.resource_assignments).joinedload(JobResourceAssignment.resource),
    ).filter(Job.id == job_id).first()
    if not job: raise HTTPException(404, "Job non trovato")

    # v3.4.33 — Ore + costo derivati dai BOOKING (canonico)
    bk_data = _bookings_hours_cost(job_id, db)

    # v3.4.38 (R3.5): Timesheet legacy DROPPATO dal cost report.
    # Memoria architetturale: cost report = quote+booking+hardcost (binario
    # cliente/finance), Timesheet/TimePunch = HR (binario consulente del lavoro).
    # I campi `hours_cost_legacy_timesheet` e `hours_cost` non sono più
    # esposti nel response da v3.4.38.

    # Spese
    total_expenses = db.query(func.sum(Expense.amount)).filter(
        Expense.job_id == job_id).scalar() or 0

    # Fatturato
    invoiced = db.query(func.sum(Invoice.total)).filter(
        Invoice.job_id == job_id,
        Invoice.status.in_([InvoiceStatus.sent, InvoiceStatus.paid])
    ).scalar() or 0
    paid = db.query(func.sum(Invoice.total)).filter(
        Invoice.job_id == job_id, Invoice.status == InvoiceStatus.paid
    ).scalar() or 0

    # Totali cost lines
    total_quoted = sum(l.total_quoted for l in job.cost_lines)
    total_accrued = sum(l.total_accrued for l in job.cost_lines)
    total_expected = sum(l.total_expected for l in job.cost_lines)

    # v3.5.0-alpha.66.11 — Hardcost ORE DELIVERABLE per cost report INTERNO.
    # Per ogni JobDeliverable del job, somma (booking.hours × Resource.internal_cost_hourly).
    # Aggregato per JobCostLine per mostrare il valore per riga + totale job.
    # Il cost report CLIENTE non vede questo dato (solo qty + descrizione).
    deliverables = (
        db.query(JobDeliverable)
        .filter(
            JobDeliverable.job_id == job_id,
            JobDeliverable.deleted_at.is_(None),
        )
        .all()
    )
    deliverable_hardcost_by_jcl: dict[int, float] = {}
    deliverable_hours_by_jcl: dict[int, float] = {}
    deliverable_count_by_jcl: dict[int, int] = {}
    total_deliverable_hardcost = 0.0
    total_deliverable_hours = 0.0
    for d in deliverables:
        # Conta i deliverable per JCL (anche se ore=0)
        if d.job_cost_line_id:
            deliverable_count_by_jcl[d.job_cost_line_id] = (
                deliverable_count_by_jcl.get(d.job_cost_line_id, 0) + 1
            )
        # Aggrega ore + costo dai booking attribuiti
        bks_d = (
            db.query(Booking)
            .filter(
                Booking.job_deliverable_id == d.id,
                Booking.status != BookingStatus.cancelled,
            )
            .all()
        )
        for b in bks_d:
            for a in (b.assignments or []):
                if not (a.start_datetime and a.end_datetime):
                    continue
                hours = (a.end_datetime - a.start_datetime).total_seconds() / 3600.0
                cost_h = a.resource.internal_cost_hourly if a.resource else None
                eur = hours * (cost_h or 0.0)
                total_deliverable_hours += hours
                total_deliverable_hardcost += eur
                if d.job_cost_line_id:
                    deliverable_hardcost_by_jcl[d.job_cost_line_id] = (
                        deliverable_hardcost_by_jcl.get(d.job_cost_line_id, 0.0) + eur
                    )
                    deliverable_hours_by_jcl[d.job_cost_line_id] = (
                        deliverable_hours_by_jcl.get(d.job_cost_line_id, 0.0) + hours
                    )

    # v3.5.0-alpha.60: 3 colonne basate su slice. Pre-fetch in singola query.
    from app.services.billing_slice_guard import billed_locked_bulk, three_column_view
    line_ids = [l.id for l in job.cost_lines]
    billed_map = billed_locked_bulk(db, line_ids)

    # v3.5.0-alpha.64: pre-fetch quote-line referrals per badge "↪ Riferita".
    # Una sola query JOIN per tutte le cost-line del job.
    from app.models import Quote as _Quote, QuoteLine as _QuoteLine
    refs_rows = []
    if line_ids:
        refs_rows = (
            db.query(_QuoteLine, _Quote)
            .join(_Quote, _QuoteLine.quote_id == _Quote.id)
            .filter(
                _QuoteLine.referred_from_jcl_id.in_(line_ids),
                _Quote.deleted_at.is_(None),
            )
            .all()
        )
    refs_map: dict[int, list[dict]] = {}
    for ql, q in refs_rows:
        refs_map.setdefault(ql.referred_from_jcl_id, []).append({
            "quote_id": q.id,
            "quote_number": q.number,
            "quote_version": q.version,
            "quote_status": (q.status.value if hasattr(q.status, "value") else q.status),
            "quote_url": f"/quotes#{q.id}",
            "line_id": ql.id,
            "line_total": ql.total,
        })
    sum_billed_locked = round(sum(billed_map.get(lid, 0.0) for lid in line_ids), 2)
    sum_accrued_post_period = round(max(0.0, total_accrued - sum_billed_locked), 2)
    sum_forecast_future = round(max(0.0, total_expected - total_accrued), 2)

    # v3.5.0-alpha.65 — Pending OT per riga JCL: ore overtime in attesa di
    # approvazione, non ancora pesate. Mostrate in tooltip nel cost-report
    # ("+€X pending") senza alterare i numeri certi (decisione semantica:
    # "solo APPROVED applica moltiplicatori"). Aggrega i booking del JCL con
    # overtime_status=pending e somma le ore via compute_assignment_breakdown
    # (campo pending_overtime_hours del breakdown).
    pending_ot_by_jcl: dict[int, float] = {}
    if line_ids:
        from app.services.booking_cost import compute_assignment_breakdown
        pending_bookings = (
            db.query(Booking)
            .options(joinedload(Booking.assignments).joinedload(BookingAssignment.resource))
            .filter(
                Booking.job_cost_line_id.in_(line_ids),
                Booking.overtime_status == BookingOvertimeStatus.pending,
                Booking.status != BookingStatus.cancelled,
            )
            .all()
        )
        _hols_cache: dict = {}
        for b in pending_bookings:
            for a in b.assignments:
                if not a.resource:
                    continue
                policy = _resource_policy_for_cost(a.resource, db)
                if not policy:
                    continue
                key = (id(policy), a.start_datetime.year, a.end_datetime.year)
                if key not in _hols_cache:
                    _hols_cache[key] = get_holidays(
                        policy, a.start_datetime.year, a.end_datetime.year
                    )
                br = compute_assignment_breakdown(a, policy, _hols_cache[key], b)
                pending_ot_by_jcl[b.job_cost_line_id] = (
                    pending_ot_by_jcl.get(b.job_cost_line_id, 0.0)
                    + br.pending_overtime_hours
                )

    # v3.5.0-alpha.68 — Σ fatture passive linkate al job (diretto o via
    # project) non cancelled. Contribuisce al margine reale completo
    # come hardcost esterno (costo per il tenant pagato a fornitori).
    sup_q = db.query(SupplierInvoice).filter(
        SupplierInvoice.tenant_id == current_tenant_id(),
        SupplierInvoice.deleted_at.is_(None),
        SupplierInvoice.payment_status != SupplierInvoiceStatus.cancelled,
    )
    if job.project_id:
        from sqlalchemy import or_ as _or
        sup_q = sup_q.filter(_or(
            SupplierInvoice.job_id == job_id,
            SupplierInvoice.project_id == job.project_id,
        ))
    else:
        sup_q = sup_q.filter(SupplierInvoice.job_id == job_id)
    total_supplier_invoices = sum(
        (i.amount_total or 0) for i in sup_q.all()
    )

    # v3.4.36 (R1.4): margine dinamico = Σ JobCostLine.total_quoted (vivo)
    # − (costo booking + spese). Non più contro Job.budget_quoted statico
    # (quello resta come riferimento "originale" all'approvazione, ma può
    # divergere da total_quoted quando si aggiungono extra cost lines).
    estimated_cost = bk_data["total_cost"] + (total_expenses or 0)
    margin = total_quoted - estimated_cost

    return {
        "job": {
            "id": job.id, "code": job.code, "title": job.title,
            "status": job.status,
            "client": job.client.name if job.client else None,
            "client_id": job.client_id,
            # v3.5.0-alpha.48: project_id necessario per trasmissione billing
            "project_id": job.project_id,
            # v3.5.0-alpha.88: project_title per scope label in drill risorse
            "project_title": (job.project.title if getattr(job, "project", None) else None),
            "budget_quoted": job.budget_quoted,
            "start_date": str(job.start_date) if job.start_date else None,
            "end_date": str(job.end_date) if job.end_date else None,
            # v3.5.0-alpha.65 — Pass-through OT al cliente (opt-in).
            "weighted_revenue": bool(getattr(job, "weighted_revenue", False)),
        },
        "summary": {
            "budget_quoted": round(job.budget_quoted, 2),
            "total_quoted": round(total_quoted, 2),
            "total_accrued": round(total_accrued, 2),
            "total_expected": round(total_expected, 2),
            # v3.5.0-alpha.66.21 — α.67 cost-side risorsa.
            # total_cost_accrued = Σ JCL.total_cost_accrued (ore done ×
            #   Resource.internal_cost_hourly). 0 se nessuna risorsa con
            #   cost_type configurato → "real_margin" sarà = total_accrued.
            # real_margin = total_accrued − total_cost_accrued (margine reale
            #   stimato basato sui rate interni configurati).
            "total_cost_accrued": round(sum((l.total_cost_accrued or 0) for l in job.cost_lines), 2),
            "real_margin": round(
                total_accrued - sum((l.total_cost_accrued or 0) for l in job.cost_lines), 2
            ),
            # v3.5.0-alpha.68 — Fatture passive (Supplier) linkate al job
            # o al project. real_margin_full sottrae anche queste.
            "total_supplier_invoices": round(total_supplier_invoices, 2),
            "real_margin_full": round(
                total_accrued
                - sum((l.total_cost_accrued or 0) for l in job.cost_lines)
                - total_supplier_invoices, 2
            ),
            # v3.5.0-alpha.55: due viste di Over/Under.
            # NOW = maturato − quotato (extracosto certo, base fatturazione).
            # FORECAST = stima − quotato (sforamento previsto su base
            # pianificato, base report cliente).
            # Convenzione segno: positivo = OVER (sforamento, problema),
            # negativo = UNDER (sotto budget, ok).
            "over_under_now": round(total_accrued - total_quoted, 2),
            "over_under_forecast": round(total_expected - total_quoted, 2),
            # Back-compat: vecchio campo `over_under` lasciato come alias di
            # forecast (con segno invertito ex-API). Da non usare in nuovi
            # consumer: leggere over_under_now / over_under_forecast.
            "over_under": round(total_expected - total_quoted, 2),
            # v3.5.0-alpha.60: 3 colonne aggregate per job.
            # billed_locked = Σ slice (chiuso in fattura, immutabile).
            # accrued_post_period = maturato non ancora fatturato (≈ ore done
            #   senza slice → prossimo candidato di trasmissione).
            # forecast_future = stima ulteriori ore ancora da lavorare.
            "billed_locked": sum_billed_locked,
            "accrued_post_period": sum_accrued_post_period,
            "forecast_future": sum_forecast_future,
            "total_expenses": round(total_expenses, 2),
            # Canonico v3.4.33
            "bookings_hours": bk_data["total_hours"],
            "bookings_hours_cost": bk_data["total_cost"],
            "estimated_cost": round(estimated_cost, 2),
            "margin": round(margin, 2),
            "invoiced": round(invoiced, 2),
            "paid": round(paid, 2),
            # v3.5.0-alpha.66.11 — Hardcost ore deliverable (solo INTERNO).
            # Cliente non vede. Da mostrare in /cost-report e /jobs/{id} solo
            # con permesso view_finance.
            "deliverable_hardcost_internal": round(total_deliverable_hardcost, 2),
            "deliverable_hours_internal": round(total_deliverable_hours, 2),
            "deliverable_count": len(deliverables),
        },
        # v3.4.33 — Breakdown per fascia + per-risorsa dai booking
        "bookings_breakdown": bk_data["breakdown_total"],
        "bookings_by_resource": bk_data["by_resource"],
        "cost_lines": [
            {
                "id": l.id, "description": l.description, "unit": l.unit,
                # v3.5.0-alpha.66.11 — Categorizzazione unit per cost report split.
                # time-based (day/hr) → mostra ore al cliente. non-time (pc/min/TB/...)
                # → cliente vede solo qty + descrizione, no monte ore.
                "unit_is_time_based": is_time_based_unit(l.unit),
                # Hardcost ore deliverable attribuiti a questa JCL (interno only)
                "deliverable_hardcost_internal": round(deliverable_hardcost_by_jcl.get(l.id, 0.0), 2),
                "deliverable_hours_internal": round(deliverable_hours_by_jcl.get(l.id, 0.0), 2),
                "deliverable_count": deliverable_count_by_jcl.get(l.id, 0),
                "quantity_quoted": l.quantity_quoted,
                "quantity_actual": l.quantity_actual,
                # v3.5.0-alpha.55: quantity_planned derivata (qty_planned =
                # total_expected / unit_price quando unit_price > 0).
                "quantity_planned": (round(l.total_expected / l.unit_price, 4)
                                     if l.unit_price else 0.0),
                "unit_price": l.unit_price,
                "total_quoted": l.total_quoted,
                "total_accrued": l.total_accrued,
                "total_expected": l.total_expected,
                # v3.5.0-alpha.66.21 — α.67 cost-side
                "total_cost_accrued": round(l.total_cost_accrued or 0.0, 2),
                "real_margin": round((l.total_accrued or 0.0) - (l.total_cost_accrued or 0.0), 2),
                # v3.5.0-alpha.55: Now = maturato−quotato; Forecast = stima−quotato.
                # Positivo = OVER (sforamento), negativo = UNDER (sotto budget).
                "over_under_now": round(l.total_accrued - l.total_quoted, 2),
                "over_under_forecast": round(l.total_expected - l.total_quoted, 2),
                # Alias back-compat (= forecast). Da non usare in nuovi consumer.
                "over_under": round(l.total_expected - l.total_quoted, 2),
                "is_billable": l.is_billable,
                "is_extra": l.is_extra,
                "category": l.price_item.category.name if l.price_item and l.price_item.category else None,
                # v3.5.0-alpha.48 — Step 3 Cost Report → Billing flow:
                # esponiamo lo stato fatturazione + l'importo realmente
                # fatturato (può divergere da total_accrued se manager ha
                # ridotto in approvazione, delta finisce in LossEntry).
                "billing_status": l.billing_status.value if l.billing_status else "not_billed",
                "billing_batch_id": l.billing_batch_id,
                "billed_amount": l.billed_amount,
                # v3.5.0-alpha.60: 3 colonne per riga (slice-based).
                **three_column_view(l, billed_map.get(l.id, 0.0)),
                # v3.5.0-alpha.64: lista quote-line che referenziano questa JCL
                # (refer-to-sales). UI mostra badge "↪ Riferita su Q-NNN-NN v2".
                "referrals": refs_map.get(l.id, []),
                # v3.5.0-alpha.65: ore overtime pending sul JCL — informativa
                # in tooltip ("+€X pending"), NON conteggiate nel maturato.
                # L'amount è una stima del delta di maturato post-approvazione
                # SE il job ha weighted_revenue=True (altrimenti = 0, perché
                # OT non gonfia il revenue lineare).
                "pending_overtime_hours": round(pending_ot_by_jcl.get(l.id, 0.0), 2),
                "pending_overtime_amount": round(
                    (pending_ot_by_jcl.get(l.id, 0.0)
                     / (HOURS_PER_DAY if (l.unit or "").lower() in ("day","giorno","giornate","giornata","d") else 1.0)
                     * (l.unit_price or 0.0))
                    if getattr(job, "weighted_revenue", False) else 0.0,
                    2,
                ),
            }
            for l in job.cost_lines
        ],
        # v3.5.0-alpha.48: elenco BillingBatch del progetto per dare contesto
        # nel widget Fatturazione del cost report (totali per stato + link).
        # Solo se il job ha un progetto associato. Ordine cronologico desc.
        "billing_batches": _billing_batches_for_job(db, job),
        # Aggregati per la card Fatturazione
        "billing_summary": _billing_summary_for_job(job),
        "resource_assignments": [
            {
                "resource": a.resource.name if a.resource else None,
                "type": a.resource.type if a.resource else None,
                "role": a.role_in_project,
                "planned_days": a.planned_days,
                "planned_hours": a.planned_hours,
                "agreed_daily_rate": a.agreed_daily_rate,
                "agreed_hourly_rate": a.agreed_hourly_rate,
            }
            for a in job.resource_assignments
        ],
        # v3.4.38 (R3.5): timesheet_summary rimosso. Le ore lavorate sono in
        # bookings_breakdown e bookings_by_resource (canonico Booking).
    }


@router.put("/api/job/{job_id}/weighted-revenue")
async def set_weighted_revenue(
    job_id: int,
    enabled: bool = Form(...),
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.65 — Toggle pass-through OT al cliente per il job.

    Quando attivato, il maturato cliente (`JobCostLine.total_accrued`) viene
    ricalcolato usando il `weighted_factor` di compute_assignment_breakdown:
    overtime/notte/domenica/festivo gonfiano il revenue. Default OFF (giornate
    fisiche). Disattivare ripristina il calcolo lineare.

    Effetto: il flag è persistito sul Job + automatic reconcile-actuals di
    tutte le cost-line (così il cliente vede subito i nuovi numeri). Idempotente.
    """
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job non trovato")
    job.weighted_revenue = bool(enabled)
    db.flush()
    from app.services.cost_line_sync import recompute_for_job
    result = recompute_for_job(db, job_id)
    db.commit()
    return {
        "job_id": job_id,
        "weighted_revenue": job.weighted_revenue,
        "lines_updated": result.get("lines_updated", 0),
    }


@router.post("/api/job/{job_id}/reconcile-actuals")
async def reconcile_actuals(job_id: int, db: Session = Depends(get_db)):
    """v3.4.41 — Ricomputa quantity_actual + total_accrued di tutte le
    JobCostLine del job aggregando i Booking con execution_status=done.
    Utile per:
      - Fix retroattivo su DB esistenti (booking marcati done prima della
        v3.4.41 quando il sync non era ancora hookato).
      - Riconciliazione manuale da UI ("Aggiorna ore") quando si vuole
        sincronizzare il consuntivo dal flusso operativo.
    Idempotente."""
    from app.services.cost_line_sync import recompute_for_job
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job non trovato")
    result = recompute_for_job(db, job_id)
    db.commit()
    return result


# v3.5.0-alpha.113 — bulk reconcile per allineare lista CR.
# Q5: la lista mostra `total_accrued` stored; il dettaglio chiama
# reconcile e aggiorna stored → ritornando alla lista i numeri appaiono
# allineati. Causa: hook recompute_for_booking non copre tutti i path
# che modificano il booking (es. shift bulk multi-select, import,
# punch import). Fix: bulk reconcile lazy all'apertura della lista CR.
@router.post("/api/reconcile-all")
async def reconcile_all_jobs(db: Session = Depends(get_db)):
    """Ricomputa tutti i job del tenant. Idempotente. Più costoso del
    singolo job ma deterministico — invoke su page-load lista CR."""
    from app.services.cost_line_sync import recompute_for_job
    job_ids = [
        j.id for j in db.query(Job.id).filter(
            Job.tenant_id == current_tenant_id(),
            Job.client_id.isnot(None),
        ).all()
    ]
    total_updated = 0
    for jid in job_ids:
        r = recompute_for_job(db, jid)
        total_updated += r.get("lines_updated", 0)
    db.commit()
    return {"jobs": len(job_ids), "lines_updated": total_updated}


@router.put("/api/job/{job_id}/cost-lines/{line_id}")
async def update_cost_line(
    job_id: int, line_id: int,
    quantity_actual: Optional[float] = Form(None),
    total_accrued: Optional[float] = Form(None),
    total_expected: Optional[float] = Form(None),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.10: rimosso override manuale di `quantity_actual` /
    `total_accrued` (decisione architetturale Matteo). I valori sono sempre
    derivati dai booking marcati `done` (cost_line_sync). Se passati, 422.
    Restano editabili `total_expected` (forecast finance) e `notes`.
    """
    if quantity_actual is not None or total_accrued is not None:
        raise HTTPException(
            422,
            "Le ore lavorate / il maturato non sono più modificabili manualmente. "
            "Derivano sempre dai booking marcati 'done'. La fatturazione di "
            "extra/scontistica/banca-ore passerà dal flusso fatturazione (in roadmap)."
        )
    line = db.query(JobCostLine).filter(
        JobCostLine.id == line_id, JobCostLine.job_id == job_id).first()
    if not line: raise HTTPException(404)
    if total_expected is not None: line.total_expected = total_expected
    if notes is not None: line.notes = notes
    db.commit()
    return {"id": line.id, "total_accrued": line.total_accrued, "total_expected": line.total_expected}


@router.post("/api/job/{job_id}/assign-resource")
async def assign_resource(
    job_id: int,
    request: Request,
    resource_id: int = Form(...),
    role_in_project: Optional[str] = Form(None),
    planned_days: Optional[float] = Form(None),
    planned_hours: Optional[float] = Form(None),
    agreed_daily_rate: Optional[float] = Form(None),
    agreed_hourly_rate: Optional[float] = Form(None),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.10: gate `assign_resources` (admin/manager/producer).
    Editor/operator NON possono assegnare risorse a un job."""
    from app.services.rbac import current_user_optional, can_assign_resources
    user = current_user_optional(request)
    if not can_assign_resources(user):
        raise HTTPException(403, "Non hai il permesso di assegnare risorse")
    assignment = JobResourceAssignment(
        job_id=job_id, resource_id=resource_id,
        role_in_project=role_in_project,
        planned_days=planned_days, planned_hours=planned_hours,
        agreed_daily_rate=agreed_daily_rate, agreed_hourly_rate=agreed_hourly_rate,
        notes=notes,
    )
    db.add(assignment); db.commit(); db.refresh(assignment)
    return {"id": assignment.id}


@router.delete("/api/job/{job_id}/assign-resource/{assignment_id}")
async def remove_resource_assignment(
    job_id: int, assignment_id: int, request: Request, db: Session = Depends(get_db)):
    """v3.5.0-alpha.10: gate `assign_resources`."""
    from app.services.rbac import current_user_optional, can_assign_resources
    user = current_user_optional(request)
    if not can_assign_resources(user):
        raise HTTPException(403, "Non hai il permesso di rimuovere risorse")
    a = db.query(JobResourceAssignment).filter(
        JobResourceAssignment.id == assignment_id, JobResourceAssignment.job_id == job_id).first()
    if not a: raise HTTPException(404)
    db.delete(a); db.commit()
    return {"ok": True}


# ── Booking executive cost report (v3.4.32) ─────────────────────────
# Aggrega i booking di un job per fascia (regular / overtime / pending /
# notturno / festivo / domenica) + pool not_done (booking con
# execution_status=not_done & count_in_costs=False, escluse dai totali).
#
# Nota strategica (vedi memoria project_costreport_vs_timesheet.md):
# il cost report parte dai BOOKING, non dai TimePunch. Le timbrature sono
# binario HR separato. Questo endpoint è il primo passo verso il rifacimento
# del cost report; per ora coabita con il vecchio /api/job/{id} basato
# su Timesheet.

def _resource_policy(resource: Resource, db: Session) -> Optional[WorkingHoursPolicy]:
    if resource.working_hours_policy_id and resource.working_hours_policy:
        return resource.working_hours_policy
    return db.query(WorkingHoursPolicy).filter(
        WorkingHoursPolicy.is_default == True  # noqa: E712
    ).first()


@router.get("/api/job/{job_id}/booking-summary")
async def job_booking_summary(
    job_id: int,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    resource_id: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Aggrega i booking del job per fascia oraria + pool not_done.

    Ritorna:
      - totals: BookingBreakdown cumulato per booking conteggiati
      - pending_overtime: lista booking con overtime_status=pending (azione richiesta)
      - not_done_pool: lista booking not_done con count_in_costs=False (pool)
      - by_resource: breakdown per risorsa (per riga finanziaria)

    v3.5.0-alpha.86 — Filtri opzionali sezione (S3.3):
    - from_date/to_date: limita ai booking che intersecano il range
    - resource_id: comma-separated list di Resource.id; le ore di altre risorse
      sono escluse dal calcolo (per drill-down su specifiche persone/sale)
    """
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job non trovato")

    # Parse resource_id list (comma-separated)
    resource_ids: Optional[set[int]] = None
    if resource_id:
        try:
            resource_ids = {int(x.strip()) for x in resource_id.split(",") if x.strip()}
        except ValueError:
            resource_ids = None

    q = (
        db.query(Booking)
        .options(
            joinedload(Booking.assignments).joinedload(BookingAssignment.resource),
            joinedload(Booking.cost_line),
        )
        .filter(
            Booking.job_id == job_id,
            Booking.status != BookingStatus.cancelled,
        )
    )
    if from_date:
        q = q.filter(Booking.end_datetime >= datetime.combine(from_date, time.min))
    if to_date:
        q = q.filter(Booking.start_datetime <= datetime.combine(to_date, time.max))
    bookings = q.all()

    totals = BookingBreakdown()
    by_resource: dict[int, dict] = {}
    pending_overtime = []
    not_done_pool = []

    # holidays cache per anno (compute una sola volta)
    holidays_cache: dict[tuple[int, int], set] = {}

    def _get_hol(policy, y0, y1):
        key = (id(policy), y0, y1)
        if key not in holidays_cache:
            holidays_cache[key] = get_holidays(policy, y0, y1)
        return holidays_cache[key]

    for b in bookings:
        # Pending overtime: tracciato a parte (mostra ore stimate ma non sommate)
        if b.overtime_status == BookingOvertimeStatus.pending:
            pending_overtime.append({
                "booking_id": b.id,
                "start": b.start_datetime.isoformat() if b.start_datetime else None,
                "end": b.end_datetime.isoformat() if b.end_datetime else None,
                "cost_line": b.cost_line.description if b.cost_line else None,
                "resources": [a.resource.name for a in b.assignments if a.resource],
                "notes": b.notes,
            })

        # Pool not_done: tracciato a parte
        if (b.execution_status == BookingExecutionStatus.not_done
            and not b.count_in_costs):
            tot_h = sum(
                (a.end_datetime - a.start_datetime).total_seconds() / 3600.0
                for a in b.assignments
            )
            not_done_pool.append({
                "booking_id": b.id,
                "start": b.start_datetime.isoformat() if b.start_datetime else None,
                "end": b.end_datetime.isoformat() if b.end_datetime else None,
                "cost_line": b.cost_line.description if b.cost_line else None,
                "resources": [a.resource.name for a in b.assignments if a.resource],
                "reason": b.not_done_reason,
                "total_hours": round(tot_h, 2),
            })
            # Saltato dal totals (vedi compute_assignment_breakdown che riconosce il pool)

        # Aggrega comunque tramite breakdown (gestisce internamente pool/pending)
        for a in b.assignments:
            if not a.resource:
                continue
            # v3.5.0-alpha.86 S3.3 — filtro risorse opzionale
            if resource_ids is not None and a.resource.id not in resource_ids:
                continue
            policy = _resource_policy(a.resource, db)
            if not policy:
                continue
            hols = _get_hol(policy, a.start_datetime.year, a.end_datetime.year)
            br = compute_assignment_breakdown(a, policy, hols, b)
            totals.add(br)
            rmap = by_resource.setdefault(a.resource.id, {
                "resource_id": a.resource.id,
                "resource_name": a.resource.name,
                "rate": a.resource.daily_rate or a.resource.hourly_rate or 0,
                "breakdown": BookingBreakdown(),
            })
            rmap["breakdown"].add(br)

    return {
        "job": {"id": job.id, "code": job.code, "title": job.title},
        "totals": totals.as_dict(),
        "by_resource": [
            {
                "resource_id": r["resource_id"],
                "resource_name": r["resource_name"],
                "rate": r["rate"],
                "breakdown": r["breakdown"].as_dict(),
            }
            for r in by_resource.values()
        ],
        "pending_overtime": pending_overtime,
        "not_done_pool": not_done_pool,
    }


@router.get("/api/resource/{resource_id}/jobs")
async def resource_job_drill(
    resource_id: int,
    period_start: Optional[date] = None,
    period_end: Optional[date] = None,
    project_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.69.2 — Drill-down inverso: per una risorsa, mostra i
    job lavorati con breakdown ore + costo per ogni job. Reverse della
    vista per-line nel cost-report (`bookings_by_resource` aggrega per
    risorsa dentro un job; questo aggrega per job di una risorsa).

    v3.5.0-alpha.88: filtro opzionale project_id per limitare al solo
    progetto corrente quando il drill parte da /cost-report."""
    resource = db.query(Resource).filter(
        Resource.id == resource_id,
        Resource.tenant_id == current_tenant_id(),
    ).first()
    if not resource:
        raise HTTPException(404, "Risorsa non trovata")
    q = (
        db.query(Booking)
        .options(
            joinedload(Booking.assignments).joinedload(BookingAssignment.resource),
            joinedload(Booking.job).joinedload(Job.client),
            joinedload(Booking.cost_line),
        )
        .join(BookingAssignment, BookingAssignment.booking_id == Booking.id)
        .join(Job, Job.id == Booking.job_id)
        .filter(
            BookingAssignment.resource_id == resource_id,
            Booking.status != BookingStatus.cancelled,
            Booking.job_id.isnot(None),
        )
    )
    if project_id:
        q = q.filter(Job.project_id == project_id)
    if period_start:
        q = q.filter(Booking.start_datetime >= period_start)
    if period_end:
        from datetime import timedelta as _td
        q = q.filter(Booking.start_datetime < period_end + _td(days=1))
    bookings = q.all()
    by_job: dict[int, dict] = {}
    holidays_cache: dict = {}
    def _hols(policy, y0, y1):
        key = (id(policy), y0, y1)
        if key not in holidays_cache:
            holidays_cache[key] = get_holidays(policy, y0, y1)
        return holidays_cache[key]
    for b in bookings:
        if not b.job:
            continue
        for a in b.assignments:
            if a.resource_id != resource_id or not a.resource:
                continue
            policy = _resource_policy_for_cost(a.resource, db)
            if not policy:
                continue
            hols = _hols(policy, a.start_datetime.year, a.end_datetime.year)
            br = compute_assignment_breakdown(a, policy, hols, b)
            entry = by_job.setdefault(b.job.id, {
                "job_id": b.job.id,
                "job_code": b.job.code,
                "job_title": b.job.title,
                "job_status": b.job.status,
                "client_name": b.job.client.name if b.job.client else None,
                "client_id": b.job.client_id,
                "breakdown": BookingBreakdown(),
                "bookings_count": 0,
                "first_date": None,
                "last_date": None,
            })
            entry["breakdown"].add(br)
            entry["bookings_count"] += 1
            d = a.start_datetime.date()
            if entry["first_date"] is None or d < entry["first_date"]:
                entry["first_date"] = d
            if entry["last_date"] is None or d > entry["last_date"]:
                entry["last_date"] = d
    rate_h = resource.hourly_rate or (
        resource.daily_rate / 8 if resource.daily_rate else 0
    )
    cost_h = resource.internal_cost_hourly or 0
    out = []
    for e in by_job.values():
        bd = e["breakdown"].as_dict()
        weighted = bd.get("weighted_factor", 0) or 0
        est_cost = round(weighted * rate_h, 2)
        real_cost = round(weighted * cost_h, 2) if cost_h else None
        out.append({
            "job_id": e["job_id"],
            "job_code": e["job_code"],
            "job_title": e["job_title"],
            "job_status": e["job_status"],
            "client_name": e["client_name"],
            "client_id": e["client_id"],
            "first_date": str(e["first_date"]) if e["first_date"] else None,
            "last_date": str(e["last_date"]) if e["last_date"] else None,
            "bookings_count": e["bookings_count"],
            "breakdown": bd,
            "total_hours": round(bd.get("total_hours", 0) or 0, 2),
            "weighted_hours": round(weighted, 2),
            "cost_estimated_sell": est_cost,
            "cost_estimated_internal": real_cost,
        })
    # Sort: most recent first
    out.sort(key=lambda r: r.get("last_date") or "", reverse=True)
    return {
        "resource_id": resource_id,
        "resource_name": resource.name,
        "resource_type": resource.type if hasattr(resource, "type") else None,
        "rate_hourly_sell": rate_h,
        "rate_hourly_internal_cost": cost_h or None,
        "jobs_count": len(out),
        "jobs": out,
        "scoped_project_id": project_id,  # v3.5.0-alpha.88 — filtro applicato
    }


@router.post("/api/job/{job_id}/not-done-pool/{booking_id}/discard")
async def discard_not_done_pool_booking(
    job_id: int, booking_id: int, db: Session = Depends(get_db),
):
    """Scarta definitivamente un booking not_done dal pool: cancella il booking
    (status=cancelled). Le ore non risulteranno mai conteggiate."""
    b = db.query(Booking).filter(
        Booking.id == booking_id, Booking.job_id == job_id,
    ).first()
    if not b:
        raise HTTPException(404, "Booking non trovato")
    b.status = BookingStatus.cancelled
    db.commit()
    return {"ok": True, "id": b.id, "status": "cancelled"}


@router.post("/api/job/{job_id}/not-done-pool/discard-all")
async def discard_all_not_done_pool(job_id: int, db: Session = Depends(get_db)):
    """v3.5.0-alpha.56 — Scarta in blocco TUTTI i booking del pool not-done
    (execution_status=not_done & count_in_costs=False) del job. Cancella i
    booking (status=cancelled). Idempotente: se il pool è vuoto ritorna 0."""
    bookings = db.query(Booking).filter(
        Booking.job_id == job_id,
        Booking.execution_status == BookingExecutionStatus.not_done,
        Booking.count_in_costs == False,  # noqa: E712
        Booking.status != BookingStatus.cancelled,
    ).all()
    n = 0
    for b in bookings:
        b.status = BookingStatus.cancelled
        n += 1
    db.commit()
    return {"ok": True, "discarded": n}


# ── PDF cliente (v3.4.33) ─────────────────────────────────────────
# Esporta una versione del cost report **filtrata per il cliente**: solo
# lavorazioni quote + extra, ore previste vs consuntivate, NIENTE hardcost,
# NIENTE rate risorsa, NIENTE costi/margine. Solo cosa serve al cliente per
# verificare l'avanzamento del lavoro.

async def _client_filtered_report(job_id: int, db: Session) -> dict:
    """v3.5.0-alpha.66.11 — Vista cliente del cost report: ricalcola
    `bookings_hours` escludendo le ore dei booking attribuiti a deliverable
    non-time-based (produzione interna del file → hardcost interno, no fatt.
    al cliente). Rimuove `deliverable_hardcost_internal` e altri campi
    internal-only dal `summary`. Riusato da PDF/CSV/XLSX cliente."""
    report = await job_cost_report(job_id, db)
    bk_client = _bookings_hours_cost(job_id, db, client_view=True)
    if "summary" in report:
        report["summary"] = dict(report["summary"])
        report["summary"]["bookings_hours"] = bk_client["total_hours"]
        report["summary"]["bookings_hours_cost"] = bk_client["total_cost"]
        for k in ("deliverable_hardcost_internal", "deliverable_hours_internal",
                  "deliverable_count", "estimated_cost", "margin"):
            report["summary"].pop(k, None)
    report["bookings_breakdown"] = bk_client["breakdown_total"]
    report["bookings_by_resource"] = bk_client["by_resource"]
    return report


@router.get("/api/job/{job_id}/client-pdf")
async def cost_report_client_pdf(
    job_id: int,
    rendiconto: int = 0,
    vista: str = "now",
    db: Session = Depends(get_db),
):
    """Genera il PDF cliente del cost report. Riusa `job_cost_report` per
    raccogliere i dati e poi li passa a `generate_client_cost_report_pdf`
    filtrando i campi sensibili (hardcost/rate/margine NON vengono mai
    serializzati nel PDF — vedi pdf_export.py).

    v3.5.0-alpha.16: parametro `rendiconto` (0/1). Se 1, il PDF mostra
    Quotato/Maturato/Stimato + Over/Under per riga + totali.
    v3.5.0-alpha.55: parametro `vista` (now|forecast). Now = Over/Under
    su maturato (base fatturazione). Forecast = su stima (base report).
    """
    report = await _client_filtered_report(job_id, db)
    from app.services.branding import get_branding
    branding = get_branding(db)
    pdf_bytes = generate_client_cost_report_pdf(
        report, rendiconto=bool(rendiconto), vista=vista, branding=branding,
    )
    job_code = (report.get("job") or {}).get("code") or f"job-{job_id}"
    suffix = "_rendiconto" if rendiconto else ""
    filename = f"rendicontazione_{job_code}{suffix}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


def _client_export_rows(report: dict, rendiconto: bool, vista: str = "now") -> tuple[list[str], list[list]]:
    """Costruisce header + rows tabellari per export CSV/XLSX dal report.

    Modalità rendiconto include 6 colonne (Quotato/Maturato/Stimato/Over-Under).
    Modalità "stato" include solo quantità + stato lavorazione.

    v3.5.0-alpha.55: parametro `vista` ∈ {"now", "forecast"}.
    - "now": Over/Under = maturato − quotato (extracosto certo, base fatturazione).
    - "forecast": Over/Under = stima − quotato (sforamento previsto, base report cliente).
    Convenzione segno: positivo = OVER (sforamento), negativo = UNDER.
    """
    quoted_lines = [l for l in (report.get("cost_lines") or []) if not l.get("is_extra")]
    extra_lines = [l for l in (report.get("cost_lines") or []) if l.get("is_extra")]
    rows: list[list] = []
    is_forecast = (vista == "forecast")
    ov_field = "over_under_forecast" if is_forecast else "over_under_now"
    ou_label = f"Over/Under ({'Stima' if is_forecast else 'Maturato'} vs Quotato)"
    if rendiconto:
        header = ["Sezione", "Descrizione", "Categoria", "Unità",
                  "Q.tà preventivo", "Q.tà consuntivo",
                  "Quotato", "Maturato", "Stimato", ou_label]
        tot_q = tot_a = tot_e = 0.0
        for l in quoted_lines:
            tq = l.get("total_quoted") or 0
            ta = l.get("total_accrued") or 0
            te = l.get("total_expected") or 0
            ou = l.get(ov_field)
            if ou is None:
                ou = (te - tq) if is_forecast else (ta - tq)
            tot_q += tq; tot_a += ta; tot_e += te
            rows.append([
                "Preventivata", l.get("description"), l.get("category") or "",
                l.get("unit") or "",
                l.get("quantity_quoted") or 0, l.get("quantity_actual") or 0,
                round(tq, 2), round(ta, 2), round(te, 2), round(ou, 2),
            ])
        if quoted_lines:
            tot_ou = (tot_e - tot_q) if is_forecast else (tot_a - tot_q)
            rows.append([
                "TOTALE preventivate", "", "", "", "", "",
                round(tot_q, 2), round(tot_a, 2), round(tot_e, 2),
                round(tot_ou, 2),
            ])
        for l in extra_lines:
            tq = l.get("total_quoted") or 0
            ta = l.get("total_accrued") or 0
            te = l.get("total_expected") or 0
            ou = l.get(ov_field)
            if ou is None:
                ou = (te - tq) if is_forecast else (ta - tq)
            rows.append([
                "Extra", l.get("description"), l.get("category") or "",
                l.get("unit") or "",
                "", l.get("quantity_actual") or 0,
                round(tq, 2), round(ta, 2), round(te, 2), round(ou, 2),
            ])
    else:
        header = ["Sezione", "Descrizione", "Categoria", "Unità",
                  "Q.tà preventivo", "Q.tà consuntivo", "Stato"]
        for l in quoted_lines:
            qq = l.get("quantity_quoted") or 0
            qa = l.get("quantity_actual") or 0
            if qa == 0 and qq > 0: stato = "Da fare"
            elif qa < qq: stato = "In corso"
            elif qa == qq: stato = "Completata"
            else: stato = "Sforamento"
            rows.append(["Preventivata", l.get("description"), l.get("category") or "",
                        l.get("unit") or "", qq, qa, stato])
        for l in extra_lines:
            rows.append(["Extra", l.get("description"), l.get("category") or "",
                        l.get("unit") or "", "", l.get("quantity_actual") or 0, "Extra"])

    # Blocco ore (sempre incluso, sotto i righe)
    bd = report.get("bookings_breakdown") or {}
    summary = report.get("summary") or {}
    bk_hours = summary.get("bookings_hours", 0)
    if bk_hours and bk_hours > 0:
        rows.append([])
        rows.append(["RIEPILOGO ORE LAVORATE"] + [""] * (len(header) - 1))
        for k_label, k_field in [
            ("Ore regolari", "regular_hours"),
            ("Ore straordinarie", "overtime_hours"),
            ("Ore notturne", "night_hours"),
            ("Ore domenicali", "sunday_hours"),
            ("Ore festive", "holiday_hours"),
        ]:
            v = bd.get(k_field) or 0
            if v:
                rows.append([k_label] + [""] * (len(header) - 2) + [f"{v:.2f}h"])
        rows.append(["Totale ore"] + [""] * (len(header) - 2) + [f"{bk_hours:.2f}h"])
    return header, rows


@router.get("/api/job/{job_id}/client-csv")
async def cost_report_client_csv(
    job_id: int,
    rendiconto: int = 0,
    vista: str = "now",
    db: Session = Depends(get_db),
):
    """Export CSV (UTF-8 + BOM per Excel) del cost report cliente."""
    import csv
    import io
    report = await _client_filtered_report(job_id, db)
    header, rows = _client_export_rows(report, bool(rendiconto), vista=vista)
    buf = io.StringIO()
    buf.write("﻿")  # BOM per Excel
    w = csv.writer(buf, delimiter=";")
    job = report.get("job") or {}
    w.writerow([f"Cost Report — {job.get('code', '')} · {job.get('title', '')}"])
    w.writerow([f"Cliente: {job.get('client') or '—'}"])
    w.writerow([f"Periodo: {job.get('start_date') or '—'} → {job.get('end_date') or '—'}"])
    vista_lbl = "Over/Under su Stima vs Quotato (forecast)" if vista == "forecast" else "Over/Under su Maturato vs Quotato"
    w.writerow([f"Modalità: {('Rendiconto — ' + vista_lbl) if rendiconto else 'Stato lavorazioni'}"])
    w.writerow([])
    w.writerow(header)
    for r in rows:
        w.writerow(r)
    suffix = "_rendiconto" if rendiconto else ""
    filename = f"rendicontazione_{job.get('code', f'job-{job_id}')}{suffix}.csv"
    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/job/{job_id}/client-xlsx")
async def cost_report_client_xlsx(
    job_id: int,
    rendiconto: int = 0,
    vista: str = "now",
    db: Session = Depends(get_db),
):
    """Export XLSX (Excel) del cost report cliente. Usa openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    report = await _client_filtered_report(job_id, db)
    header, rows = _client_export_rows(report, bool(rendiconto), vista=vista)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rendicontazione"
    job = report.get("job") or {}
    bold = Font(bold=True)
    indigo = PatternFill(start_color="6272F5", end_color="6272F5", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    grey_fill = PatternFill(start_color="EEF0FB", end_color="EEF0FB", fill_type="solid")

    ws.append([f"Cost Report — {job.get('code', '')} · {job.get('title', '')}"])
    ws["A1"].font = Font(bold=True, size=14, color="6272F5")
    ws.append([f"Cliente: {job.get('client') or '—'}"])
    ws.append([f"Periodo: {job.get('start_date') or '—'} → {job.get('end_date') or '—'}"])
    vista_lbl = "Over/Under su Stima vs Quotato (forecast)" if vista == "forecast" else "Over/Under su Maturato vs Quotato"
    ws.append([f"Modalità: {('Rendiconto — ' + vista_lbl) if rendiconto else 'Stato lavorazioni'}"])
    ws.append([])
    ws.append(header)
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.font = white_bold
        cell.fill = indigo
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r if r else [""])
        # Highlight TOTALE preventivate
        if r and isinstance(r[0], str) and r[0].startswith("TOTALE"):
            for cell in ws[ws.max_row]:
                cell.font = bold
                cell.fill = grey_fill

    # Larghezze colonne
    widths = [16, 50, 18, 12, 16, 16, 14, 14, 14, 14]
    for i, w in enumerate(widths[:len(header)], start=1):
        col = ws.cell(row=header_row_idx, column=i).column_letter
        ws.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    suffix = "_rendiconto" if rendiconto else ""
    filename = f"rendicontazione_{job.get('code', f'job-{job_id}')}{suffix}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
