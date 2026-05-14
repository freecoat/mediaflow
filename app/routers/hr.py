"""
Router HR — gestione timbrature e ore di lavoro.

Sezione amministrativa per la rendicontazione delle ore di lavoro di tutte le
risorse umane (interne + freelance). Modello `TimePunch` separato dai Booking:
booking = intenzione di pianificazione, time_punch = presenza effettiva.

MVP: CRUD + lista filtrabile + totali per kind. Aggregazioni avanzate (report
mensile, costo orario × ore, esportazione cedolino) in iterazione successiva.
"""
from datetime import date, datetime, time, timedelta
from typing import Optional

from fastapi import APIRouter, Cookie, Depends, HTTPException, Request, Form
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import (
    Resource, ResourceType, TimePunch, PunchKind, Job, JobCostLine, User,
    WorkingHoursPolicy, ResourceUnavailability, UnavailabilityKind, UnavailabilityStatus,
    Booking, BookingAssignment, BookingKind, BookingStatus,
)
from app.services.auth import get_current_user_from_token
from app.services.overtime import compute_overtime, compute_punch_breakdown
from app.services.booking_cost import compute_assignment_breakdown, BookingBreakdown
from app.services.working_hours import get_holidays
from app.services.rbac import is_elevated, scope_resource_id, current_user_optional

router = APIRouter(prefix="/hr", tags=["hr"])

PERSON_TYPES = (
    ResourceType.person_internal,
    ResourceType.person_freelance,
    ResourceType.person,  # legacy
)

# Colori per tipologia di timbratura nel calendario / badge.
KIND_COLOR = {
    PunchKind.shift: None,       # usa il colore della risorsa
    PunchKind.idle: "#9ca3af",   # grigio
    PunchKind.leave: "#c084fc",  # lavanda
    PunchKind.sick: "#f87171",   # rosso chiaro
    PunchKind.break_: "#fbbf24", # giallo
    PunchKind.overtime: "#fb923c",  # arancione
}

KIND_LABEL = {
    PunchKind.shift: "Turno",
    PunchKind.idle: "Idle",
    PunchKind.leave: "Ferie/Permesso",
    PunchKind.sick: "Malattia",
    PunchKind.break_: "Pausa",
    PunchKind.overtime: "Straordinario",
}


def _tpl():
    from app.main import templates
    return templates


# v3.5.0-alpha.66.14.2: alias verso il singleton in app.services.auth.
# La logica fail-closed (settings.auth_required=True → no fallback) vive lì.
from app.services.auth import resolve_current_user as _resolve_current_user  # noqa: E402,F401
from app.context import current_tenant_id


def _punch_dict(p: TimePunch, *, fullcalendar: bool = False) -> dict:
    """Serializza una timbratura. Se fullcalendar=True, formato compatibile FullCalendar."""
    duration_h = None
    duration_h_gross = None
    break_min = int(getattr(p, "break_minutes", 0) or 0)
    if p.end_datetime:
        delta = p.end_datetime - p.start_datetime
        gross_h = delta.total_seconds() / 3600.0
        duration_h_gross = round(gross_h, 2)
        # v3.5.0-alpha.22: durata netta = lordo − pausa pranzo (solo per shift).
        if p.kind == PunchKind.shift and break_min > 0:
            net_h = max(0.0, gross_h - break_min / 60.0)
        else:
            net_h = gross_h
        duration_h = round(net_h, 2)

    if fullcalendar:
        kind_color = KIND_COLOR.get(p.kind)
        color = kind_color or (p.resource.color if p.resource else "#6272f5")
        title_parts = [p.resource.name if p.resource else "?"]
        title_parts.append(KIND_LABEL.get(p.kind, str(p.kind)))
        if p.job:
            title_parts.append(p.job.title)
        return {
            "id": f"punch-{p.id}",
            "title": " · ".join(title_parts),
            "start": p.start_datetime.isoformat(),
            "end": p.end_datetime.isoformat() if p.end_datetime else None,
            "color": color,
            "extendedProps": {
                "source": "punch",
                "punch_id": p.id,
                "resource_id": p.resource_id,
                "job_id": p.job_id,
                "job_cost_line_id": p.job_cost_line_id,
                "kind": p.kind.value if hasattr(p.kind, "value") else p.kind,
                "notes": p.notes,
                "duration_h": duration_h,
            },
        }

    return {
        "id": p.id,
        "tenant_id": p.tenant_id,
        "resource_id": p.resource_id,
        "resource_name": p.resource.name if p.resource else None,
        "resource_color": p.resource.color if p.resource else None,
        "job_id": p.job_id,
        "job_title": p.job.title if p.job else None,
        "job_code": p.job.code if p.job else None,
        "job_cost_line_id": p.job_cost_line_id,
        "cost_line_description": p.cost_line.description if p.cost_line else None,
        "start_datetime": p.start_datetime.isoformat(),
        "end_datetime": p.end_datetime.isoformat() if p.end_datetime else None,
        "duration_h": duration_h,
        "duration_h_gross": duration_h_gross,
        "break_minutes": break_min,
        "kind": p.kind.value if hasattr(p.kind, "value") else p.kind,
        "kind_label": KIND_LABEL.get(p.kind, str(p.kind)),
        "notes": p.notes,
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }


# ── Pagina HTML ──────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def hr_page(request: Request, db: Session = Depends(get_db)):
    user = current_user_optional(request)
    elevated = is_elevated(user)
    scoped_rid = None if elevated else scope_resource_id(db, user)

    persons_q = (
        db.query(Resource)
        .filter(
            Resource.tenant_id == current_tenant_id(),
            Resource.is_active == True,
            Resource.type.in_(PERSON_TYPES),
        )
    )
    if not elevated:
        if scoped_rid is None:
            persons_q = persons_q.filter(Resource.id == -1)  # nessuna risorsa visibile
        else:
            persons_q = persons_q.filter(Resource.id == scoped_rid)
    persons = persons_q.order_by(Resource.name).all()

    jobs = (
        db.query(Job)
        .filter(Job.client_id.isnot(None))
        .order_by(Job.created_at.desc())
        .limit(200)
        .all()
    )
    kinds = [{"value": k.value, "label": KIND_LABEL[k], "color": KIND_COLOR.get(k)} for k in PunchKind]
    # Categorie del filtro Tipo nella pagina (v3.5.0-alpha.16): non più i raw
    # PunchKind ma le categorie del breakdown straordinari + ferie/malattia.
    return _tpl().TemplateResponse(
        "pages/hr.html",
        {"request": request, "persons": persons, "jobs": jobs, "kinds": kinds,
         "categories": TIMELINE_CATEGORIES,
         "user_is_elevated": elevated, "scoped_resource_id": scoped_rid},
    )


# ── API ──────────────────────────────────────────────────────

def _enforce_scope(request: Request, db: Session, requested_resource_id: Optional[int]) -> Optional[int]:
    """Per staff/viewer forza il filtro sulla propria risorsa.

    Ritorna l'ID risorsa effettivo da usare nelle query. Solleva 403 se l'utente
    non-elevated tenta di accedere a una risorsa diversa dalla sua. Se è un
    elevato, lascia passare il valore richiesto (anche None = nessun filtro).
    """
    user = current_user_optional(request)
    if is_elevated(user):
        return requested_resource_id
    own = scope_resource_id(db, user)
    if own is None:
        raise HTTPException(403, "Nessuna risorsa associata a questo utente")
    if requested_resource_id and requested_resource_id != own:
        raise HTTPException(403, "Permesso negato sulla risorsa richiesta")
    return own


@router.get("/api/punches")
async def list_punches(
    request: Request,
    resource_id: Optional[int] = None,
    job_id: Optional[int] = None,
    kind: Optional[PunchKind] = None,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client_id: Optional[int] = None,
    project_id: Optional[int] = None,
    department_id: Optional[int] = None,
    format: str = "json",
    db: Session = Depends(get_db),
):
    resource_id = _enforce_scope(request, db, resource_id)
    q = (
        db.query(TimePunch)
        .options(joinedload(TimePunch.resource), joinedload(TimePunch.job))
        .filter(TimePunch.tenant_id == current_tenant_id())
    )
    if resource_id:
        q = q.filter(TimePunch.resource_id == resource_id)
    if job_id:
        q = q.filter(TimePunch.job_id == job_id)
    if kind:
        q = q.filter(TimePunch.kind == kind)
    if from_date:
        q = q.filter(
            (TimePunch.end_datetime.is_(None)) | (TimePunch.end_datetime >= from_date)
        )
    if to_date:
        q = q.filter(TimePunch.start_datetime <= to_date)
    if client_id or project_id:
        q = q.join(Job, TimePunch.job_id == Job.id)
        if client_id:
            q = q.filter(Job.client_id == client_id)
        if project_id:
            q = q.filter(Job.project_id == project_id)
    if department_id:
        q = q.join(Resource, TimePunch.resource_id == Resource.id).filter(
            Resource.department_id == department_id
        )
    # v3.5.0-alpha.14: ordine ASC (cronologico, allineato col calendario).
    # Pre-alpha.14 era desc() = più recente prima. Matteo (5 mag): preferisce
    # ordine calendario.
    q = q.order_by(TimePunch.start_datetime.asc())
    punches = q.all()

    fc = format == "fullcalendar"
    return [_punch_dict(p, fullcalendar=fc) for p in punches]


@router.post("/api/punches")
async def create_punch(
    request: Request,
    resource_id: int = Form(...),
    start_datetime: datetime = Form(...),
    end_datetime: Optional[datetime] = Form(None),
    kind: PunchKind = Form(PunchKind.shift),
    break_minutes: int = Form(0),
    job_id: Optional[int] = Form(None),
    job_cost_line_id: Optional[int] = Form(None),
    notes: Optional[str] = Form(None),
    access_token: Optional[str] = Cookie(None),
    db: Session = Depends(get_db),
):
    resource_id = _enforce_scope(request, db, resource_id)
    r = db.query(Resource).filter(
        Resource.id == resource_id,
        Resource.tenant_id == current_tenant_id(),
    ).first()
    if not r:
        raise HTTPException(404, "Risorsa non trovata")
    if r.type not in PERSON_TYPES:
        raise HTTPException(400, f"Le timbrature sono solo per persone (resource type={r.type})")

    if end_datetime and end_datetime <= start_datetime:
        raise HTTPException(400, "end_datetime deve essere successivo a start_datetime")

    # v3.5.0-alpha.14: blocco overlap con altre timbrature della stessa risorsa.
    # Una timbratura "in corso" (no end) si considera estesa fino al momento
    # del check; in pratica blocchiamo se start_datetime ricade dentro un'altra.
    overlap_q = db.query(TimePunch).filter(
        TimePunch.resource_id == resource_id,
        TimePunch.tenant_id == current_tenant_id(),
    )
    if end_datetime:
        # Nuova chiusa: cerca overlap [start, end) vs [s, e)
        # Punch in corso (end NULL) → overlap se p.start < end
        overlap = overlap_q.filter(
            (
                (TimePunch.end_datetime.is_(None)) & (TimePunch.start_datetime < end_datetime)
            ) | (
                (TimePunch.end_datetime.isnot(None))
                & (TimePunch.start_datetime < end_datetime)
                & (TimePunch.end_datetime > start_datetime)
            )
        ).first()
    else:
        # Nuova in corso: blocca se esiste una qualsiasi punch che copre `start`
        overlap = overlap_q.filter(
            (
                (TimePunch.end_datetime.is_(None))
            ) | (
                (TimePunch.end_datetime.isnot(None))
                & (TimePunch.end_datetime > start_datetime)
            )
        ).first()
    if overlap:
        raise HTTPException(
            409,
            f"Timbratura sovrapposta a #{overlap.id} ({overlap.start_datetime.strftime('%d/%m/%Y %H:%M')}"
            f"{' → ' + overlap.end_datetime.strftime('%H:%M') if overlap.end_datetime else ' (in corso)'}). "
            f"Non puoi avere due timbrature sovrapposte sulla stessa risorsa.",
        )

    # v3.5.0-alpha.22: blocco timbratura su giorno con ferie/malattia approvata.
    # Una giornata di assenza dichiarata non può convivere con una presenza
    # registrata sulla stessa data: il dato sarebbe incoerente nel rendiconto.
    if kind == PunchKind.shift:
        check_end = end_datetime or start_datetime
        unav = db.query(ResourceUnavailability).filter(
            ResourceUnavailability.resource_id == resource_id,
            ResourceUnavailability.status == UnavailabilityStatus.approved,
            ResourceUnavailability.start_date <= check_end.date(),
            ResourceUnavailability.end_date >= start_datetime.date(),
        ).first()
        if unav:
            kv = unav.kind.value if hasattr(unav.kind, "value") else unav.kind
            kv_label = {"vacation": "ferie", "sick": "malattia", "other": "permesso"}.get(kv, kv)
            raise HTTPException(
                409,
                f"Risorsa in {kv_label} dal {unav.start_date.strftime('%d/%m/%Y')} "
                f"al {unav.end_date.strftime('%d/%m/%Y')} (richiesta #{unav.id}). "
                f"Annulla l'assenza prima di registrare la timbratura.",
            )

    if job_id:
        j = db.query(Job).filter(Job.id == job_id).first()
        if not j:
            raise HTTPException(404, "Job non trovato")
    if job_cost_line_id:
        line = db.query(JobCostLine).filter(JobCostLine.id == job_cost_line_id).first()
        if not line:
            raise HTTPException(404, "Lavorazione non trovata")
        if job_id and line.job_id != job_id:
            raise HTTPException(400, f"Lavorazione #{job_cost_line_id} non appartiene al job #{job_id}")
        # Se non c'è job_id ma c'è cost_line, deduco il job dalla riga
        if not job_id:
            job_id = line.job_id

    u = _resolve_current_user(db, access_token)
    # break_minutes: clamp [0, 240] step 15. Solo per shift.
    bm = max(0, min(240, int(break_minutes or 0)))
    if kind != PunchKind.shift:
        bm = 0
    p = TimePunch(
        tenant_id=current_tenant_id(),
        resource_id=resource_id,
        job_id=job_id,
        job_cost_line_id=job_cost_line_id,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        kind=kind,
        break_minutes=bm,
        notes=notes,
        created_by_user_id=u.id if u else None,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return _punch_dict(p)


@router.put("/api/punches/{punch_id}")
async def update_punch(
    request: Request,
    punch_id: int,
    start_datetime: Optional[datetime] = Form(None),
    end_datetime: Optional[datetime] = Form(None),
    kind: Optional[PunchKind] = Form(None),
    break_minutes: Optional[int] = Form(None),
    job_id: Optional[int] = Form(None),
    job_cost_line_id: Optional[int] = Form(None),
    notes: Optional[str] = Form(None),
    clear_end: bool = Form(False),
    clear_job: bool = Form(False),
    clear_cost_line: bool = Form(False),
    db: Session = Depends(get_db),
):
    p = db.query(TimePunch).filter(
        TimePunch.id == punch_id,
        TimePunch.tenant_id == current_tenant_id(),
    ).first()
    if not p:
        raise HTTPException(404, "Timbratura non trovata")
    _enforce_scope(request, db, p.resource_id)

    if start_datetime is not None:
        p.start_datetime = start_datetime
    if end_datetime is not None:
        p.end_datetime = end_datetime
    elif clear_end:
        p.end_datetime = None
    if kind is not None:
        p.kind = kind
    if break_minutes is not None:
        bm = max(0, min(240, int(break_minutes or 0)))
        if (kind or p.kind) != PunchKind.shift:
            bm = 0
        p.break_minutes = bm
    if job_id is not None:
        p.job_id = job_id
    elif clear_job:
        p.job_id = None
        p.job_cost_line_id = None  # cancellando job, cancello anche la lavorazione
    if job_cost_line_id is not None:
        p.job_cost_line_id = job_cost_line_id
    elif clear_cost_line:
        p.job_cost_line_id = None
    if notes is not None:
        p.notes = notes

    if p.end_datetime and p.end_datetime <= p.start_datetime:
        raise HTTPException(400, "end_datetime deve essere successivo a start_datetime")

    db.commit()
    db.refresh(p)
    return _punch_dict(p)


@router.delete("/api/punches/{punch_id}")
async def delete_punch(punch_id: int, request: Request, db: Session = Depends(get_db)):
    p = db.query(TimePunch).filter(
        TimePunch.id == punch_id,
        TimePunch.tenant_id == current_tenant_id(),
    ).first()
    if not p:
        raise HTTPException(404, "Timbratura non trovata")
    _enforce_scope(request, db, p.resource_id)
    db.delete(p)
    db.commit()
    return {"ok": True}


@router.get("/api/summary")
async def punches_summary(
    request: Request,
    resource_id: Optional[int] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Totali ore per kind nel periodo. Esclude timbrature in corso (end NULL)."""
    resource_id = _enforce_scope(request, db, resource_id)
    q = db.query(TimePunch).filter(
        TimePunch.tenant_id == current_tenant_id(),
        TimePunch.end_datetime.isnot(None),
    )
    if resource_id:
        q = q.filter(TimePunch.resource_id == resource_id)
    if from_date:
        q = q.filter(TimePunch.start_datetime >= datetime.combine(from_date, datetime.min.time()))
    if to_date:
        q = q.filter(TimePunch.start_datetime <= datetime.combine(to_date, datetime.max.time()))

    totals = {k.value: 0.0 for k in PunchKind}
    grand_total = 0.0
    for p in q.all():
        hours = (p.end_datetime - p.start_datetime).total_seconds() / 3600.0
        kv = p.kind.value if hasattr(p.kind, "value") else p.kind
        totals[kv] = totals.get(kv, 0.0) + hours
        grand_total += hours

    return {
        "totals": {k: round(v, 2) for k, v in totals.items()},
        "grand_total": round(grand_total, 2),
        "labels": {k.value: KIND_LABEL[k] for k in PunchKind},
        "colors": {k.value: KIND_COLOR.get(k) for k in PunchKind},
    }


## ── Timeline unificata: TimePunch + Unavailability + breakdown (v3.5.0-alpha.16) ──
# Le categorie del filtro Tipo nella pagina HR. Sostituiscono i raw PunchKind
# nel dropdown perché solo `shift`/`break` vengono creati dal modal: l'overtime
# è una conseguenza del breakdown dell'engine, non un kind di punch. Le ferie/
# malattia sono Unavailability separati e non punch.
TIMELINE_CATEGORIES = [
    {"value": "regular",  "label": "Regolari",      "color": "#6272f5"},
    {"value": "overtime", "label": "Straordinari",  "color": "#fb923c"},
    {"value": "night",    "label": "Notturne",      "color": "#a78bfa"},
    {"value": "holiday",  "label": "Festivo",       "color": "#ef4444"},
    {"value": "sunday",   "label": "Domenicali",    "color": "#fb923c"},
    {"value": "break",    "label": "Pausa",         "color": "#fbbf24"},
    {"value": "vacation", "label": "Ferie",         "color": "#6272f5"},
    {"value": "sick",     "label": "Malattia",      "color": "#ef4444"},
    {"value": "other",    "label": "Permesso",      "color": "#9ca3af"},
]


def _entry_matches_category(entry: dict, cat: str) -> bool:
    """True se un'entry timeline rientra nella categoria filtro selezionata."""
    if not cat:
        return True
    if entry["source"] == "unavailability":
        return entry["unav_kind"] == cat
    # source == 'punch'
    if cat == "break":
        return entry["kind"] == "break"
    bd = entry.get("breakdown") or {}
    if cat == "regular":
        return (bd.get("regular_h") or 0.0) > 0
    if cat == "overtime":
        return (bd.get("overtime_h") or 0.0) > 0
    if cat == "night":
        return (bd.get("night_h") or 0.0) > 0
    if cat == "holiday":
        return bool(bd.get("is_holiday"))
    if cat == "sunday":
        return bool(bd.get("is_sunday"))
    return False


@router.get("/api/timeline")
async def hr_timeline(
    request: Request,
    resource_id: Optional[int] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Lista unificata TimePunch + ResourceUnavailability con breakdown per-punch.

    Sostituisce le 2 chiamate separate /api/punches + /api/summary nella pagina
    HR. Ritorna entries cronologiche miste e totali per categoria. Il filtro
    `category` agisce su entrambe le sorgenti (es. "vacation" filtra solo le
    unavailability di tipo ferie; "overtime" filtra solo i punch con quota di
    straordinario > 0).
    """
    resource_id = _enforce_scope(request, db, resource_id)

    # Punch nel periodo (chiusi e in corso, ma il breakdown salta gli in-corso)
    pq = (
        db.query(TimePunch)
        .options(joinedload(TimePunch.resource), joinedload(TimePunch.job))
        .filter(TimePunch.tenant_id == current_tenant_id())
    )
    if resource_id:
        pq = pq.filter(TimePunch.resource_id == resource_id)
    if from_date:
        pq = pq.filter(
            (TimePunch.end_datetime.is_(None))
            | (TimePunch.end_datetime >= datetime.combine(from_date, datetime.min.time()))
        )
    if to_date:
        pq = pq.filter(TimePunch.start_datetime <= datetime.combine(to_date, datetime.max.time()))
    punches = pq.order_by(TimePunch.start_datetime.asc()).all()

    # Calcolo breakdown raggruppato per (resource_id, policy)
    breakdown_by_punch: dict[int, dict] = {}
    if punches:
        from collections import defaultdict
        by_res: dict[int, list[TimePunch]] = defaultdict(list)
        for p in punches:
            by_res[p.resource_id].append(p)
        for rid, group in by_res.items():
            res = db.query(Resource).filter(Resource.id == rid).first()
            policy = _resolve_policy_for_resource(db, res) if res else None
            if not policy:
                continue
            bd_map = compute_punch_breakdown(group, policy)
            for pid, pb in bd_map.items():
                breakdown_by_punch[pid] = pb.as_dict()

    # Unavailability approvate nel periodo (espanse 1 entry per giorno)
    uav_entries: list[dict] = []
    if from_date and to_date:
        uq = (
            db.query(ResourceUnavailability)
            .options(joinedload(ResourceUnavailability.resource))
            .filter(
                ResourceUnavailability.status == UnavailabilityStatus.approved,
                ResourceUnavailability.end_date >= from_date,
                ResourceUnavailability.start_date <= to_date,
            )
        )
        if resource_id:
            uq = uq.filter(ResourceUnavailability.resource_id == resource_id)
        # daily hours per assenza dipende dalla policy della risorsa
        for u in uq.all():
            policy = _resolve_policy_for_resource(db, u.resource) if u.resource else None
            daily_h = (policy.daily_hours_threshold if policy else 8.0) or 8.0
            kv = u.kind.value if hasattr(u.kind, "value") else u.kind
            s = max(u.start_date, from_date)
            e = min(u.end_date, to_date)
            cur = s
            while cur <= e:
                # Salta i weekend per ferie? No — ricalca quanto fa /api/calendar
                # (anche lì conta tutti i giorni). L'utente vede esattamente le
                # date in cui era assente. Eventuale skip weekend va deciso in UI.
                uav_entries.append({
                    "source": "unavailability",
                    "id": f"u{u.id}-{cur.isoformat()}",
                    "unav_id": u.id,
                    "resource_id": u.resource_id,
                    "resource_name": u.resource.name if u.resource else None,
                    "resource_color": u.resource.color if u.resource else None,
                    "date": cur.isoformat(),
                    "start_datetime": datetime.combine(cur, time(9, 0)).isoformat(),
                    "end_datetime": datetime.combine(cur, time(9, 0)).isoformat(),
                    "duration_h": daily_h,
                    "unav_kind": kv,
                    "unav_kind_label": {
                        "vacation": "Ferie",
                        "sick": "Malattia",
                        "other": "Permesso",
                    }.get(kv, kv),
                    "reason": u.reason,
                })
                cur += timedelta(days=1)

    # Compongo entries punch
    punch_entries: list[dict] = []
    for p in punches:
        kv = p.kind.value if hasattr(p.kind, "value") else p.kind
        bd = breakdown_by_punch.get(p.id)
        duration_h = None
        duration_h_gross = None
        bm = int(getattr(p, "break_minutes", 0) or 0)
        if p.end_datetime:
            gross = (p.end_datetime - p.start_datetime).total_seconds() / 3600.0
            duration_h_gross = round(gross, 2)
            net = gross - (bm / 60.0) if (p.kind == PunchKind.shift and bm > 0) else gross
            duration_h = round(max(0.0, net), 2)
        punch_entries.append({
            "source": "punch",
            "id": f"p{p.id}",
            "punch_id": p.id,
            "resource_id": p.resource_id,
            "resource_name": p.resource.name if p.resource else None,
            "resource_color": p.resource.color if p.resource else None,
            "job_id": p.job_id,
            "job_title": p.job.title if p.job else None,
            "job_code": p.job.code if p.job else None,
            "start_datetime": p.start_datetime.isoformat(),
            "end_datetime": p.end_datetime.isoformat() if p.end_datetime else None,
            "duration_h": duration_h,
            "duration_h_gross": duration_h_gross,
            "break_minutes": bm,
            "kind": kv,
            "kind_label": KIND_LABEL.get(p.kind, str(p.kind)),
            "notes": p.notes,
            "in_progress": p.end_datetime is None,
            "breakdown": bd,
        })

    # Filtro categoria
    cat = (category or "").strip() or None
    all_entries = punch_entries + uav_entries
    if cat:
        all_entries = [e for e in all_entries if _entry_matches_category(e, cat)]

    # Sort cronologico
    all_entries.sort(key=lambda e: e.get("start_datetime") or e.get("date") or "")

    # Totali per categoria, ricalcolati sul subset filtrato
    totals = {c["value"]: 0.0 for c in TIMELINE_CATEGORIES}
    has_policy_warning = False
    for e in all_entries:
        if e["source"] == "unavailability":
            totals[e["unav_kind"]] = totals.get(e["unav_kind"], 0.0) + (e["duration_h"] or 0.0)
        else:
            if e["kind"] == "break":
                totals["break"] += (e["duration_h"] or 0.0)
                continue
            bd = e.get("breakdown")
            if bd is None:
                # Fallback: nessuna WorkingHoursPolicy default → tutte le ore
                # vanno in "regular" (niente split overtime/notte/festivo).
                totals["regular"] += (e["duration_h"] or 0.0)
                if e["kind"] == "shift" and e.get("end_datetime"):
                    has_policy_warning = True
                continue
            totals["regular"] += bd.get("regular_h") or 0.0
            totals["overtime"] += bd.get("overtime_h") or 0.0
            totals["night"]    += bd.get("night_h") or 0.0
            if bd.get("is_holiday"):
                totals["holiday"] += bd.get("duration_h") or 0.0
            elif bd.get("is_sunday"):
                totals["sunday"] += bd.get("duration_h") or 0.0
    grand_total = sum(totals.values())

    return {
        "entries": all_entries,
        "totals": {k: round(v, 2) for k, v in totals.items()},
        "grand_total": round(grand_total, 2),
        "categories": TIMELINE_CATEGORIES,
        "filter_category": cat,
        "warning": (
            "Nessuna WorkingHoursPolicy default configurata: il breakdown "
            "regular/overtime/notte/festivo non è disponibile. Vai in /settings#hours "
            "per impostare una policy default."
        ) if has_policy_warning else None,
    }


@router.get("/api/calendar")
async def calendar_summary(
    request: Request,
    from_date: date,
    to_date: date,
    resource_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Riepilogo giorno-per-giorno per la vista calendario complessiva (v3.4.30).

    Ritorna per ogni giorno del periodo:
    - regular_h, overtime_h, night_h: split delle ore lavorate (timbrature shift/overtime)
    - vacation_h, sick_h, other_h: ore da ResourceUnavailability approvate (calcolate
      come daily_hours_threshold della policy applicabile)
    - has_unavailability, unav_kinds: marker per evidenziazione celle
    - resource_count: quanti utenti hanno timbrato/erano in ferie/malattia in quel giorno

    Se `resource_id` è specificato, il calcolo split overtime usa la policy della
    risorsa via `compute_overtime`. Senza resource_id, somma cross-risorsa (ore
    totali shift+overtime, niente split notturno fine).
    """
    resource_id = _enforce_scope(request, db, resource_id)
    if to_date < from_date:
        raise HTTPException(400, "to_date < from_date")

    # Singola risorsa: usa compute_overtime per il breakdown preciso
    days_map: dict[str, dict] = {}
    cur = from_date
    while cur <= to_date:
        days_map[cur.isoformat()] = {
            "date": cur.isoformat(),
            "regular_h": 0.0, "overtime_h": 0.0, "night_h": 0.0,
            "vacation_h": 0.0, "sick_h": 0.0, "other_h": 0.0,
            "resource_ids": set(),
            "unav_kinds": set(),
        }
        cur += timedelta(days=1)

    pq = db.query(TimePunch).options(joinedload(TimePunch.resource)).filter(
        TimePunch.tenant_id == current_tenant_id(),
        TimePunch.end_datetime.isnot(None),
        TimePunch.start_datetime >= datetime.combine(from_date, datetime.min.time()),
        TimePunch.start_datetime <= datetime.combine(to_date, datetime.max.time()),
    )
    if resource_id:
        pq = pq.filter(TimePunch.resource_id == resource_id)
    punches = pq.all()

    if resource_id:
        # Single resource → breakdown via compute_overtime per giorno
        res = db.query(Resource).filter(Resource.id == resource_id).first()
        policy = _resolve_policy_for_resource(db, res)
        if policy:
            # Raggruppa punch per giorno
            from collections import defaultdict
            by_day = defaultdict(list)
            for p in punches:
                by_day[p.start_datetime.date().isoformat()].append(p)
            for day_iso, day_punches in by_day.items():
                if day_iso not in days_map:
                    continue
                bd = compute_overtime(day_punches, policy)
                days_map[day_iso]["regular_h"] = round(bd.regular_hours, 2)
                # somma overtime giornaliero+settimanale per la cella
                ot = (getattr(bd, "overtime_daily_hours", 0.0) or 0.0) + \
                     (getattr(bd, "overtime_weekly_hours", 0.0) or 0.0)
                days_map[day_iso]["overtime_h"] = round(ot, 2)
                days_map[day_iso]["night_h"] = round(getattr(bd, "night_hours", 0.0) or 0.0, 2)
                days_map[day_iso]["resource_ids"].add(resource_id)
        else:
            # No policy: somma tutto come regular
            for p in punches:
                d = p.start_datetime.date().isoformat()
                if d not in days_map:
                    continue
                hrs = (p.end_datetime - p.start_datetime).total_seconds() / 3600.0
                days_map[d]["regular_h"] += hrs
                days_map[d]["resource_ids"].add(p.resource_id)
    else:
        # All resources → aggregato semplice (no split overtime preciso cross-policy)
        for p in punches:
            d = p.start_datetime.date().isoformat()
            if d not in days_map:
                continue
            hrs = (p.end_datetime - p.start_datetime).total_seconds() / 3600.0
            kv = p.kind.value if hasattr(p.kind, "value") else p.kind
            if kv == "overtime":
                days_map[d]["overtime_h"] += hrs
            else:
                days_map[d]["regular_h"] += hrs
            days_map[d]["resource_ids"].add(p.resource_id)

    # Ferie/malattia/permessi: ResourceUnavailability approvate
    uq = db.query(ResourceUnavailability).options(joinedload(ResourceUnavailability.resource)).filter(
        ResourceUnavailability.status == UnavailabilityStatus.approved,
        ResourceUnavailability.end_date >= from_date,
        ResourceUnavailability.start_date <= to_date,
    )
    if resource_id:
        uq = uq.filter(ResourceUnavailability.resource_id == resource_id)
    # Default daily hours per assenze: usa policy della risorsa o fallback 8
    default_policy = db.query(WorkingHoursPolicy).filter(
        WorkingHoursPolicy.tenant_id == current_tenant_id(),
        WorkingHoursPolicy.is_default == True,  # noqa: E712
    ).first()
    fallback_daily = (default_policy.daily_hours_threshold if default_policy else 8.0) or 8.0

    for u in uq.all():
        # daily hours specifici alla risorsa
        rp = _resolve_policy_for_resource(db, u.resource) if u.resource else default_policy
        daily_h = (rp.daily_hours_threshold if rp else fallback_daily) or fallback_daily
        kind_v = u.kind.value if hasattr(u.kind, "value") else u.kind
        s = max(u.start_date, from_date)
        e = min(u.end_date, to_date)
        cur = s
        while cur <= e:
            iso = cur.isoformat()
            if iso in days_map:
                if kind_v == "vacation":
                    days_map[iso]["vacation_h"] += daily_h
                elif kind_v == "sick":
                    days_map[iso]["sick_h"] += daily_h
                else:
                    days_map[iso]["other_h"] += daily_h
                days_map[iso]["resource_ids"].add(u.resource_id)
                days_map[iso]["unav_kinds"].add(kind_v)
            cur += timedelta(days=1)

    # Round + serializzazione + totali periodo
    days_out = []
    period_totals = {
        "regular_h": 0.0, "overtime_h": 0.0, "night_h": 0.0,
        "vacation_h": 0.0, "sick_h": 0.0, "other_h": 0.0,
        "total_h": 0.0,
    }
    for iso in sorted(days_map.keys()):
        d = days_map[iso]
        for k in ("regular_h", "overtime_h", "night_h", "vacation_h", "sick_h", "other_h"):
            d[k] = round(d[k], 2)
            period_totals[k] += d[k]
        d["total_h"] = round(
            d["regular_h"] + d["overtime_h"] + d["vacation_h"] + d["sick_h"] + d["other_h"], 2
        )
        d["resource_count"] = len(d.pop("resource_ids"))
        d["unav_kinds"] = sorted(d["unav_kinds"])
        days_out.append(d)
    period_totals["total_h"] = round(
        period_totals["regular_h"] + period_totals["overtime_h"]
        + period_totals["vacation_h"] + period_totals["sick_h"] + period_totals["other_h"],
        2,
    )
    for k in period_totals:
        period_totals[k] = round(period_totals[k], 2)

    return {
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "resource_id": resource_id,
        "days": days_out,
        "totals": period_totals,
    }


def _resolve_policy_for_resource(db: Session, resource: Optional[Resource]) -> Optional[WorkingHoursPolicy]:
    """Override per-risorsa, fallback alla policy default del tenant."""
    if resource and resource.working_hours_policy_id:
        p = db.query(WorkingHoursPolicy).filter(
            WorkingHoursPolicy.id == resource.working_hours_policy_id,
            WorkingHoursPolicy.tenant_id == current_tenant_id(),
        ).first()
        if p:
            return p
    return db.query(WorkingHoursPolicy).filter(
        WorkingHoursPolicy.tenant_id == current_tenant_id(),
        WorkingHoursPolicy.is_default == True,  # noqa: E712
    ).first()


@router.get("/api/overtime")
async def overtime_breakdown(
    request: Request,
    resource_id: int,
    from_date: date,
    to_date: date,
    db: Session = Depends(get_db),
):
    resource_id = _enforce_scope(request, db, resource_id)
    """Breakdown auto-straordinari per una risorsa nel periodo.

    Calcola ore regolari, overtime giornaliero/settimanale, fascia notturna,
    domenica e festivi a partire dai TimePunch chiusi (kind shift+overtime),
    applicando la policy della risorsa (override) o quella default del tenant.
    """
    res = db.query(Resource).filter(
        Resource.id == resource_id,
        Resource.tenant_id == current_tenant_id(),
    ).first()
    if not res:
        raise HTTPException(404, "Risorsa non trovata")

    policy = _resolve_policy_for_resource(db, res)
    if not policy:
        # v3.5.0-alpha.9: degradazione graceful invece di 400.
        # Senza policy non possiamo splittare regular/overtime/notturno, ma
        # possiamo comunque tornare le ore totali della risorsa nel periodo
        # + un warning visibile in UI. Evita di rompere /hr quando il tenant
        # non ha (ancora) configurato nessuna WorkingHoursPolicy default.
        punches = db.query(TimePunch).filter(
            TimePunch.tenant_id == current_tenant_id(),
            TimePunch.resource_id == resource_id,
            TimePunch.end_datetime.isnot(None),
            TimePunch.start_datetime >= datetime.combine(from_date, datetime.min.time()),
            TimePunch.start_datetime <= datetime.combine(to_date, datetime.max.time()),
        ).all()
        total_h = sum(
            (p.end_datetime - p.start_datetime).total_seconds() / 3600.0 for p in punches
        )
        return {
            "resource_id": resource_id,
            "resource_name": res.name,
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "policy": None,
            "warning": (
                "Nessuna WorkingHoursPolicy default configurata. Il breakdown "
                "regular/overtime/notturno non è disponibile. Vai in /settings#hours "
                "per impostare una policy default."
            ),
            "breakdown": {
                "regular_hours": round(total_h, 2),
                "overtime_daily_hours": 0.0,
                "overtime_weekly_hours": 0.0,
                "night_hours": 0.0,
                "sunday_hours": 0.0,
                "holiday_hours": 0.0,
                "total_hours": round(total_h, 2),
                "weighted_factor": round(total_h, 2),
            },
            "unavailability": {
                "vacation_days": 0, "sick_days": 0, "other_days": 0,
                "vacation_hours": 0.0, "sick_hours": 0.0, "other_hours": 0.0,
            },
            "grand_total_hours": round(total_h, 2),
        }

    punches = db.query(TimePunch).filter(
        TimePunch.tenant_id == current_tenant_id(),
        TimePunch.resource_id == resource_id,
        TimePunch.end_datetime.isnot(None),
        TimePunch.start_datetime >= datetime.combine(from_date, datetime.min.time()),
        TimePunch.start_datetime <= datetime.combine(to_date, datetime.max.time()),
    ).all()

    breakdown = compute_overtime(punches, policy)

    # Conteggio ore ferie/malattia/altro nel periodo (per la rendicontazione amministrativa).
    # Convertiamo i giorni di unavailability approvata in ore usando daily_hours_threshold della policy.
    daily_h = policy.daily_hours_threshold or 8.0
    unav = db.query(ResourceUnavailability).filter(
        ResourceUnavailability.resource_id == resource_id,
        ResourceUnavailability.status == UnavailabilityStatus.approved,
        ResourceUnavailability.end_date >= from_date,
        ResourceUnavailability.start_date <= to_date,
    ).all()
    vacation_days = sick_days = other_days = 0
    for u in unav:
        s = max(u.start_date, from_date)
        e = min(u.end_date, to_date)
        if e < s:
            continue
        ndays = (e - s).days + 1
        k = u.kind.value if hasattr(u.kind, "value") else u.kind
        if k == "vacation":
            vacation_days += ndays
        elif k == "sick":
            sick_days += ndays
        else:
            other_days += ndays
    # v3.5.0-alpha.111.19 — Ore di permesso (other) con moltiplicatore ROL
    # opzionale per report HR consulente lavoro. NON tocca costi.
    permit_mult = float(getattr(policy, "permit_multiplier", 1.0) or 1.0)
    other_hours_raw = round(other_days * daily_h, 2)
    permit_weighted = round(other_hours_raw * permit_mult, 2)

    unavailability = {
        "vacation_days": vacation_days,
        "sick_days": sick_days,
        "other_days": other_days,
        "vacation_hours": round(vacation_days * daily_h, 2),
        "sick_hours": round(sick_days * daily_h, 2),
        "other_hours": other_hours_raw,
        # Ore permesso ponderate (ROL). Coincide con other_hours quando
        # permit_multiplier=1.0. Sempre presenti: la UI decide se mostrare
        # entrambi i valori solo se mult != 1.0.
        "permit_hours_weighted": permit_weighted,
        "permit_multiplier": permit_mult,
    }
    grand_total = round(
        breakdown.total_hours
        + unavailability["vacation_hours"]
        + unavailability["sick_hours"]
        + unavailability["other_hours"],
        2,
    )

    return {
        "resource_id": resource_id,
        "resource_name": res.name,
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "policy": {
            "id": policy.id,
            "name": policy.name,
            "daily_hours_threshold": policy.daily_hours_threshold,
            "weekly_hours_threshold": policy.weekly_hours_threshold,
            "overtime_multiplier": policy.overtime_multiplier,
            "night_multiplier": policy.night_multiplier,
            "sunday_multiplier": policy.sunday_multiplier,
            "holiday_multiplier": policy.holiday_multiplier,
            "permit_multiplier": permit_mult,
            "night_start": policy.night_start.strftime("%H:%M") if policy.night_start else None,
            "night_end": policy.night_end.strftime("%H:%M") if policy.night_end else None,
        },
        "breakdown": breakdown.as_dict(),
        "unavailability": unavailability,
        "grand_total_hours": grand_total,
    }


# ── BOOKING INTERNI: monte ore non-progetto (v3.5.0-alpha.65) ───────
# Manutenzione, R&D, training. Non hanno cost-line cliente, quindi sono
# fuori dal cost-report ma rendicontano comunque ore-uomo. Aggrega per
# risorsa con ore lineari e weighted (multiplier holiday/sunday/overtime/night
# della WorkingHoursPolicy della risorsa).

INTERNAL_KIND_LABEL = {
    BookingKind.internal_maintenance: "Manutenzione",
    BookingKind.internal_research:    "R&D / Test",
    BookingKind.internal_training:    "Formazione",
}


@router.get("/api/internal-bookings-report")
async def internal_bookings_report(
    request: Request,
    from_date: date,
    to_date: date,
    db: Session = Depends(get_db),
):
    """Aggrega i booking con `kind != project` nel periodo, raggruppa per
    risorsa e per kind interno, e calcola il monte ore (lineare + weighted).

    Restituisce:
      - by_resource: list of {resource_id, resource_name, total_h, weighted_h,
        by_kind: {kind: {hours_linear, hours_weighted}}}
      - by_kind: aggregato {kind_label: {hours_linear, hours_weighted, count}}
      - totals: {hours_linear, hours_weighted}
    """
    if to_date < from_date:
        raise HTTPException(400, "to_date precedente a from_date")

    bookings = (
        db.query(Booking)
        .options(joinedload(Booking.assignments).joinedload(BookingAssignment.resource))
        .filter(
            Booking.kind != BookingKind.project,
            Booking.status != BookingStatus.cancelled,
            Booking.start_datetime >= datetime.combine(from_date, time(0, 0)),
            Booking.start_datetime < datetime.combine(to_date + timedelta(days=1), time(0, 0)),
        )
        .all()
    )

    by_resource: dict[int, dict] = {}
    by_kind: dict[str, dict] = {}
    holidays_cache: dict = {}

    def _hols(policy, y0, y1):
        key = (id(policy), y0, y1)
        if key not in holidays_cache:
            holidays_cache[key] = get_holidays(policy, y0, y1)
        return holidays_cache[key]

    total_linear = 0.0
    total_weighted = 0.0

    for b in bookings:
        kind_key = b.kind.value if hasattr(b.kind, "value") else str(b.kind)
        kind_label = INTERNAL_KIND_LABEL.get(b.kind, kind_key)
        by_kind.setdefault(kind_label, {
            "kind": kind_key,
            "label": kind_label,
            "hours_linear": 0.0,
            "hours_weighted": 0.0,
            "count": 0,
        })
        by_kind[kind_label]["count"] += 1
        for a in b.assignments:
            if not a.resource or not a.start_datetime or not a.end_datetime:
                continue
            linear_h = max(0.0, (a.end_datetime - a.start_datetime).total_seconds() / 3600.0)
            policy = _resolve_policy_for_resource(db, a.resource)
            weighted_h = linear_h
            if policy:
                hols = _hols(policy, a.start_datetime.year, a.end_datetime.year)
                br = compute_assignment_breakdown(a, policy, hols, b)
                # weighted_factor copre regular + overtime/notte/dom/festivo
                # con i moltiplicatori della policy. Per booking interni
                # overtime_status è quasi sempre `none` (non c'è cliente da
                # avvisare), quindi l'OT eventualmente speso è già pesato.
                # Fallback al lineare se il breakdown è vuoto (es. assignment
                # non valutabile). Pending OT su booking interni è raro: se
                # presente lo aggiungiamo al weighted come ore "approved-eq"
                # perché non c'è un workflow d'approvazione per il manutenzione.
                weighted_h = br.weighted_factor + br.pending_overtime_hours
                if weighted_h <= 0:
                    weighted_h = linear_h

            rmap = by_resource.setdefault(a.resource.id, {
                "resource_id": a.resource.id,
                "resource_name": a.resource.name,
                "resource_type": (a.resource.type.value if hasattr(a.resource.type, "value")
                                  else str(a.resource.type)),
                "hours_linear": 0.0,
                "hours_weighted": 0.0,
                "by_kind": {},
            })
            rmap["hours_linear"] += linear_h
            rmap["hours_weighted"] += weighted_h
            kmap = rmap["by_kind"].setdefault(kind_label, {
                "kind": kind_key, "label": kind_label,
                "hours_linear": 0.0, "hours_weighted": 0.0,
            })
            kmap["hours_linear"] += linear_h
            kmap["hours_weighted"] += weighted_h

            by_kind[kind_label]["hours_linear"] += linear_h
            by_kind[kind_label]["hours_weighted"] += weighted_h
            total_linear += linear_h
            total_weighted += weighted_h

    # Round + sort
    by_resource_out = []
    for r in sorted(by_resource.values(), key=lambda x: -x["hours_weighted"]):
        r["hours_linear"] = round(r["hours_linear"], 2)
        r["hours_weighted"] = round(r["hours_weighted"], 2)
        r["by_kind"] = [
            {**v, "hours_linear": round(v["hours_linear"], 2),
             "hours_weighted": round(v["hours_weighted"], 2)}
            for v in r["by_kind"].values()
        ]
        by_resource_out.append(r)
    by_kind_out = [
        {**v, "hours_linear": round(v["hours_linear"], 2),
         "hours_weighted": round(v["hours_weighted"], 2)}
        for v in sorted(by_kind.values(), key=lambda x: -x["hours_weighted"])
    ]

    return {
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "totals": {
            "hours_linear": round(total_linear, 2),
            "hours_weighted": round(total_weighted, 2),
            "bookings_count": len(bookings),
        },
        "by_resource": by_resource_out,
        "by_kind": by_kind_out,
    }


# ── Export ore per consulente lavoro (v3.5.0-alpha.111.21) ──────────
# CSV + XLSX per singolo lavoratore / reparto / azienda.
# Permission gate: scope='resource' senza filtri = solo manager+.
# Output 1 riga per punch + 1 riga per giorno di unavailability nel range.

_PUNCH_KIND_LABEL = {
    "shift": "Turno",
    "overtime": "Straordinario",
    "idle": "Idle (presente non allocato)",
    "leave": "Permesso",
    "sick": "Malattia",
    "break_": "Pausa",
}
_UNAV_KIND_LABEL = {
    "vacation": "Ferie",
    "sick": "Malattia",
    "holiday": "Festività",
    "other": "Permesso",
    "weekend": "Weekend",
}


def _export_rows_for_scope(
    db: Session,
    *,
    from_date: date,
    to_date: date,
    resource_id: Optional[int] = None,
    department_id: Optional[int] = None,
) -> tuple[list[str], list[list]]:
    """Costruisce header + rows per export ore.
    Una riga per ogni TimePunch nel range + una riga per ogni giorno di
    ResourceUnavailability (ferie/malattia/permesso) nel range.
    """
    q_res = db.query(Resource).filter(
        Resource.tenant_id == 1,
        Resource.type == ResourceType.person_internal,
    )
    if resource_id:
        q_res = q_res.filter(Resource.id == resource_id)
    if department_id:
        q_res = q_res.filter(Resource.department_id == department_id)
    resources = q_res.all()
    if not resources:
        return [], []
    res_ids = [r.id for r in resources]
    res_by_id = {r.id: r for r in resources}

    # Pre-fetch dept names + policy permit_multiplier
    from app.models.models import Department as _Dept
    dept_by_id = {d.id: d.name for d in db.query(_Dept).all()}
    policy = db.query(WorkingHoursPolicy).filter(
        WorkingHoursPolicy.tenant_id == 1,
        WorkingHoursPolicy.is_default == True,  # noqa: E712
    ).first()
    permit_mult = float(getattr(policy, "permit_multiplier", 1.0) or 1.0) if policy else 1.0
    daily_h = 8.0
    if policy and policy.daily_hours_threshold:
        daily_h = float(policy.daily_hours_threshold)

    # Punches in range
    punches = (
        db.query(TimePunch)
        .filter(
            TimePunch.tenant_id == 1,
            TimePunch.resource_id.in_(res_ids),
            TimePunch.start_datetime >= datetime.combine(from_date, time.min),
            TimePunch.start_datetime <= datetime.combine(to_date, time.max),
        )
        .order_by(TimePunch.start_datetime)
        .all()
    )
    # Job + cost line lookup
    job_ids = {p.job_id for p in punches if p.job_id}
    jcl_ids = {p.job_cost_line_id for p in punches if p.job_cost_line_id}
    job_by_id = {j.id: j for j in db.query(Job).filter(Job.id.in_(job_ids)).all()} if job_ids else {}
    jcl_by_id = {j.id: j for j in db.query(JobCostLine).filter(JobCostLine.id.in_(jcl_ids)).all()} if jcl_ids else {}

    # Unavailability in range
    unavs = (
        db.query(ResourceUnavailability)
        .filter(
            ResourceUnavailability.resource_id.in_(res_ids),
            ResourceUnavailability.start_date <= to_date,
            ResourceUnavailability.end_date >= from_date,
        )
        .all()
    )

    header = [
        "Data", "Risorsa", "Reparto", "Tipo",
        "Inizio", "Fine", "Ore", "Ore ponderate",
        "Pausa (min)", "Job", "Lavorazione", "Note",
    ]
    rows: list[list] = []

    for p in punches:
        r = res_by_id.get(p.resource_id)
        if not r:
            continue
        kind_str = p.kind.value if hasattr(p.kind, "value") else str(p.kind)
        kind_lbl = _PUNCH_KIND_LABEL.get(kind_str, kind_str)
        start_iso = p.start_datetime.strftime("%Y-%m-%d")
        start_t = p.start_datetime.strftime("%H:%M")
        end_t = p.end_datetime.strftime("%H:%M") if p.end_datetime else ""
        hours = 0.0
        if p.end_datetime:
            ms = (p.end_datetime - p.start_datetime).total_seconds() / 3600.0
            hours = round(max(0.0, ms - (p.break_minutes or 0) / 60.0), 2)
        job = job_by_id.get(p.job_id) if p.job_id else None
        jcl = jcl_by_id.get(p.job_cost_line_id) if p.job_cost_line_id else None
        rows.append([
            start_iso,
            r.name,
            dept_by_id.get(r.department_id, ""),
            kind_lbl,
            start_t,
            end_t,
            hours,
            hours,  # ore ponderate = identiche per turno/overtime (multiplier applicato in altre viste)
            p.break_minutes or 0,
            job.code if job else "",
            (jcl.description or "")[:80] if jcl else "",
            p.notes or "",
        ])

    for u in unavs:
        r = res_by_id.get(u.resource_id)
        if not r:
            continue
        kind_str = u.kind.value if hasattr(u.kind, "value") else str(u.kind)
        kind_lbl = _UNAV_KIND_LABEL.get(kind_str, kind_str)
        is_permit = (kind_str == "other")
        # 1 riga per giorno nel range
        d = max(u.start_date, from_date)
        e = min(u.end_date, to_date)
        while d <= e:
            ore = daily_h
            ore_pond = round(ore * permit_mult, 2) if is_permit else ore
            rows.append([
                d.isoformat(),
                r.name,
                dept_by_id.get(r.department_id, ""),
                kind_lbl,
                "", "",
                ore, ore_pond,
                0, "", "",
                u.reason or "",
            ])
            d += timedelta(days=1)

    # Sort by Data + Risorsa
    rows.sort(key=lambda x: (x[0], x[1]))
    return header, rows


def _build_csv(header: list[str], rows: list[list]) -> bytes:
    import csv
    from io import StringIO
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(header)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM Excel-friendly


def _build_xlsx(header: list[str], rows: list[list], title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14, color="6272F5")
    ws.append([f"Generato: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.append([])
    ws.append(header)
    header_row = ws.max_row
    indigo = PatternFill(start_color="6272F5", end_color="6272F5", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.font = white_bold
        cell.fill = indigo
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    widths = [11, 26, 22, 22, 7, 7, 7, 11, 9, 14, 40, 30]
    for i, w in enumerate(widths[:len(header)], start=1):
        col = ws.cell(row=header_row, column=i).column_letter
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/api/export/timesheet")
async def export_timesheet(
    request: Request,
    scope: str = "resource",  # 'resource' | 'department' | 'all'
    resource_id: Optional[int] = None,
    department_id: Optional[int] = None,
    from_date: str = "",
    to_date: str = "",
    format: str = "csv",  # 'csv' | 'xlsx'
    db: Session = Depends(get_db),
):
    """Export ore per consulente lavoro. Scope: singolo / reparto / azienda.

    v3.5.0-alpha.111.21 — CSV (UTF-8 BOM) o XLSX. Una riga per TimePunch +
    una riga per giorno ResourceUnavailability nel range.

    Permission gate:
    - scope='resource' + resource_id == proprio user.resource_id → sempre OK
    - scope='resource' altrui / 'department' / 'all' → solo manager+ (`is_elevated`).
    """
    from fastapi.responses import Response
    # Parse date
    try:
        fd = date.fromisoformat(from_date) if from_date else date.today().replace(day=1)
        td = date.fromisoformat(to_date) if to_date else date.today()
    except ValueError:
        raise HTTPException(400, "Date non valide (formato YYYY-MM-DD)")
    if td < fd:
        raise HTTPException(400, "to_date precedente a from_date")

    user = current_user_optional(request)
    # Determina scope effettivo
    if scope == "resource" and resource_id:
        # Self vs altri
        my_res_id = scope_resource_id(user)
        if my_res_id != resource_id and not is_elevated(user):
            raise HTTPException(403, "Servono permessi manager+ per esportare ore altrui")
    elif scope in ("department", "all"):
        if not is_elevated(user):
            raise HTTPException(403, "Servono permessi manager+ per esportare reparto/azienda")
    else:
        raise HTTPException(400, "scope non valido o resource_id mancante")

    # Filtra args
    rid = resource_id if scope == "resource" else None
    did = department_id if scope == "department" else None

    header, rows = _export_rows_for_scope(
        db, from_date=fd, to_date=td, resource_id=rid, department_id=did,
    )
    if not header:
        raise HTTPException(404, "Nessun dato per il range/scope selezionato")

    # Build title + filename
    if scope == "resource":
        r = db.query(Resource).filter(Resource.id == resource_id).first()
        title = f"Ore lavoro · {r.name if r else f'#{resource_id}'} · {fd}→{td}"
        fname_stem = f"timesheet_{r.name.replace(' ', '_') if r else resource_id}_{fd}_{td}"
    elif scope == "department":
        from app.models.models import Department
        d = db.query(Department).filter(Department.id == department_id).first()
        title = f"Ore lavoro · Reparto {d.name if d else f'#{department_id}'} · {fd}→{td}"
        fname_stem = f"timesheet_dept_{d.name.replace(' ', '_') if d else department_id}_{fd}_{td}"
    else:
        title = f"Ore lavoro · Azienda · {fd}→{td}"
        fname_stem = f"timesheet_azienda_{fd}_{td}"

    if format == "xlsx":
        body = _build_xlsx(header, rows, title)
        return Response(
            content=body,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname_stem}.xlsx"'},
        )
    else:
        body = _build_csv(header, rows)
        return Response(
            content=body,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname_stem}.csv"'},
        )
