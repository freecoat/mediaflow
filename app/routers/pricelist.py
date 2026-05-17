"""
Router listino prezzi — CRUD categorie e voci.
Fase 1-bis: aggiunti department_id e keywords su PriceItem.
"""
import json
from datetime import datetime
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, Response
from typing import Optional
from sqlalchemy.orm import Session, joinedload
from app.database import get_db
from app.models import (
    PriceCategory, PriceItem, Department,
    PricelistSnapshot, PricelistSnapshotKind,
)
from app.services import pricelist_snapshot as plsnap
from app.services.rbac import requires_permission
from app.context import current_tenant_id

router = APIRouter(prefix="/pricelist", tags=["pricelist"])

# (categorie + items + import) che non avevano alcun check (audit HIGH #4).
# Pattern identico a quotes.RequireEditQuotes / finance.RequireEditInvoices.
RequireEditPricelist = Depends(requires_permission("edit_pricelist"))


def _tpl():
    from app.main import templates
    return templates


def _parse_keywords(raw: Optional[str]) -> Optional[list]:
    """Parsa una stringa comma-separated in lista pulita di keywords."""
    if raw is None:
        return None
    raw = raw.strip()
    if not raw:
        return []
    return [k.strip().lower() for k in raw.replace(";", ",").split(",") if k.strip()]


@router.get("/", response_class=HTMLResponse)
async def pricelist_page(
    request: Request,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    categories = (
        db.query(PriceCategory)
        .filter(PriceCategory.tenant_id == current_tenant_id())
        .order_by(PriceCategory.sort_order)
        .all()
    )
    departments = (
        db.query(Department)
        .filter(Department.tenant_id == current_tenant_id(), Department.is_active == True)
        .order_by(Department.sort_order, Department.name)
        .all()
    )
    return _tpl().TemplateResponse(
        "pages/pricelist.html",
        {
            "request": request,
            "categories": categories,
            "departments": departments,
            "selected_dept_id": department_id,
        }
    )


# ── Categorie ─────────────────────────────────────────────────
@router.get("/api/categories")
async def list_categories(db: Session = Depends(get_db)):
    cats = (
        db.query(PriceCategory)
        .filter(PriceCategory.tenant_id == current_tenant_id())
        .order_by(PriceCategory.sort_order)
        .all()
    )
    return [{"id": c.id, "name": c.name, "description": c.description,
             "sort_order": c.sort_order, "item_count": len(c.items)} for c in cats]


@router.post("/api/categories", dependencies=[RequireEditPricelist])
async def create_category(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    sort_order: int = Form(0),
    db: Session = Depends(get_db),
):
    c = PriceCategory(
        tenant_id=current_tenant_id(),
        name=name.strip(), description=description, sort_order=sort_order,
    )
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "name": c.name}


@router.put("/api/categories/{cat_id}", dependencies=[RequireEditPricelist])
async def update_category(
    cat_id: int,
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    sort_order: Optional[int] = Form(None),
    db: Session = Depends(get_db),
):
    c = db.query(PriceCategory).filter(
        PriceCategory.id == cat_id,
        PriceCategory.tenant_id == current_tenant_id()
    ).first()
    if not c: raise HTTPException(404, "Categoria non trovata")
    if name is not None: c.name = name.strip()
    if description is not None: c.description = description
    if sort_order is not None: c.sort_order = sort_order
    db.commit()
    return {"id": c.id, "name": c.name}


@router.delete("/api/categories/{cat_id}", dependencies=[RequireEditPricelist])
async def delete_category(cat_id: int, db: Session = Depends(get_db)):
    c = db.query(PriceCategory).filter(
        PriceCategory.id == cat_id,
        PriceCategory.tenant_id == current_tenant_id()
    ).first()
    if not c: raise HTTPException(404)
    if c.items:
        raise HTTPException(400, f"Impossibile eliminare: {len(c.items)} voci collegate")
    db.delete(c); db.commit()
    return {"ok": True}


# ── Voci listino ──────────────────────────────────────────────
@router.get("/api/items")
async def list_items(
    category_id: Optional[int] = None,
    department_id: Optional[int] = None,
    active_only: bool = True,
    db: Session = Depends(get_db),
):
    q = (
        db.query(PriceItem)
        .options(joinedload(PriceItem.category), joinedload(PriceItem.department))
        .filter(PriceItem.tenant_id == current_tenant_id())
    )
    if category_id: q = q.filter(PriceItem.category_id == category_id)
    if department_id: q = q.filter(PriceItem.department_id == department_id)
    if active_only: q = q.filter(PriceItem.is_active == True)
    items = q.all()
    return [
        {
            "id": i.id, "category_id": i.category_id,
            "category": i.category.name if i.category else None,
            "department_id": i.department_id,
            "department": i.department.name if i.department else None,
            "department_color": i.department.color if i.department else None,
            "name": i.name, "description": i.description,
            "unit_pre": i.unit_pre, "unit": i.unit,
            "price_list": i.price_list, "price_average": i.price_average,
            "price_low": i.price_low, "hardcosts": i.hardcosts,
            "keywords": i.keywords or [],
            # v3.5.0-alpha.163 — Voce trasversale (cross_dept) per UI badge + filter
            "cross_dept": bool(getattr(i, "cross_dept", False)),
        }
        for i in items
    ]


@router.get("/api/items/{item_id}")
async def get_item(item_id: int, db: Session = Depends(get_db)):
    i = db.query(PriceItem).filter(
        PriceItem.id == item_id,
        PriceItem.tenant_id == current_tenant_id()
    ).first()
    if not i: raise HTTPException(404, "Voce non trovata")
    return {
        "id": i.id, "category_id": i.category_id, "department_id": i.department_id,
        "name": i.name, "description": i.description,
        "unit_pre": i.unit_pre, "unit": i.unit,
        "price_list": i.price_list, "price_average": i.price_average,
        "price_low": i.price_low, "hardcosts": i.hardcosts,
        "keywords": i.keywords or [],
        "is_active": i.is_active,
    }


@router.post("/api/items", dependencies=[RequireEditPricelist])
async def create_item(
    category_id: int = Form(...),
    department_id: Optional[int] = Form(None),
    name: str = Form(...),
    description: Optional[str] = Form(None),
    unit: str = Form("day"),
    unit_pre: str = Form("per"),
    price_list: Optional[float] = Form(None),
    price_average: Optional[float] = Form(None),
    price_low: Optional[float] = Form(None),
    hardcosts: Optional[float] = Form(None),
    keywords: Optional[str] = Form(None),
    # v3.5.0-alpha.163 — Voce trasversale (Production Management, Overhead...)
    cross_dept: Optional[bool] = Form(False),
    db: Session = Depends(get_db),
):
    item = PriceItem(
        tenant_id=current_tenant_id(),
        category_id=category_id,
        department_id=department_id,
        name=name.strip(),
        description=description,
        unit=unit, unit_pre=unit_pre,
        price_list=price_list,
        price_average=price_average,
        price_low=price_low,
        hardcosts=hardcosts,
        keywords=_parse_keywords(keywords),
        cross_dept=bool(cross_dept),
    )
    db.add(item); db.commit(); db.refresh(item)
    return {"id": item.id, "name": item.name}


@router.put("/api/items/{item_id}", dependencies=[RequireEditPricelist])
async def update_item(
    item_id: int,
    category_id: Optional[int] = Form(None),
    department_id: Optional[int] = Form(None),
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    unit: Optional[str] = Form(None),
    unit_pre: Optional[str] = Form(None),
    price_list: Optional[float] = Form(None),
    price_average: Optional[float] = Form(None),
    price_low: Optional[float] = Form(None),
    hardcosts: Optional[float] = Form(None),
    keywords: Optional[str] = Form(None),
    is_active: Optional[bool] = Form(None),
    cross_dept: Optional[bool] = Form(None),
    db: Session = Depends(get_db),
):
    i = db.query(PriceItem).filter(
        PriceItem.id == item_id,
        PriceItem.tenant_id == current_tenant_id()
    ).first()
    if not i: raise HTTPException(404, "Voce non trovata")
    if category_id is not None: i.category_id = category_id
    if department_id is not None: i.department_id = department_id or None
    if name is not None: i.name = name.strip()
    if description is not None: i.description = description
    if unit is not None: i.unit = unit
    if unit_pre is not None: i.unit_pre = unit_pre
    if price_list is not None: i.price_list = price_list
    if price_average is not None: i.price_average = price_average
    if price_low is not None: i.price_low = price_low
    if hardcosts is not None: i.hardcosts = hardcosts
    if keywords is not None: i.keywords = _parse_keywords(keywords)
    if is_active is not None: i.is_active = is_active
    if cross_dept is not None: i.cross_dept = bool(cross_dept)
    db.commit()
    return {"id": i.id, "name": i.name}


@router.delete("/api/items/{item_id}", dependencies=[RequireEditPricelist])
async def delete_item(item_id: int, db: Session = Depends(get_db)):
    i = db.query(PriceItem).filter(
        PriceItem.id == item_id,
        PriceItem.tenant_id == current_tenant_id()
    ).first()
    if not i: raise HTTPException(404)
    i.is_active = False; db.commit()
    return {"ok": True}


# ── Export / Import ───────────────────────────────────────────

EXPORT_VERSION = "1.0"


@router.get("/api/export")
async def export_pricelist(db: Session = Depends(get_db)):
    """Scarica un dump JSON portabile del listino corrente.
    Categorie e reparti sono referenziati per nome/codice (non per ID), così
    il file può essere reimportato anche su un'altra installazione.
    """
    cats = (
        db.query(PriceCategory)
        .filter(PriceCategory.tenant_id == current_tenant_id())
        .order_by(PriceCategory.sort_order)
        .all()
    )
    items = (
        db.query(PriceItem)
        .options(joinedload(PriceItem.category), joinedload(PriceItem.department))
        .filter(PriceItem.tenant_id == current_tenant_id())
        .all()
    )
    payload = {
        "version": EXPORT_VERSION,
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "tenant_id": current_tenant_id(),
        "categories": [
            {
                "name": c.name,
                "description": c.description,
                "sort_order": c.sort_order,
            }
            for c in cats
        ],
        "items": [
            {
                "category": i.category.name if i.category else None,
                "department_code": i.department.code if i.department else None,
                "name": i.name,
                "description": i.description,
                "unit_pre": i.unit_pre,
                "unit": i.unit,
                "price_list": i.price_list,
                "price_average": i.price_average,
                "price_low": i.price_low,
                "hardcosts": i.hardcosts,
                "keywords": i.keywords or [],
                "is_active": i.is_active,
            }
            for i in items
        ],
    }
    body = json.dumps(payload, ensure_ascii=False, indent=2)
    filename = f"mediaflow_listino_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pricelist_rows_for_export(db: Session) -> tuple[list[str], list[list]]:
    """Header + righe tabellari per export CSV/Excel del listino del tenant."""
    items = (
        db.query(PriceItem)
        .options(joinedload(PriceItem.category), joinedload(PriceItem.department))
        .filter(PriceItem.tenant_id == current_tenant_id())
        .order_by(PriceItem.category_id, PriceItem.name)
        .all()
    )
    headers = [
        "Categoria", "Reparto", "Nome", "Descrizione",
        "Unità pre", "Unità", "Prezzo €", "Prezzo medio €", "Prezzo basso €",
        "Hardcosts €", "Keywords", "Attivo",
    ]
    rows = []
    for i in items:
        rows.append([
            i.category.name if i.category else "",
            i.department.name if i.department else "",
            i.name,
            i.description or "",
            i.unit_pre or "",
            i.unit or "",
            i.price_list if i.price_list is not None else "",
            i.price_average if i.price_average is not None else "",
            i.price_low if i.price_low is not None else "",
            i.hardcosts if i.hardcosts is not None else "",
            ", ".join(i.keywords or []),
            "Sì" if i.is_active else "No",
        ])
    return headers, rows


@router.get("/api/export.csv")
async def export_pricelist_csv(db: Session = Depends(get_db)):
    """Export del listino in CSV (UTF-8 con BOM per compatibilità Excel)."""
    import csv, io
    headers, rows = _pricelist_rows_for_export(db)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    body = "﻿" + buf.getvalue()  # BOM così Excel apre correttamente UTF-8
    filename = f"mediaflow_listino_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        content=body, media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/export.xlsx")
async def export_pricelist_xlsx(db: Session = Depends(get_db)):
    """Export del listino in Excel .xlsx con header bold e larghezze auto."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers, rows = _pricelist_rows_for_export(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Listino"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="6272f5")
    bold_white = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = bold_white
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    # Larghezze auto basate sul contenuto (max 60 char)
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for r in rows:
            v = r[col_idx - 1]
            if v is None: continue
            l = len(str(v))
            if l > max_len: max_len = l
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    filename = f"mediaflow_listino_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/api/import", dependencies=[RequireEditPricelist])
async def import_pricelist(
    file: UploadFile = File(...),
    mode: str = Form("merge"),
    db: Session = Depends(get_db),
):
    """Importa un listino da file JSON.

    mode = "merge"   → categorie e voci con stesso nome vengono aggiornate; nuove
                       voci vengono aggiunte. I dati esistenti non in import sono preservati.
    mode = "replace" → CANCELLA tutte le voci e categorie del tenant, poi importa.
    """
    if mode not in ("merge", "replace"):
        raise HTTPException(400, "mode deve essere 'merge' o 'replace'")

    raw = await file.read()
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise HTTPException(400, f"File non valido: {e}")

    if not isinstance(payload, dict) or "items" not in payload:
        raise HTTPException(400, "File non riconosciuto: manca il campo 'items'")

    if mode == "replace":
        # Soft-delete dei dati esistenti per evitare cascade su QuoteLine
        db.query(PriceItem).filter(PriceItem.tenant_id == current_tenant_id()).delete()
        # Categorie senza voci collegate possono essere eliminate
        cat_ids = [c.id for c in db.query(PriceCategory).filter(PriceCategory.tenant_id == current_tenant_id()).all()]
        for cid in cat_ids:
            db.query(PriceCategory).filter(PriceCategory.id == cid).delete()
        db.flush()

    # Mappa departments per codice (servono per assegnare le voci)
    dept_map = {
        d.code: d.id
        for d in db.query(Department).filter(Department.tenant_id == current_tenant_id()).all()
    }

    # Crea/aggiorna categorie
    cat_map = {}  # nome -> id
    for c_data in payload.get("categories", []):
        name = c_data.get("name")
        if not name: continue
        existing = (
            db.query(PriceCategory)
            .filter(PriceCategory.tenant_id == current_tenant_id(), PriceCategory.name == name)
            .first()
        )
        if existing:
            existing.description = c_data.get("description") or existing.description
            existing.sort_order = c_data.get("sort_order", existing.sort_order)
            cat_map[name] = existing.id
        else:
            new_cat = PriceCategory(
                tenant_id=current_tenant_id(),
                name=name,
                description=c_data.get("description"),
                sort_order=c_data.get("sort_order", 100),
            )
            db.add(new_cat); db.flush()
            cat_map[name] = new_cat.id
    db.flush()

    # Aggiorna mappa con categorie già esistenti non incluse nel file
    for c in db.query(PriceCategory).filter(PriceCategory.tenant_id == current_tenant_id()).all():
        cat_map.setdefault(c.name, c.id)

    created = updated = skipped = 0
    for item_data in payload.get("items", []):
        cat_name = item_data.get("category")
        cat_id = cat_map.get(cat_name) if cat_name else None
        if not cat_id:
            skipped += 1
            continue

        name = item_data.get("name")
        if not name:
            skipped += 1
            continue

        dept_code = item_data.get("department_code")
        dept_id = dept_map.get(dept_code) if dept_code else None

        existing = (
            db.query(PriceItem)
            .filter(
                PriceItem.tenant_id == current_tenant_id(),
                PriceItem.category_id == cat_id,
                PriceItem.name == name,
            )
            .first()
        )

        if existing and mode == "merge":
            existing.description = item_data.get("description")
            existing.unit_pre = item_data.get("unit_pre", "per")
            existing.unit = item_data.get("unit", "day")
            existing.price_list = item_data.get("price_list")
            existing.price_average = item_data.get("price_average")
            existing.price_low = item_data.get("price_low")
            existing.hardcosts = item_data.get("hardcosts")
            existing.keywords = item_data.get("keywords") or []
            existing.department_id = dept_id
            existing.is_active = item_data.get("is_active", True)
            updated += 1
        else:
            db.add(PriceItem(
                tenant_id=current_tenant_id(),
                category_id=cat_id,
                department_id=dept_id,
                name=name,
                description=item_data.get("description"),
                unit_pre=item_data.get("unit_pre", "per"),
                unit=item_data.get("unit", "day"),
                price_list=item_data.get("price_list"),
                price_average=item_data.get("price_average"),
                price_low=item_data.get("price_low"),
                hardcosts=item_data.get("hardcosts"),
                keywords=item_data.get("keywords") or [],
                is_active=item_data.get("is_active", True),
            ))
            created += 1

    db.commit()
    return {
        "ok": True,
        "mode": mode,
        "categories": len(payload.get("categories", [])),
        "items_created": created,
        "items_updated": updated,
        "items_skipped": skipped,
    }


# ── SNAPSHOT LISTINO (v3.5.0-alpha.66.6) ──────────────────────
# Storage persistente di backup/restore del listino. Wrappato sopra il
# servizio app/services/pricelist_snapshot.py.

def _serialize_snapshot(s: PricelistSnapshot, include_payload: bool = False) -> dict:
    out = {
        "id": s.id,
        "name": s.name,
        "description": s.description,
        "kind": s.kind.value if s.kind else "manual",
        "item_count": s.item_count,
        "category_count": s.category_count,
        "department_count": s.department_count,
        "schema_version": s.schema_version,
        "source_app_version": s.source_app_version,
        "created_by_user_id": s.created_by_user_id,
        "created_at": s.created_at.isoformat() + "Z" if s.created_at else None,
        "deleted_at": s.deleted_at.isoformat() + "Z" if s.deleted_at else None,
    }
    if include_payload:
        out["payload"] = s.payload_json
    return out


def _require_edit_pricelist(request: Request):
    from app.services.rbac import current_user_optional, has_permission
    user = current_user_optional(request)
    if not has_permission(user, "edit_pricelist"):
        raise HTTPException(403, "Permesso 'edit_pricelist' richiesto")
    return user


@router.get("/api/snapshots")
async def list_snapshots(
    include_deleted: bool = False,
    db: Session = Depends(get_db),
):
    snaps = plsnap.list_snapshots(
        db,
        tenant_id=current_tenant_id(),
        include_deleted=include_deleted,
    )
    return [_serialize_snapshot(s) for s in snaps]


@router.post("/api/snapshots")
async def create_snapshot(
    request: Request,
    name: str = Form(...),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Crea uno snapshot del listino corrente del tenant."""
    user = _require_edit_pricelist(request)
    name = (name or "").strip()
    if not name:
        raise HTTPException(400, "Il nome è obbligatorio")
    snap = plsnap.create_snapshot_record(
        db,
        tenant_id=current_tenant_id(),
        name=name,
        description=description,
        kind=PricelistSnapshotKind.manual,
        user_id=getattr(user, "id", None),
        user_email=getattr(user, "email", None),
    )
    db.commit()
    return _serialize_snapshot(snap)


@router.get("/api/snapshots/{snap_id}")
async def get_snapshot(snap_id: int, db: Session = Depends(get_db)):
    s = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.id == snap_id,
            PricelistSnapshot.tenant_id == current_tenant_id(),
        )
        .first()
    )
    if not s:
        raise HTTPException(404, "Snapshot non trovato")
    return _serialize_snapshot(s, include_payload=True)


@router.get("/api/snapshots/{snap_id}/download")
async def download_snapshot(snap_id: int, db: Session = Depends(get_db)):
    s = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.id == snap_id,
            PricelistSnapshot.tenant_id == current_tenant_id(),
        )
        .first()
    )
    if not s:
        raise HTTPException(404, "Snapshot non trovato")
    body = json.dumps(s.payload_json, ensure_ascii=False, indent=2)
    safe_stem = "".join(c if c.isalnum() or c in "-_" else "_" for c in s.name)[:80] or f"snapshot-{s.id}"
    ts = s.created_at.strftime("%Y%m%d-%H%M%S") if s.created_at else "snapshot"
    fname = f"mediaflow-listino-{safe_stem}-{ts}.json"
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/api/snapshots/{snap_id}/restore")
async def restore_snapshot(
    snap_id: int,
    request: Request,
    mode: str = Form("merge"),
    db: Session = Depends(get_db),
):
    """Ripristina uno snapshot. mode: 'merge' o 'replace'.

    Se mode='replace' viene creato automaticamente un auto-snapshot del
    listino corrente PRIMA dell'overwrite, per permettere rollback.
    """
    user = _require_edit_pricelist(request)
    if mode not in ("merge", "replace"):
        raise HTTPException(400, "mode deve essere 'merge' o 'replace'")
    s = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.id == snap_id,
            PricelistSnapshot.tenant_id == current_tenant_id(),
            PricelistSnapshot.deleted_at.is_(None),
        )
        .first()
    )
    if not s:
        raise HTTPException(404, "Snapshot non trovato")
    try:
        stats = plsnap.apply_snapshot_payload(
            db,
            tenant_id=current_tenant_id(),
            payload=s.payload_json or {},
            mode=mode,  # type: ignore[arg-type]
            auto_backup=True,
            auto_backup_user_id=getattr(user, "id", None),
        )
    except ValueError as e:
        db.rollback()
        raise HTTPException(400, str(e))
    db.commit()
    return {
        "ok": True,
        "snapshot_id": s.id,
        "snapshot_name": s.name,
        **stats,
    }


@router.delete("/api/snapshots/{snap_id}")
async def delete_snapshot(
    snap_id: int,
    request: Request,
    hard: bool = False,
    db: Session = Depends(get_db),
):
    """Soft-delete (default) o hard-delete (?hard=true) di uno snapshot.

    I preset built-in (kind=preset) non sono cancellabili — vengono
    ricreati al boot da app/data/pricelist_presets/.
    """
    _require_edit_pricelist(request)
    s = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.id == snap_id,
            PricelistSnapshot.tenant_id == current_tenant_id(),
        )
        .first()
    )
    if not s:
        raise HTTPException(404, "Snapshot non trovato")
    if s.kind == PricelistSnapshotKind.preset and hard:
        raise HTTPException(400, "I preset built-in non sono cancellabili definitivamente.")
    if hard:
        plsnap.hard_delete_snapshot(db, s)
    else:
        plsnap.soft_delete_snapshot(db, s)
    db.commit()
    return {"ok": True, "id": snap_id, "hard": hard}


@router.post("/api/snapshots/{snap_id}/restore-deleted")
async def restore_deleted_snapshot(
    snap_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    _require_edit_pricelist(request)
    s = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.id == snap_id,
            PricelistSnapshot.tenant_id == current_tenant_id(),
        )
        .first()
    )
    if not s:
        raise HTTPException(404, "Snapshot non trovato")
    plsnap.restore_deleted_snapshot(db, s)
    db.commit()
    return _serialize_snapshot(s)


@router.post("/api/snapshots/upload")
async def upload_snapshot(
    request: Request,
    file: UploadFile = File(...),
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Carica un file .json esportato (schema 1.0 o 1.1) come snapshot manuale.
    Non applica nulla al listino: lo snapshot resta in lista e va
    ripristinato esplicitamente con `/restore`.
    """
    user = _require_edit_pricelist(request)
    raw = await file.read()
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise HTTPException(400, f"File non valido: {e}")
    if not isinstance(payload, dict) or "items" not in payload:
        raise HTTPException(400, "File non riconosciuto: manca il campo 'items'")
    snap_name = (name or "").strip() or f"Importato {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
    snap = plsnap.create_snapshot_record(
        db,
        tenant_id=current_tenant_id(),
        name=snap_name,
        description=description or f"Importato da file {file.filename!r}",
        kind=PricelistSnapshotKind.manual,
        user_id=getattr(user, "id", None),
        payload=payload,
    )
    db.commit()
    return _serialize_snapshot(snap)


@router.get("/api/presets")
async def list_pricelist_presets(db: Session = Depends(get_db)):
    """Lista preset built-in disponibili in app/data/pricelist_presets/.
    Ognuno è un file .json schema 1.1 caricabile come snapshot.
    """
    presets = []
    for path in plsnap.list_preset_files():
        try:
            data = plsnap.load_preset_payload(path.name)
            presets.append({
                "filename": path.name,
                "description": data.get("description") or path.stem.replace("_", " "),
                "schema_version": data.get("schema_version", "1.0"),
                "source_app_version": data.get("source_app_version"),
                "exported_at": data.get("exported_at"),
                "item_count": len(data.get("items", []) or []),
                "category_count": len(data.get("categories", []) or []),
                "department_count": len(data.get("departments", []) or []),
            })
        except Exception as e:
            presets.append({
                "filename": path.name,
                "error": str(e),
            })
    return presets


@router.post("/api/presets/load")
async def load_preset_as_snapshot(
    request: Request,
    preset_filename: str = Form(...),
    name: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Carica un preset built-in come PricelistSnapshot (kind=preset).
    Idempotente: se esiste già uno snapshot preset con stesso filename
    di origine, restituisce quello esistente senza duplicare.
    """
    user = _require_edit_pricelist(request)
    try:
        payload = plsnap.load_preset_payload(preset_filename)
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))
    snap_name = (name or "").strip() or f"Preset: {Path(preset_filename).stem}"
    # Idempotenza: cerca snapshot preset già esistente per stesso filename
    existing = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.tenant_id == current_tenant_id(),
            PricelistSnapshot.kind == PricelistSnapshotKind.preset,
            PricelistSnapshot.name == snap_name,
            PricelistSnapshot.deleted_at.is_(None),
        )
        .first()
    )
    if existing:
        return _serialize_snapshot(existing)
    snap = plsnap.create_snapshot_record(
        db,
        tenant_id=current_tenant_id(),
        name=snap_name,
        description=payload.get("description") or f"Preset {preset_filename}",
        kind=PricelistSnapshotKind.preset,
        user_id=getattr(user, "id", None),
        payload=payload,
    )
    db.commit()
    return _serialize_snapshot(snap)
