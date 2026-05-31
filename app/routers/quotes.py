"""Router quotazioni — ora ancorate al Progetto."""
from app.services.clock import now_utc
from fastapi import APIRouter, Depends, HTTPException, Request, Form, Response
from fastapi.responses import HTMLResponse
from typing import Optional, List
from datetime import date, timedelta
from sqlalchemy.orm import Session, joinedload
from app.database import get_db
from app.models import (
    Quote, QuoteLine, Job, JobStatus, QuoteStatus,
    PriceItem, PriceCategory, PriceLevel, Project, Client,
    Booking, BookingStatus, JobCostLine, TimePunch,
    DeliveryTemplate, JobDeliverable, DeliverableNature,
)
from app.services.rbac import requires_permission
from app.services.billing_slice_guard import assert_jcl_lock_safe
from app.services.delivery_bucket import template_bucket_options
from app.context import current_tenant_id

router = APIRouter(prefix="/quotes", tags=["quotes"])


def _resolve_item_unit_price(item, price_level) -> float:
    """Risolve unit_price dal PriceItem per il livello scelto.

    v3.5.0-alpha.172.142 (audit) — il pattern `{...}.get(...) or 0.0` silenziava
    a €0 le voci con prezzo None (es. voci-bucket deliverable della migrazione
    α.172.135), nascondendo un dato mancante come "lavoro gratis". Ora logga un
    warning quando il prezzo del livello è assente; il default 0.0 resta per non
    bloccare l'import, ma l'evento è tracciato."""
    chosen = {
        PriceLevel.list_price: item.price_list,
        PriceLevel.average: item.price_average,
        PriceLevel.low: item.price_low,
    }.get(price_level, item.price_list)
    if chosen is None:
        import logging
        logging.getLogger(__name__).warning(
            "PriceItem id=%s ('%s') senza prezzo per livello %s → riga a €0 "
            "(prezzo listino MANCANTE, non zero reale: verifica il listino)",
            getattr(item, "id", "?"),
            getattr(item, "name", None) or getattr(item, "description", "?"),
            getattr(price_level, "value", price_level),
        )
        return 0.0
    return float(chosen)


def _line_price_to_base(db, quote, entered_price: float, from_price_item: bool) -> float:
    """Converte un prezzo riga in valuta base. Prezzo da listino = già base.
    Prezzo digitato manualmente in quote estera = in valuta cliente -> /converti."""
    from app.services import fx, currency as cur
    from app.models import Tenant
    tenant = db.query(Tenant).filter(Tenant.id == current_tenant_id()).first()
    base = (tenant.default_currency if tenant else "EUR").upper()
    ccy = (getattr(quote, "currency", None) or base).upper()
    if from_price_item or ccy == base:
        return float(entered_price)
    rate = fx.get_fx_rate(db, ccy, base)
    if rate is None:
        raise HTTPException(422, "Tasso di cambio non disponibile per la conversione")
    return cur.to_base(float(entered_price), rate)


def _currency_block_for_quote(db, quote) -> dict:
    """Blocco valuta per il payload quote: valuta target + tasso LIVE corrente
    (indicativo) + disclaimer. Importi restano in base; il frontend converte."""
    from app.services import fx, currency as cur
    from app.models import Tenant as _Tenant
    tenant = db.query(_Tenant).filter(_Tenant.id == current_tenant_id()).first()
    base = (tenant.default_currency if tenant else "EUR").upper()
    ccy = (getattr(quote, "currency", None) or base).upper()
    if ccy == base:
        return {
            "currency": base, "base_currency": base, "live_rate": 1.0,
            "symbol": cur.symbol(base), "disclaimer": None, "rate_available": True,
        }
    live = fx.get_fx_rate(db, ccy, base)  # quanti base per 1 ccy
    today = now_utc().strftime("%d/%m/%Y")
    disc = cur.disclaimer(base, ccy, live, today) if live is not None else None
    return {
        "currency": ccy, "base_currency": base, "live_rate": live,
        "symbol": cur.symbol(ccy), "disclaimer": disc,
        "rate_available": live is not None,
    }


# v3.5.0-alpha.66.14.5 — Dependency riusabile per i mutator quote.
# Sostituisce i check inline `if not has_permission(user, "edit_quotes")` e
# li applica anche agli 11 mutator che ne erano sprovvisti (audit HIGH #4).
# Importato come module-level per essere usato in `dependencies=[...]` dei
# decoratori router. La dependency raise 403 se permesso mancante.
RequireEditQuotes = Depends(requires_permission("edit_quotes"))

# v3.5.0-alpha.66.15.2 — Costante locale per leggibilità. Single-tenant attuale
# → 1. In Fase 7 sostituire con `Depends(get_tenant_id)` su ogni endpoint.

CATEGORY_FALLBACK = "Altro"

# v3.5.0-alpha.172.93 (Bundle K2) — Auto-classify DeliverableNature da PriceItem.
# Match su name/keywords case-insensitive. Se hit → nature=physical.
# Cover LTO, HDD, CRU, Blu-Ray/Bluray, DVD, tape/nastro, USB drive, shuttle.
_PHYSICAL_KEYWORDS = (
    "lto", "hdd", "cru", "tape", "nastro", "nastri",
    "blu-ray", "bluray", "blu ray", "dvd",
    "shuttle", "usb drive", "harddisk", "hard disk", "hard-disk",
    "drive consegna", "disco rigido", "supporto fisico",
)


def _infer_deliverable_nature(price_item: Optional[PriceItem]) -> DeliverableNature:
    """Inferisce digital vs physical da nome + keywords del PriceItem.
    Default = digital. Match case-insensitive su substring.
    """
    if not price_item:
        return DeliverableNature.digital
    haystack_parts = [(price_item.name or ""), (price_item.description or "")]
    kw = price_item.keywords
    if isinstance(kw, list):
        haystack_parts.extend(str(x) for x in kw)
    elif isinstance(kw, str):
        haystack_parts.append(kw)
    hay = " ".join(haystack_parts).lower()
    for needle in _PHYSICAL_KEYWORDS:
        if needle in hay:
            return DeliverableNature.physical
    return DeliverableNature.digital


def _next_job_code(db: Session, project: Project) -> str:
    """v3.5.0-alpha.66.14.8 — Genera codice job '{PROJECT_CODE}-J{N}'
    progressivo per il progetto.
    v3.5.0-alpha.116 — Cabling NumberingConfig "job". Variabili supportate:
    YYYY/.../NNN/PROJECT_CODE. Fallback al pattern legacy se config assente
    o se produce collision (job già esiste con quel code).
    """
    base = (project.code or f"P{project.id}").strip()
    # include_deleted=True così i job in cestino non liberano il code
    existing = (
        db.query(Job)
        .execution_options(include_deleted=True)
        .filter(Job.project_id == project.id)
        .all()
    )
    used = {j.code for j in existing if j.code}

    # Try NumberingConfig first
    try:
        from app.services.numbering import gen_doc_code
        code, _ = gen_doc_code(
            db, "job",
            tenant_id=current_tenant_id(),
            project_code=base,
        )
        # Se collision con un job esistente, fallback al while-loop
        if code not in used:
            return code
    except Exception as _e:
        print(f"[job_numbering] gen_doc_code failed, fallback: {_e}")

    # Fallback legacy: while-loop pattern {BASE}-J{N}
    n = 1
    while f"{base}-J{n}" in used:
        n += 1
    return f"{base}-J{n}"


def _create_job_from_quote(db: Session, q: Quote, user_id: Optional[int] = None) -> Job:
    """Crea il Job dalla Quote approvata + JobCostLine da ogni QuoteLine.

    Eredita titolo dal progetto (non dalla quote: spesso coincidono ma il
    riferimento canonico è il progetto). Codice auto-generato {PROJECT}-J{N}.
    Idempotenza: se la quote ha già `q.job` ritorna quello.

    v3.5.0-alpha.144 — Hook materializzazione QuoteAdvanceSchedule → AdvancePayment(pending)
    + AdvancePaymentAllocation (mappa QuoteLine→JCL) + Notification admin/manager.
    Idempotente: skip schedule già materializzati.
    """
    if q.job:
        # Job già collegato: se cancelled lo ri-attivo (riapprovazione della stessa quote
        # dopo un rollback). Se in qualunque altro stato lo ritorno così com'è.
        if q.job.status == JobStatus.cancelled:
            q.job.status = JobStatus.approved
        # v3.5.0-alpha.144 — Anche su re-converti, materializza schedule (idempotente)
        try:
            from app.services.advance_schedule_to_payment import materialize_schedules
            db.flush()
            materialize_schedules(db, q, q.job, user_id, current_tenant_id())
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception(f"materialize_schedules failed (re-converti): {e}")
        return q.job
    project = q.project
    if not project:
        raise HTTPException(400, "Quote senza progetto: impossibile promuovere a job")

    job = Job(
        code=_next_job_code(db, project),
        title=project.title,
        project_id=q.project_id,
        client_id=q.client_id,
        quote_id=q.id,
        status=JobStatus.approved,
        budget_quoted=q.total_after_discount,
    )
    db.add(job)
    db.flush()
    # v3.5.0-alpha.172.2 Restructure — branching unit time-based vs non-time.
    # v3.5.0-alpha.172.14 — Revisione spawn rule per nature:
    #   - time_based (hr/day) → JobCostLine
    #   - deliverable_qty (pc/lot/shot/version) → N row, 1 per ogni unità.
    #     Es. quote "3 DCP" → 3 JobDeliverable distinti (qty_planned=1.0 each)
    #     per permettere conferma individuale + link asset 1:1.
    #   - deliverable_volume (TB/GB) → 1 row, qty_planned = qty quote.
    #     Es. quote "10 TB backup" → 1 deliverable qty_planned=10.0, qty_delivered
    #     incrementato via MHL Yoyotta o manuale.
    #   - manual_allow (allow/lump/fix) → 1 row, qty_planned = qty quote.
    #     Es. quote "1 allow rinegoziazione" → 1 deliverable qty_planned=1.0.
    #     Forfait/lump → 1 conferma manuale.
    from app.services.cost_line_sync import unit_nature_for
    from app.models import JobDeliverable, DeliverableUnitNature, DeliverableBillingStatus
    TIME_UNITS = ("hr", "day")
    # Nature per cui spawn 1 row per qty unitaria (consegne discrete).
    SPAWN_PER_UNIT_NATURES = ("deliverable_qty",)
    for line in q.lines:
        unit_l = (line.unit or "").strip().lower()
        if unit_l in TIME_UNITS:
            db.add(JobCostLine(
                job_id=job.id,
                quote_line_id=line.id,
                price_item_id=line.price_item_id,
                description=line.description,
                quantity_quoted=line.quantity,
                unit=line.unit,
                unit_price=line.unit_price,
                total_quoted=line.total,
                total_expected=line.total,
            ))
        else:
            nature_code = unit_nature_for(line.unit)
            nature = DeliverableUnitNature(nature_code)
            qty_total = float(line.quantity or 0.0)
            up = float(line.unit_price or 0.0)
            if nature_code in SPAWN_PER_UNIT_NATURES:
                # Discreto: 1 row per unità (es. 3 DCP separati).
                n_rows = max(1, int(round(qty_total)))
                per_row_qty = 1.0
            else:
                # Volume/forfait: 1 row aggregato (TB cumulativo o lump sum).
                n_rows = 1
                per_row_qty = qty_total if qty_total > 0 else 1.0
            # v3.5.0-alpha.172.93 (Bundle K2) — auto-classify digital/physical
            pi = db.query(PriceItem).filter(PriceItem.id == line.price_item_id).first() if line.price_item_id else None
            phys_nature = _infer_deliverable_nature(pi)
            for idx in range(n_rows):
                db.add(JobDeliverable(
                    tenant_id=q.tenant_id,
                    job_id=job.id,
                    job_cost_line_id=None,
                    quote_line_id=line.id,
                    price_item_id=line.price_item_id,
                    name=line.description,
                    nature=phys_nature,
                    unit=line.unit,
                    unit_price=up,
                    unit_nature=nature,
                    quantity_planned=per_row_qty,
                    quantity_delivered=0.0,
                    total_quoted=round(per_row_qty * up, 2),
                    total_accrued=0.0,
                    total_cost_accrued=0.0,
                    billing_status=DeliverableBillingStatus.not_billed,
                ))
    db.flush()  # Necessario: JCL.id + Deliverable.id servono al materialize_schedules

    # v3.5.0-alpha.144 — Materializza AdvancePayment(pending) da schedule quote.
    # Fail-soft: errori loggati, non bloccano la conversione quote→job.
    try:
        from app.services.advance_schedule_to_payment import materialize_schedules
        result = materialize_schedules(db, q, job, user_id, current_tenant_id())
        if result["created"]:
            import logging
            logging.getLogger(__name__).info(
                f"[quote→job] materialized {len(result['created'])} AdvancePayment(pending) "
                f"from quote {q.number} (notified {result['notified']} users)"
            )
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception(f"materialize_schedules failed: {e}")

    return job


def _respawn_line_artifacts(db: Session, line: QuoteLine, job: Optional[Job]) -> dict:
    """v3.5.0-alpha.172.99 — Re-spawn JCL/JobDeliverable per una singola QuoteLine.

    Trigger: cambio `unit_nature` (volume↔qty↔manual_allow) O cambio numero
    row spawn-per-unit (qty cambia per `deliverable_qty`). In tutti questi
    casi gli artifacts esistenti NON rappresentano più la quote line.

    Pre-check sicurezza prima di delete:
      - JCL: 0 booking non-cancelled
      - JobDeliverable: quantity_delivered=0 + no link booking_deliverables +
        billing_status=not_billed + confirmed_at=None
      - Altrimenti 409 con messaggio "crea nuova versione quote"

    Out-of-scope MVP: cambio cross-table time↔non-time (JCL↔Deliverable).
    Per ora se nature passa da/a time_based ritorna no-op con warning — l'edit
    delle altre property (description/unit_price) viene già propagato dal
    chiamante. La transizione completa cross-table va via nuova versione quote.
    """
    from app.models import (
        BookingDeliverable, DeliverableBillingStatus,
        DeliverableUnitNature,
    )
    from app.services.cost_line_sync import unit_nature_for, TIME_UNITS

    if job is None:
        return {"respawned": False, "reason": "no_job_yet"}

    unit_l = (line.unit or "").strip().lower()
    target_is_time = unit_l in TIME_UNITS
    target_nature_code = "time_based" if target_is_time else unit_nature_for(line.unit)

    # Artifacts esistenti
    existing_jcl = db.query(JobCostLine).filter(
        JobCostLine.quote_line_id == line.id
    ).all()
    existing_deliv = db.query(JobDeliverable).filter(
        JobDeliverable.quote_line_id == line.id
    ).all()

    current_is_time = bool(existing_jcl) and not existing_deliv
    current_is_deliv = bool(existing_deliv) and not existing_jcl

    # Cross-table change: out-of-scope (NEEDS new quote version)
    if current_is_time != target_is_time and (existing_jcl or existing_deliv):
        raise HTTPException(
            409,
            f"Cambio unit '{line.unit}' richiede passaggio time↔deliverable: "
            f"non supportato a caldo. Crea una nuova versione di quote (clone "
            f"+ modifica + migrate-job)."
        )

    # Same-side change (entrambi deliverable side)
    if current_is_deliv:
        # Detect se serve respawn:
        # - nature cambia, O
        # - target nature=deliverable_qty E count rows != target qty
        current_nature_code = (existing_deliv[0].unit_nature.value
                               if existing_deliv[0].unit_nature else "deliverable_qty")
        nature_changed = (current_nature_code != target_nature_code)
        qty_total = float(line.quantity or 0.0)
        target_n_rows = (max(1, int(round(qty_total)))
                         if target_nature_code == "deliverable_qty" else 1)
        rows_changed = (len(existing_deliv) != target_n_rows)
        if not (nature_changed or rows_changed):
            return {"respawned": False, "reason": "no_change_needed"}

        # SAFETY: tutti i deliverable devono essere "vergini"
        for d in existing_deliv:
            if (d.quantity_delivered or 0.0) > 0.0:
                raise HTTPException(
                    409,
                    f"Consegna '{d.name}' ha quantity_delivered={d.quantity_delivered}: "
                    f"impossibile re-spawn. Annulla la consegna parziale o crea "
                    f"nuova versione di quote."
                )
            if d.confirmed_at:
                raise HTTPException(
                    409,
                    f"Consegna '{d.name}' confermata: impossibile re-spawn. "
                    f"Annulla conferma o crea nuova versione di quote."
                )
            if d.billing_status in (
                DeliverableBillingStatus.in_batch,
                DeliverableBillingStatus.billed,
                DeliverableBillingStatus.paid,
            ):
                raise HTTPException(
                    409,
                    f"Consegna '{d.name}' è {d.billing_status.value}: "
                    f"impossibile re-spawn. Crea nuova versione di quote."
                )
            n_links = db.query(BookingDeliverable).filter(
                BookingDeliverable.job_deliverable_id == d.id
            ).count()
            if n_links > 0:
                raise HTTPException(
                    409,
                    f"Consegna '{d.name}' ha {n_links} booking linkati: "
                    f"impossibile re-spawn. Scollegali o crea nuova versione."
                )

        # SAFE: delete vecchi + spawn nuovi
        for d in existing_deliv:
            db.delete(d)
        db.flush()

        target_nature = DeliverableUnitNature(target_nature_code)
        up = float(line.unit_price or 0.0)
        per_row_qty = (1.0 if target_nature_code == "deliverable_qty"
                       else (qty_total if qty_total > 0 else 1.0))
        pi = (db.query(PriceItem).filter(PriceItem.id == line.price_item_id).first()
              if line.price_item_id else None)
        phys_nature = _infer_deliverable_nature(pi)
        for _idx in range(target_n_rows):
            db.add(JobDeliverable(
                tenant_id=job.tenant_id,
                job_id=job.id,
                job_cost_line_id=None,
                quote_line_id=line.id,
                price_item_id=line.price_item_id,
                name=line.description,
                nature=phys_nature,
                unit=line.unit,
                unit_price=up,
                unit_nature=target_nature,
                quantity_planned=per_row_qty,
                quantity_delivered=0.0,
                total_quoted=round(per_row_qty * up, 2),
                total_accrued=0.0,
                total_cost_accrued=0.0,
                billing_status=DeliverableBillingStatus.not_billed,
            ))
        db.flush()
        return {
            "respawned": True,
            "old_n_rows": len(existing_deliv),
            "old_nature": current_nature_code,
            "new_n_rows": target_n_rows,
            "new_nature": target_nature_code,
        }

    # current_is_time: solo update qty su JCL (già fatto dal chiamante)
    return {"respawned": False, "reason": "time_based_inline_update"}


def _job_has_activity(db: Session, job: Job) -> bool:
    """True se il job ha booking non-cancelled o TimePunch effettivi."""
    active_bk = db.query(Booking).filter(
        Booking.job_id == job.id,
        Booking.status != BookingStatus.cancelled,
    ).first()
    if active_bk:
        return True
    punch = db.query(TimePunch).filter(TimePunch.job_id == job.id).first()
    return punch is not None


def _tpl():
    from app.main import templates
    return templates


def _line_category(line: QuoteLine) -> str:
    """Categoria per il raggruppamento (editor / PDF / export).
    Se `category_override` è valorizzato, prevale su quello del price_item.
    """
    override = (line.category_override or "").strip()
    if override:
        return override
    if line.price_item and line.price_item.category:
        return line.price_item.category.name
    return CATEGORY_FALLBACK


# v3.5.0-alpha.172.4 (Sprint 4 T1) — Alias del helper centralizzato per
# evitare import cross-module ripetuti nei serializer JSON.
def _unit_nature(unit: Optional[str]) -> str:
    from app.services.cost_line_sync import unit_nature_for
    return unit_nature_for(unit)


# v3.5.0-alpha.172.18 — Quote approvata immutabile: HARD-BLOCK su tutte le
# mutazioni dirette (add/update/reorder line, batch ops). Unica via per
# modificare una quote approvata è creare una nuova versione via
# POST /api/{quote_id}/new-version → modificare la draft → migrate-job.
#
# Eccezione: i Consuntivi (quote `is_phantom=True`, anche se status=approved)
# rappresentano il cost report vivo del progetto e DEVONO restare editabili
# per propagazione automatica dalle voci approvate eliminate.
#
# `delete_quote_line` ha logica propria di propagazione a Consuntivo: il
# blocco lì è gestito separatamente (non viene mai 409, ridiretto a phantom).
def _assert_quote_mutable(quote: Quote, action: str = "modifica") -> None:
    """Raise 409 se la quote è approvata e non è un Consuntivo.

    Action è una parola breve che entra nel messaggio d'errore (es. "aggiunta
    voce", "modifica voce", "riordino"). Default "modifica".
    """
    if quote.status == QuoteStatus.approved and not quote.is_phantom:
        raise HTTPException(
            409,
            f"{action.capitalize()} bloccata: la quotazione {quote.number} è "
            f"approvata. Crea una nuova versione (POST /api/{quote.id}/new-version), "
            f"applica le modifiche alla draft e usa migrate-job per propagare al Job."
        )


def _recalc_quote(quote: Quote) -> None:
    """
    Cascata sconti:
      1. line_total = qty × unit_price × (1 + allowance) × (1 − line_discount_pct)
      2. subtotal_gross = Σ qty × unit_price × (1 + allowance)        (no sconti)
      3. categoria_subtotale = Σ line_total per categoria
      4. categoria_totale = categoria_subtotale × (1 − category_discount_pct)
      5. subtotal = Σ categoria_totale                                  (post line + cat)
      6. total_after_discount = subtotal × (1 + package_discount)       (package è negativo per retrocompat)
      7. total_with_vat = total_after_discount × (1 + vat_rate/100)

    v3.5.0-alpha.27: le righe con `is_optional=True` hanno il proprio
    `total` calcolato ma NON contribuiscono a subtotal_gross / cat_buckets /
    subtotal / total_after_discount / total_with_vat. Sono mostrate in un
    blocco "Optional aggiuntivi" separato (UI + PDF).
    """
    from app.services.cost_line_sync import unit_nature_for
    cat_disc = quote.category_discounts or {}
    subtotal_gross = 0.0
    # v3.5.0-alpha.172.4 (Sprint 4 T1) — split lordo JCL vs Deliverable.
    # JCL = unit time_based (hr/day/...). Deliverable = il resto (pc/TB/lump/...).
    subtotal_gross_jcl = 0.0
    subtotal_gross_deliverable = 0.0
    cat_buckets: dict[str, float] = {}

    for l in quote.lines:
        gross = (l.quantity or 0.0) * (l.unit_price or 0.0) * (1 + (l.allowance or 0.0))
        net_after_line = gross * (1 - (l.line_discount_pct or 0.0))
        l.total = round(net_after_line, 2)
        if l.is_optional:
            continue  # totale calcolato ma fuori dai subtotali
        subtotal_gross += gross
        if unit_nature_for(l.unit) == "time_based":
            subtotal_gross_jcl += gross
        else:
            subtotal_gross_deliverable += gross
        cat_key = _line_category(l)
        cat_buckets[cat_key] = cat_buckets.get(cat_key, 0.0) + net_after_line

    subtotal_after_cat = sum(
        bucket * (1 - float(cat_disc.get(cat_key, 0.0)))
        for cat_key, bucket in cat_buckets.items()
    )

    total_after = subtotal_after_cat * (1 + (quote.package_discount or 0.0))
    total_with_vat = total_after * (1 + (quote.vat_rate or 0.0) / 100)

    quote.subtotal_gross = round(subtotal_gross, 2)
    quote.subtotal_gross_jcl = round(subtotal_gross_jcl, 2)
    quote.subtotal_gross_deliverable = round(subtotal_gross_deliverable, 2)
    quote.subtotal = round(subtotal_after_cat, 2)
    quote.total_after_discount = round(total_after, 2)
    quote.total_with_vat = round(total_with_vat, 2)


@router.get("/", response_class=HTMLResponse)
async def quotes_page(request: Request, db: Session = Depends(get_db)):
    # v3.5.0-alpha.66.15.2 — tenant scope (R1)
    quotes = db.query(Quote).filter(
        Quote.tenant_id == current_tenant_id(),
    ).options(
        joinedload(Quote.client),
        joinedload(Quote.project),
    ).order_by(Quote.created_at.desc()).all()
    projects = db.query(Project).filter(
        Project.tenant_id == current_tenant_id(),
    ).options(joinedload(Project.client)).order_by(Project.title).all()
    return _tpl().TemplateResponse(
        "pages/quotes.html",
        {"request": request, "quotes": quotes, "projects": projects},
    )


@router.get("/api")
async def list_quotes(
    project_id: Optional[int] = None,
    include_superseded: bool = False,
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.156 — Default: nasconde versioni superseded.
    v3.5.0-alpha.172.12 — Default: collassa la catena versioni per `parent_quote_id`,
    mostra solo la LEAF (ultima versione, no child) + `versions_count` nella
    risposta. Param `include_superseded=true` espande la storia completa.

    Una quote è "leaf" se:
      - nessun'altra quote ha `parent_quote_id == self.id`
      - AND `superseded_by_id IS NULL`
    """
    # v3.5.0-alpha.66.15.2 — tenant scope (R1)
    q = db.query(Quote).filter(
        Quote.tenant_id == current_tenant_id(),
    ).options(joinedload(Quote.client), joinedload(Quote.project))
    if project_id:
        q = q.filter(Quote.project_id == project_id)
    all_qs = q.order_by(Quote.created_at.desc()).all()

    # Pre-calcolo: per ogni quote, c'è una figlia che la referenzia come parent?
    has_child: dict[int, bool] = {qq.id: False for qq in all_qs}
    parent_map: dict[int, Optional[int]] = {qq.id: qq.parent_quote_id for qq in all_qs}
    for qq in all_qs:
        if qq.parent_quote_id and qq.parent_quote_id in has_child:
            has_child[qq.parent_quote_id] = True

    def _is_leaf(qq) -> bool:
        return (not has_child.get(qq.id)) and (qq.superseded_by_id is None)

    # Catena: count delle versioni risalendo parent_quote_id fino a root + figli.
    # Pre-calcolo root_id per evitare hop ripetuti.
    root_of: dict[int, int] = {}
    def _root(qid: int) -> int:
        if qid in root_of:
            return root_of[qid]
        cur = qid
        seen = set()
        while cur not in seen:
            seen.add(cur)
            p = parent_map.get(cur)
            if not p or p == cur:
                root_of[qid] = cur
                return cur
            cur = p
        root_of[qid] = cur
        return cur

    chain_size: dict[int, int] = {}
    for qq in all_qs:
        r = _root(qq.id)
        chain_size[r] = chain_size.get(r, 0) + 1

    visible = all_qs if include_superseded else [qq for qq in all_qs if _is_leaf(qq)]
    # v3.5.0-alpha.172.97 — base_code + version_number derivati da Quote.number
    # per folder-view UI (raggruppa per base_code, ordina per version desc).
    from app.services.numbering import split_version_suffix
    return [
        {
            "id": qq.id, "number": qq.number, "version": qq.version,
            "title": qq.title, "status": qq.status,
            "project_id": qq.project_id,
            "project_title": qq.project.title if qq.project else None,
            "client": qq.client.name if qq.client else None,
            "issue_date": str(qq.issue_date),
            "valid_until": str(qq.valid_until) if qq.valid_until else None,
            "subtotal": qq.subtotal,
            "total_after_discount": qq.total_after_discount,
            "total_with_vat": qq.total_with_vat,
            "has_job": qq.job is not None,
            "from_deliverables": qq.generated_from_deliverables,
            # v3.5.0-alpha.172.12 — Catena versioni collassata
            "versions_count": chain_size.get(_root(qq.id), 1),
            "root_quote_id": _root(qq.id),
            "is_leaf": _is_leaf(qq),
            # v3.5.0-alpha.172.97 — parent + created_at + base_code + version_number per folder-view
            "parent_quote_id": qq.parent_quote_id,
            "created_at": qq.created_at.isoformat() if qq.created_at else None,
            "base_code": split_version_suffix(qq.number)[0],
            "version_number": split_version_suffix(qq.number)[1],
        }
        for qq in visible
    ]


@router.get("/api/{quote_id}/booking-lines")
async def get_quote_booking_lines(
    quote_id: int,
    dept_ids: Optional[str] = None,  # CSV "1,2,3"
    db: Session = Depends(get_db),
):
    """Linee selezionabili in un booking, partendo dalla quote (v3.4.53).

    Ritorna le lavorazioni della quote filtrate per reparto delle risorse
    selezionate (passate in `dept_ids` come CSV). Job nascosto: l'UI booking
    parla solo in termini di quote+lavorazione.

    Modalità:
    - quote `approved` con Job: ritorna `JobCostLine` (kind=cost_line)
    - quote `draft|sent`: ritorna `QuoteLine` (kind=quote_line) — il booking save
      farà reverse-attach implicito al primo uso (approva la quote, crea il
      JobCostLine, link al booking)

    Filtro dept: include linee con `price_item.department_id ∈ dept_ids` o NULL
    (voce non assegnata a reparto = visibile a tutti). Se `dept_ids` vuoto,
    nessun filtro reparto applicato.
    """
    quote = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item),
        joinedload(Quote.job).joinedload(Job.cost_lines),
    ).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(404, "Quote non trovata")

    dept_set: Optional[set[int]] = None
    if dept_ids:
        try:
            dept_set = {int(x) for x in dept_ids.split(",") if x.strip()}
        except ValueError:
            raise HTTPException(400, "dept_ids deve essere CSV di interi")

    def _dept_match(price_item) -> bool:
        if dept_set is None:
            return True
        if price_item is None:
            return True  # riga senza price_item → sempre visibile
        # v3.5.0-alpha.163 — Voce trasversale (Production Management, Overhead...):
        # accetta qualsiasi reparto risorsa, no filtro.
        if getattr(price_item, "cross_dept", False):
            return True
        pid = getattr(price_item, "department_id", None)
        return pid is None or pid in dept_set

    out = []
    if quote.status == QuoteStatus.approved and quote.job:
        # JobCostLines (canoniche per quote già approvata)
        # Carico esplicitamente price_item su ogni cost line (la relationship esiste in v3.4.33)
        from app.models import PriceItem
        pi_ids = {l.price_item_id for l in quote.job.cost_lines if l.price_item_id}
        pi_map = {}
        if pi_ids:
            pis = db.query(PriceItem).options(joinedload(PriceItem.department)).filter(PriceItem.id.in_(pi_ids)).all()
            pi_map = {p.id: p for p in pis}
        for ln in quote.job.cost_lines:
            pi = pi_map.get(ln.price_item_id) if ln.price_item_id else None
            if not _dept_match(pi):
                continue
            dept = getattr(pi, "department", None) if pi else None
            out.append({
                "id": ln.id, "kind": "cost_line",
                "description": ln.description,
                "unit": ln.unit, "quantity_quoted": ln.quantity_quoted,
                "is_extra": ln.is_extra,
                "department_id": dept.id if dept else None,
                "department_name": dept.name if dept else None,
                "department_color": dept.color if dept else None,
            })
    else:
        # QuoteLines (la quote è ancora in trattativa). Il booking save attiverà
        # reverse-attach implicit-approval per materializzare la JobCostLine.
        for ln in quote.lines:
            pi = ln.price_item
            if not _dept_match(pi):
                continue
            dept = getattr(pi, "department", None) if pi else None
            out.append({
                "id": ln.id, "kind": "quote_line",
                "description": ln.description,
                "unit": ln.unit, "quantity_quoted": ln.quantity,
                "is_extra": False,
                "department_id": dept.id if dept else None,
                "department_name": dept.name if dept else None,
                "department_color": dept.color if dept else None,
            })

    return {
        "quote_id": quote.id, "quote_number": quote.number,
        "quote_status": quote.status.value if hasattr(quote.status, "value") else str(quote.status),
        "is_phantom": getattr(quote, "is_phantom", False),
        "has_job": quote.job is not None,
        "job_id": quote.job.id if quote.job else None,
        "job_code": quote.job.code if quote.job else None,
        "filter_dept_ids": sorted(dept_set) if dept_set else None,
        "lines": out,
    }


@router.post("/api/{quote_id}/promote-line-to-cost-line", dependencies=[RequireEditQuotes])
async def promote_line_to_cost_line(
    quote_id: int,
    request: Request,
    quote_line_id: int = Form(...),
    db: Session = Depends(get_db),
):
    """Materializza una QuoteLine → JobCostLine al volo (v3.4.53).

    Usato dal modal booking quando l'utente seleziona una linea di una quote
    in trattativa (`draft|sent`) come bersaglio del booking. Effetti:
    - Se la quote è in `draft|sent`: approva implicitamente + notifica AM
      (kind=quote_reverse_approval, severity=action_required)
    - Crea il Job (forward-flow standard) se manca
    - Crea la JobCostLine corrispondente alla QuoteLine se manca
    - Idempotente: se tutto esiste già, no-op + ritorna gli ID

    Ritorna `{quote_id, job_id, cost_line_id, was_implicit_approval}` per il
    binding diretto del booking.
    """
    from app.services.rbac import current_user_optional
    actor = current_user_optional(request)

    quote = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item),
        joinedload(Quote.job).joinedload(Job.cost_lines),
        joinedload(Quote.project),
    ).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(404, "Quote non trovata")

    line = next((ln for ln in quote.lines if ln.id == quote_line_id), None)
    if not line:
        raise HTTPException(404, "QuoteLine non trovata in questa quote")

    was_implicit = False
    prev_status = quote.status
    if quote.status in (QuoteStatus.draft, QuoteStatus.sent):
        quote.status = QuoteStatus.approved
        was_implicit = True

    job = _create_job_from_quote(db, quote)  # idempotente
    db.flush()

    # Trova o crea la JobCostLine corrispondente alla QuoteLine
    jcl = db.query(JobCostLine).filter(
        JobCostLine.quote_line_id == line.id, JobCostLine.job_id == job.id
    ).first()
    if not jcl:
        jcl = JobCostLine(
            job_id=job.id,
            quote_line_id=line.id,
            price_item_id=line.price_item_id,
            description=line.description,
            quantity_quoted=line.quantity,
            unit=line.unit,
            unit_price=line.unit_price,
            total_quoted=line.total,
            total_expected=line.total,
        )
        db.add(jcl)
        db.flush()

    db.commit()
    db.refresh(quote); db.refresh(job); db.refresh(jcl)

    if was_implicit:
        try:
            from app.services.notifications import notify_permission
            from app.models import NotificationKind, NotificationSeverity
            notify_permission(
                db,
                permission="edit_quotes",
                exclude_user_ids=[actor.id] if actor else None,
                kind=NotificationKind.quote_reverse_approval.value,
                severity=NotificationSeverity.action_required.value,
                title=f"Quote {quote.number} approvata implicitamente (booking → linea)",
                body=(
                    f"Booking su progetto '{quote.project.title if quote.project else '?'}' "
                    f"ha attivato la linea '{line.description}' (€ {line.total:.2f}). "
                    f"Stato precedente: {prev_status.value}. Verifica e attiva eventualmente "
                    f"procedure di migrate-job o versioning standard."
                ),
                link=f"/quotes#{quote.id}",
                payload={
                    "quote_id": quote.id, "quote_number": quote.number,
                    "job_id": job.id, "line_id": line.id,
                    "previous_status": prev_status.value,
                },
                actor_user_id=actor.id if actor else None,
            )
        except Exception as e:
            print(f"[promote-line] notify failed: {e}")

    return {
        "quote_id": quote.id, "quote_number": quote.number,
        "quote_status": quote.status.value,
        "previous_status": prev_status.value,
        "job_id": job.id, "job_code": job.code,
        "cost_line_id": jcl.id,
        "cost_line_description": jcl.description,
        "was_implicit_approval": was_implicit,
    }


@router.post("/api/reverse-attach", dependencies=[RequireEditQuotes])
async def reverse_attach(
    request: Request,
    project_id: int = Form(...),
    mode: str = Form(...),  # "attach_existing" | "create_phantom"
    target_quote_id: Optional[int] = Form(None),
    price_item_id: int = Form(...),
    booking_hours: float = Form(0.0),  # somma ore-persona del booking corrente
    quantity_override: Optional[float] = Form(None),
    phantom_title: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Reverse-flow v3.4.52: booking su progetto senza quote attiva.

    - mode="attach_existing": riga aggiunta a quote draft|sent → approvazione
      implicita → ensure Job → notifica account managers (edit_quotes).
    - mode="create_phantom": crea Quote(is_phantom=True, status=approved) +
      una line + Job auto-creato → notifica account managers.

    Quantità: se quantity_override è valorizzato lo usa, altrimenti la deriva
    da booking_hours secondo l'unità della voce listino.
    """
    from app.services.rbac import current_user_optional
    from app.services.reverse_quote import (
        attach_to_pending_quote, compute_quantity_from_hours,
        create_phantom_quote_with_line,
    )
    actor = current_user_optional(request)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Progetto non trovato")
    pi = db.query(PriceItem).filter(PriceItem.id == price_item_id).first()
    if not pi:
        raise HTTPException(404, "Voce listino non trovata")

    qty = float(quantity_override) if quantity_override is not None else compute_quantity_from_hours(
        float(booking_hours or 0.0), pi.unit or "day"
    )
    if qty <= 0:
        raise HTTPException(400, "quantity calcolata <= 0: passa booking_hours o quantity_override > 0")

    if mode == "attach_existing":
        if not target_quote_id:
            raise HTTPException(400, "target_quote_id richiesto in mode=attach_existing")
        result = attach_to_pending_quote(
            db, target_quote_id, project_id, price_item_id, qty, actor=actor,
        )
    elif mode == "create_phantom":
        result = create_phantom_quote_with_line(
            db, project, price_item_id, qty, title=phantom_title, actor=actor,
        )
    else:
        raise HTTPException(400, "mode deve essere 'attach_existing' o 'create_phantom'")

    result["client_name"] = project.client.name if project.client else None
    result["project_code"] = project.code
    result["project_title"] = project.title
    result["computed_quantity"] = qty
    result["price_item_unit"] = pi.unit
    return result


@router.post("/api", dependencies=[RequireEditQuotes])
async def create_quote(
    number: Optional[str] = Form(None),
    project_id: int = Form(...),
    title: str = Form(...),
    issue_date: date = Form(...),
    valid_until: Optional[date] = Form(None),
    production_material: Optional[str] = Form(None),
    length_minutes: Optional[float] = Form(None),
    fps: Optional[str] = Form(None),
    delivery_format: Optional[str] = Form(None),
    shooting_days: Optional[int] = Form(None),
    package_discount: float = Form(0.0),
    vat_rate: float = Form(22.0),
    notes: Optional[str] = Form(None),
    payment_terms: Optional[str] = Form(None),
    # v3.5.0-alpha.137 — Multi-currency quote
    currency: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Progetto non trovato")
    # v3.5.0-alpha.172.140 — Numero auto da naming convention (NumberingConfig)
    # se non fornito: la creazione da /projects non prefilla il campo numero.
    # Source of truth lato server → ogni path (manuale, projects, futuro) eredita
    # la convenzione senza duplicare logica nel client.
    if not number or not str(number).strip():
        _cli = db.query(Client).filter(Client.id == project.client_id).first()
        number = _next_quote_number_progressive(db, project=project, client=_cli)
    # v3.5.0-alpha.172.140 — Validità default = 2 settimane se non specificata.
    if valid_until is None:
        valid_until = issue_date + timedelta(days=14)
    # v3.5.0-alpha.137 — Currency setup: default = tenant base, fx_rate fixed snapshot
    from app.models import Tenant
    from app.services.fx import get_fx_rate
    from datetime import datetime as _dt
    tenant = db.query(Tenant).filter(Tenant.id == current_tenant_id()).first()
    base_ccy = (tenant.default_currency if tenant else "EUR").upper()
    q_ccy = (currency or base_ccy).upper().strip()
    if len(q_ccy) != 3:
        q_ccy = base_ccy
    if q_ccy == base_ccy:
        fx_rate, fx_at = 1.0, None
    else:
        fx_rate = get_fx_rate(db, q_ccy, base_ccy)
        if fx_rate is None:
            # Fallback: salva quote in valuta non-base ma con rate=1.0 + warning
            # (UI può segnalare "tasso non disponibile, ricaricare").
            fx_rate, fx_at = 1.0, None
        else:
            fx_at = now_utc()
    q = Quote(
        number=number, project_id=project_id, client_id=project.client_id,
        title=title, issue_date=issue_date, valid_until=valid_until,
        production_material=production_material or project.shooting_format,
        length_minutes=length_minutes or project.length_minutes,
        fps=fps or project.fps,
        delivery_format=delivery_format or project.delivery_format,
        shooting_days=shooting_days,
        package_discount=package_discount, vat_rate=vat_rate,
        notes=notes, payment_terms=payment_terms,
        currency=q_ccy, fx_rate_to_base=fx_rate, fx_rate_fixed_at=fx_at,
    )
    db.add(q); db.commit(); db.refresh(q)
    return {"id": q.id, "number": q.number, "currency": q.currency,
            "fx_rate_to_base": q.fx_rate_to_base}


@router.get("/api/{quote_id}")
async def get_quote(quote_id: int, db: Session = Depends(get_db)):
    q = db.query(Quote).options(
        joinedload(Quote.client), joinedload(Quote.project),
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category),
    ).filter(Quote.id == quote_id).first()
    if not q:
        raise HTTPException(404, "Quotazione non trovata")
    return {
        "id": q.id, "number": q.number, "version": q.version,
        "title": q.title, "status": q.status,
        "project_id": q.project_id,
        "project_title": q.project.title if q.project else None,
        "client_id": q.client_id,
        "client": q.client.name if q.client else None,
        "issue_date": str(q.issue_date),
        "valid_until": str(q.valid_until) if q.valid_until else None,
        "production_material": q.production_material,
        "length_minutes": q.length_minutes, "fps": q.fps,
        "delivery_format": q.delivery_format, "shooting_days": q.shooting_days,
        "package_discount": q.package_discount, "vat_rate": q.vat_rate,
        "category_discounts": q.category_discounts or {},
        "category_order": q.category_order or [],
        "subtotal_gross": q.subtotal_gross,
        # v3.5.0-alpha.172.4 (Sprint 4 T1) — split JCL/Deliverable per editor tabs
        "subtotal_gross_jcl": getattr(q, "subtotal_gross_jcl", 0.0) or 0.0,
        "subtotal_gross_deliverable": getattr(q, "subtotal_gross_deliverable", 0.0) or 0.0,
        "subtotal": q.subtotal,
        "total_after_discount": q.total_after_discount,
        "total_with_vat": q.total_with_vat,
        "notes": q.notes, "payment_terms": q.payment_terms,
        "shipping_markup_pct": getattr(q, "shipping_markup_pct", 15.0),
        # v3.5.0-alpha.111 — Scadenze fatturazione (propagate a Project all'approve)
        "billing_frequency": getattr(q, "billing_frequency", None),
        "billing_terms_days": getattr(q, "billing_terms_days", None),
        # v3.5.0-alpha.137 — Multi-currency
        "currency": getattr(q, "currency", "EUR"),
        "fx_rate_to_base": getattr(q, "fx_rate_to_base", 1.0),
        "fx_rate_fixed_at": q.fx_rate_fixed_at.isoformat() if getattr(q, "fx_rate_fixed_at", None) else None,
        "currency_block": _currency_block_for_quote(db, q),
        # v3.5.0-alpha.139 — Termini di acconto definiti in quote
        "advance_schedules": _get_schedules_serialized(db, q.id),
        "generated_from_deliverables": q.generated_from_deliverables,
        "source_document_name": q.source_document_name,
        "subtotal_optional": round(
            sum((l.total or 0.0) for l in q.lines if l.is_optional), 2
        ),
        "lines": [
            {
                "id": l.id, "section": l.section, "position": l.position,
                "description": l.description, "detail": l.detail,
                "quantity": l.quantity, "unit": l.unit,
                "price_level": l.price_level, "unit_price": l.unit_price,
                "allowance": l.allowance,
                "line_discount_pct": l.line_discount_pct or 0.0,
                "total": l.total,
                "hardcosts": l.hardcosts, "sort_order": l.sort_order,
                "price_item_id": l.price_item_id,
                "category": _line_category(l),
                "category_override": l.category_override,
                "is_optional": bool(l.is_optional),
                "section_label": l.section_label or None,
                # v3.5.0-alpha.64: link a JCL d'origine se la riga è nata da
                # refer-to-sales (badge "↪ Da JCL #X" cliccabile).
                "referred_from_jcl_id": l.referred_from_jcl_id,
                # v3.5.0-alpha.172.4 (Sprint 4 T1) — nature derivata dall'unit
                # per filtro tab Lavorazioni (time_based) / Consegne (resto).
                "unit_nature": _unit_nature(l.unit),
            }
            for l in sorted(q.lines, key=lambda x: x.sort_order)
        ],
    }


@router.put("/api/{quote_id}/status", dependencies=[RequireEditQuotes])
async def update_quote_status(
    quote_id: int, status: QuoteStatus = Form(...), db: Session = Depends(get_db),
):
    """Aggiorna lo stato della quote.

    Side-effect: la transizione di stato gestisce automaticamente il Job collegato.
    - draft/sent → approved: crea il Job + JobCostLine (idempotente se già esiste)
    - approved → altro stato: cancella il Job se non ha attività; blocca con 400
      se ha booking attivi o TimePunch (preserva storico operativo).
    """
    q = (
        db.query(Quote)
        .options(joinedload(Quote.lines), joinedload(Quote.project), joinedload(Quote.job))
        .filter(Quote.id == quote_id)
        .first()
    )
    if not q:
        raise HTTPException(404)

    prev = q.status
    new = status
    promoted_job = None
    cancelled_job_id = None

    if new == QuoteStatus.approved and prev != QuoteStatus.approved:
        # v3.5.0-alpha.172.97.1 — HARD-BLOCK: se la quote ha un parent gia'
        # approved con Job collegato, approvare direttamente creerebbe un Job
        # duplicato (bypass del workflow migrate-job che e' l'unico path corretto
        # per propagare una nuova versione al Job esistente).
        # Caso reale incontrato: v2 approvata via PUT/status invece di migrate-job
        # → Job duplicato con 111 deliverable spawn-per-unit + v1+v2 entrambe
        # approved, stato incongruente non recuperabile via UI.
        if q.parent_quote_id:
            parent = (
                db.query(Quote)
                .options(joinedload(Quote.job))
                .filter(Quote.id == q.parent_quote_id).first()
            )
            if parent and parent.status == QuoteStatus.approved and parent.job:
                raise HTTPException(
                    409,
                    f"Versione collegata: usa migrate-job invece di approvare direttamente. "
                    f"La versione precedente {parent.number} è già approved con Job "
                    f"{parent.job.code}. Approvare questa versione creerebbe un Job duplicato. "
                    f"POST /quotes/api/{q.id}/migrate-job per propagare al Job esistente."
                )
        # v3.5.0-alpha.111 — Propaga scadenze fatturazione Quote → Project
        # SE Project non ha override esplicito (campi NULL).
        if q.project:
            if q.billing_frequency and not getattr(q.project, "billing_frequency", None):
                q.project.billing_frequency = q.billing_frequency
            elif q.billing_frequency:
                # Project ha già un valore: aggiorna SOLO se il Project ha il default
                # `monthly` (significa che non è stato ancora personalizzato).
                if getattr(q.project, "billing_frequency", "monthly") == "monthly":
                    q.project.billing_frequency = q.billing_frequency
            if q.billing_terms_days and not getattr(q.project, "billing_terms_days", None):
                q.project.billing_terms_days = q.billing_terms_days
        # Approvazione: crea il job se non esiste
        promoted_job = _create_job_from_quote(db, q)
        # v3.4.56 — warning non bloccante: se il job non ha JobResourceAssignment,
        # notifica chi può assegnare (producer/manager). Auto-assignment via booking
        # resta disponibile (richiesta conferma client-side).
        try:
            from app.models import JobResourceAssignment, NotificationKind, NotificationSeverity
            from app.services.notifications import notify_permission
            ass_count = db.query(JobResourceAssignment).filter(
                JobResourceAssignment.job_id == promoted_job.id
            ).count()
            if ass_count == 0:
                notify_permission(
                    db, permission="assign_resources",
                    kind=NotificationKind.quote_approved_no_resources.value,
                    severity=NotificationSeverity.action_required.value,
                    title=f"Quote {q.number} approvata: nessuna risorsa assegnata",
                    body=(
                        f"Il job {promoted_job.code} ({promoted_job.title}) è stato creato "
                        f"ma non ha ancora risorse assegnate. Aggiungile manualmente in "
                        f"/projects/{q.project_id} oppure scattano in automatico al primo "
                        f"booking (con richiesta di conferma)."
                    ),
                    link=f"/projects/{q.project_id}",
                    payload={
                        "quote_id": q.id, "quote_number": q.number,
                        "job_id": promoted_job.id, "job_code": promoted_job.code,
                        "project_id": q.project_id,
                    },
                )
        except Exception as e:
            print(f"[quote-approve] notify no_resources failed: {e}")
    elif prev == QuoteStatus.approved and new != QuoteStatus.approved and q.job:
        # Disapprovazione: cancella il job se senza attività, altrimenti blocca
        if _job_has_activity(db, q.job):
            raise HTTPException(
                400,
                f"Impossibile riportare la quote a '{new.value}': il job {q.job.code} "
                "ha attività (booking o timbrature). Cancella prima le attività."
            )
        cancelled_job_id = q.job.id
        q.job.status = JobStatus.cancelled

    q.status = new
    db.commit()

    resp = {"id": q.id, "status": q.status}
    if promoted_job:
        db.refresh(promoted_job)
        resp["job_created"] = {
            "id": promoted_job.id,
            "code": promoted_job.code,
            "title": promoted_job.title,
            "lines_count": len(q.lines),
        }
    if cancelled_job_id:
        resp["job_cancelled_id"] = cancelled_job_id
    return resp


@router.put("/api/{quote_id}")
async def update_quote(
    quote_id: int,
    request: Request,
    package_discount: Optional[float] = Form(None),
    vat_rate: Optional[float] = Form(None),
    notes: Optional[str] = Form(None),
    payment_terms: Optional[str] = Form(None),
    # v3.5.0-alpha.7.5 — rinomina di title (sempre) e number (solo draft)
    title: Optional[str] = Form(None),
    number: Optional[str] = Form(None),
    # v3.5.0-alpha.106 — Clausola ricarico spedizioni
    shipping_markup_pct: Optional[float] = Form(None),
    # v3.5.0-alpha.111 — Scadenze fatturazione (propagate a Project all'approve)
    billing_frequency: Optional[str] = Form(None),
    billing_terms_days: Optional[int] = Form(None),
    # v3.5.0-alpha.137 — Multi-currency: cambio valuta = refresh fx + snapshot
    currency: Optional[str] = Form(None),
    refresh_fx: Optional[str] = Form(None),  # "true" → forza refresh tasso
    db: Session = Depends(get_db),
):
    # v3.4.38 (R3.2): guard permission edit_quotes
    from app.services.rbac import current_user_optional, has_permission
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di modificare le quotazioni")
    q = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
    ).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404)

    if title is not None:
        new_title = title.strip()
        if not new_title:
            raise HTTPException(400, "Il titolo non può essere vuoto")
        q.title = new_title
    if number is not None:
        new_number = number.strip()
        if not new_number:
            raise HTTPException(400, "Il numero non può essere vuoto")
        if new_number != q.number:
            # Modifica del numero ammessa solo finché la quote è in draft.
            # Una volta "sent" (proposta al cliente) o "approved" (con job
            # collegato), il numero è il riferimento ufficiale e non va toccato.
            if q.status != QuoteStatus.draft:
                raise HTTPException(409,
                    f"Il numero non può essere modificato in stato '{q.status.value}'. "
                    "Solo le bozze sono rinominabili.")
            # Unicità: bypass del filter soft-delete perché le quote in cestino
            # occupano comunque il number (vincolo UNIQUE su DB).
            collision = (db.query(Quote)
                           .execution_options(include_deleted=True)
                           .filter(Quote.number == new_number, Quote.id != q.id)
                           .first())
            if collision:
                raise HTTPException(409,
                    f"Esiste già una quote con number '{new_number}' (eventualmente nel cestino)")
            q.number = new_number

    if package_discount is not None: q.package_discount = package_discount
    if vat_rate is not None: q.vat_rate = vat_rate
    if notes is not None: q.notes = notes
    if payment_terms is not None: q.payment_terms = payment_terms
    if shipping_markup_pct is not None:
        q.shipping_markup_pct = max(0.0, min(float(shipping_markup_pct), 100.0))
    if billing_frequency is not None:
        bf = billing_frequency.strip() or None
        if bf and bf not in {"monthly", "quarterly", "milestone", "on_completion", "custom"}:
            raise HTTPException(400, f"billing_frequency non valido: {bf}")
        q.billing_frequency = bf
    if billing_terms_days is not None:
        q.billing_terms_days = (
            int(billing_terms_days) if billing_terms_days else None
        )
    # v3.5.0-alpha.137 — Cambio valuta quote + refresh tasso on-demand.
    # Rifiuta cambio se quote NON in draft (immutabilità imponibile post-emissione).
    want_refresh = (refresh_fx or "").strip().lower() == "true"
    new_ccy = (currency or "").upper().strip() if currency else None
    if (new_ccy and new_ccy != (q.currency or "EUR")) or want_refresh:
        if q.status != QuoteStatus.draft:
            raise HTTPException(
                409,
                f"Valuta o tasso modificabili solo in stato draft (attuale: {q.status.value}). "
                "Per nuovo tasso: clona la quote in draft."
            )
        from app.models import Tenant
        from app.services.fx import get_fx_rate, refresh_fx_rate
        from datetime import datetime as _dt
        tenant = db.query(Tenant).filter(Tenant.id == current_tenant_id()).first()
        base_ccy = (tenant.default_currency if tenant else "EUR").upper()
        target_ccy = new_ccy or q.currency or base_ccy
        if len(target_ccy) != 3:
            raise HTTPException(400, f"Valuta non valida: {target_ccy}")
        if target_ccy == base_ccy:
            q.currency = target_ccy
            q.fx_rate_to_base = 1.0
            q.fx_rate_fixed_at = None
        else:
            rate = (refresh_fx_rate(db, target_ccy, base_ccy) if want_refresh
                    else get_fx_rate(db, target_ccy, base_ccy))
            if rate is None:
                raise HTTPException(422, "Tasso di cambio non disponibile, riprova più tardi")
            q.currency = target_ccy
            q.fx_rate_to_base = rate
            q.fx_rate_fixed_at = now_utc()
    _recalc_quote(q)
    db.commit()
    return {
        "id": q.id,
        "number": q.number,
        "title": q.title,
        "subtotal_gross": q.subtotal_gross,
        "subtotal": q.subtotal,
        "total_after_discount": q.total_after_discount,
        "total_with_vat": q.total_with_vat,
        # v3.5.0-alpha.137 — restituisce valuta + tasso aggiornato per UI
        "currency": q.currency or "EUR",
        "fx_rate_to_base": q.fx_rate_to_base or 1.0,
        "fx_rate_fixed_at": q.fx_rate_fixed_at.isoformat() if q.fx_rate_fixed_at else None,
    }


# ── Advance schedule (v3.5.0-alpha.139) ────────────────────────────
# Termini di acconto definiti in quote. Sostituisce il pattern "fattura manuale"
# di α.136-138: ora gli acconti nascono qui, strutturati con scadenze e pct,
# poi al converti quote→job (α.140) verranno auto-creati AdvancePayment pending.


def _get_schedules_serialized(db: Session, quote_id: int) -> list:
    from app.models import QuoteAdvanceSchedule
    rows = (
        db.query(QuoteAdvanceSchedule)
        .options(joinedload(QuoteAdvanceSchedule.allocations))
        .filter(QuoteAdvanceSchedule.quote_id == quote_id)
        .order_by(QuoteAdvanceSchedule.sort_order.asc(), QuoteAdvanceSchedule.id.asc())
        .all()
    )
    return [_serialize_schedule(s) for s in rows]


def _serialize_schedule(s):
    return {
        "id": s.id,
        "label": s.label,
        "pct": s.pct,
        "amount_fixed": s.amount_fixed,
        "due_anchor": s.due_anchor.value if hasattr(s.due_anchor, "value") else s.due_anchor,
        "due_offset_days": s.due_offset_days,
        "due_date": str(s.due_date) if s.due_date else None,
        "milestone_label": s.milestone_label,
        "sort_order": s.sort_order,
        "notes": s.notes,
        "allocations": [
            {"id": a.id, "quote_line_id": a.quote_line_id, "pct": a.pct}
            for a in (s.allocations or [])
        ],
    }


@router.get("/api/{quote_id}/advance-schedules")
async def list_advance_schedules(quote_id: int, db: Session = Depends(get_db)):
    from app.models import QuoteAdvanceSchedule
    q = db.query(Quote).filter(
        Quote.id == quote_id, Quote.tenant_id == current_tenant_id(),
    ).first()
    if not q:
        raise HTTPException(404, "Quote non trovata")
    schedules = (
        db.query(QuoteAdvanceSchedule)
        .options(joinedload(QuoteAdvanceSchedule.allocations))
        .filter(QuoteAdvanceSchedule.quote_id == quote_id)
        .order_by(QuoteAdvanceSchedule.sort_order.asc(), QuoteAdvanceSchedule.id.asc())
        .all()
    )
    return {"quote_id": quote_id, "schedules": [_serialize_schedule(s) for s in schedules]}


# v3.5.0-alpha.172.47 — HARD-BLOCK helper: Σ pct schedules <= 100%.
# Pre-α.172.47 utente poteva creare schedule 30% + 80% = 110% senza warn.
# Materialize → 2 AP per 110% del budget → bloccato solo a emit.
# Check anticipato al save dello schedule.
def _check_advance_schedule_total(
    db: Session, quote_id: int, *,
    new_pct: Optional[float] = None,
    new_amount_fixed: Optional[float] = None,
    exclude_id: Optional[int] = None,
) -> None:
    """Solleva HTTPException(409) se Σ schedule pct (+ amount_fixed
    convertito a pct via quote_total) eccede 1.0. `exclude_id` = id del
    record corrente (PUT) da escludere dal calcolo."""
    from app.models import QuoteAdvanceSchedule
    q = db.query(Quote).filter(Quote.id == quote_id).first()
    if not q:
        return  # quote non esiste, lascia gestire al chiamante
    quote_total = q.total_after_discount or 0.0
    existing = db.query(QuoteAdvanceSchedule).filter(
        QuoteAdvanceSchedule.quote_id == quote_id,
        QuoteAdvanceSchedule.tenant_id == current_tenant_id(),
    )
    if exclude_id is not None:
        existing = existing.filter(QuoteAdvanceSchedule.id != exclude_id)
    cumulative_pct = 0.0
    for s in existing.all():
        if s.pct:
            cumulative_pct += s.pct
        elif s.amount_fixed and quote_total > 0:
            cumulative_pct += s.amount_fixed / quote_total
    new_contrib = 0.0
    if new_pct:
        new_contrib = new_pct
    elif new_amount_fixed and quote_total > 0:
        new_contrib = new_amount_fixed / quote_total
    if cumulative_pct + new_contrib > 1.0 + 0.001:
        existing_pct_label = round(cumulative_pct * 100, 1)
        new_pct_label = round(new_contrib * 100, 1)
        total_pct_label = round((cumulative_pct + new_contrib) * 100, 1)
        raise HTTPException(
            409,
            detail={
                "message": (
                    f"Gli acconti programmati supererebbero il 100% del valore della quotazione.\n\n"
                    f"• Acconti già programmati: {existing_pct_label}%\n"
                    f"• Nuovo acconto richiesto: {new_pct_label}%\n"
                    f"• Totale risultante: {total_pct_label}%\n\n"
                    f"Per procedere: riduci la percentuale del nuovo acconto, "
                    f"oppure elimina uno degli acconti programmati esistenti."
                ),
                "existing_pct": existing_pct_label,
                "attempted_pct": new_pct_label,
                "total_pct": total_pct_label,
            },
        )


@router.post("/api/{quote_id}/advance-schedules", dependencies=[RequireEditQuotes])
async def create_advance_schedule(
    quote_id: int,
    label: str = Form(...),
    pct: Optional[float] = Form(None),
    amount_fixed: Optional[float] = Form(None),
    due_anchor: str = Form("quote_approved"),
    due_offset_days: int = Form(0),
    due_date: Optional[date] = Form(None),
    milestone_label: Optional[str] = Form(None),
    sort_order: int = Form(0),
    notes: Optional[str] = Form(None),
    # CSV "line_id:pct,line_id:pct" — allocazione opzionale a QuoteLine
    allocations: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    from app.models import QuoteAdvanceSchedule, QuoteAdvanceAllocation, AdvanceDueAnchor
    q = db.query(Quote).filter(
        Quote.id == quote_id, Quote.tenant_id == current_tenant_id(),
    ).first()
    if not q:
        raise HTTPException(404, "Quote non trovata")
    # v3.5.0-alpha.171.11 — Companion=0 dal client significa "non usato": normalizza a None.
    # UI radio invia sempre entrambi (uno con valore, l'altro a 0) per coerenza form.
    if pct is not None and pct <= 0:
        pct = None
    if amount_fixed is not None and amount_fixed <= 0:
        amount_fixed = None
    if pct is None and amount_fixed is None:
        raise HTTPException(400, "Specificare pct o amount_fixed")
    # v3.5.0-alpha.166 — Mutual exclusion. Pre-α.166: entrambi valorizzati →
    # amount_fixed prevaleva silenziosamente in _compute_amount, semantica
    # confusa. Ora rifiuta esplicito.
    if pct is not None and pct > 0 and amount_fixed is not None and amount_fixed > 0:
        raise HTTPException(
            400,
            "pct e amount_fixed mutualmente esclusivi: specifica uno o l'altro, non entrambi",
        )
    if pct is not None and (pct < 0 or pct > 1.0):
        raise HTTPException(400, "pct deve essere tra 0 e 1.0 (es. 0.30 = 30%)")
    if amount_fixed is not None and amount_fixed < 0:
        raise HTTPException(400, "amount_fixed deve essere >= 0")
    try:
        anchor_enum = AdvanceDueAnchor(due_anchor)
    except ValueError:
        raise HTTPException(400, f"due_anchor non valido: {due_anchor}")
    # v3.5.0-alpha.172.47 — HARD-BLOCK overflow Σ pct
    _check_advance_schedule_total(db, quote_id, new_pct=pct, new_amount_fixed=amount_fixed)
    sched = QuoteAdvanceSchedule(
        tenant_id=current_tenant_id(),
        quote_id=quote_id,
        label=label.strip(),
        pct=pct, amount_fixed=amount_fixed,
        due_anchor=anchor_enum, due_offset_days=due_offset_days,
        due_date=due_date, milestone_label=milestone_label,
        sort_order=sort_order, notes=notes,
    )
    db.add(sched)
    db.flush()
    # Parse e crea allocations opzionali (formato "id:pct,id:pct")
    if allocations:
        for token in allocations.split(","):
            token = token.strip()
            if not token:
                continue
            if ":" not in token:
                raise HTTPException(400, f"allocations formato invalido: '{token}' (atteso line_id:pct)")
            l_id_s, pct_s = token.split(":", 1)
            try:
                l_id = int(l_id_s.strip())
                a_pct = float(pct_s.strip())
            except ValueError:
                raise HTTPException(400, f"allocations parse error: '{token}'")
            if a_pct <= 0 or a_pct > 1.0:
                raise HTTPException(400, f"allocation pct deve essere tra 0 e 1.0 (line #{l_id})")
            alloc = QuoteAdvanceAllocation(
                schedule_id=sched.id, quote_line_id=l_id, pct=a_pct,
            )
            db.add(alloc)
    db.commit()
    db.refresh(sched)
    return _serialize_schedule(sched)


@router.put("/api/advance-schedules/{schedule_id}", dependencies=[RequireEditQuotes])
async def update_advance_schedule(
    schedule_id: int,
    label: Optional[str] = Form(None),
    pct: Optional[float] = Form(None),
    amount_fixed: Optional[float] = Form(None),
    due_anchor: Optional[str] = Form(None),
    due_offset_days: Optional[int] = Form(None),
    due_date: Optional[date] = Form(None),
    milestone_label: Optional[str] = Form(None),
    sort_order: Optional[int] = Form(None),
    notes: Optional[str] = Form(None),
    # v3.5.0-alpha.172.3 Bug 1 fix — allocations form param accettato anche
    # in update path. Pre-fix (pre α.172.3) il PUT NON accettava allocations
    # quindi la modifica delle voci coperte dall'acconto post-creazione era
    # impossibile via UI. Vedi memory project_bug_acconti_2026_05_20.
    # Formato CSV "line_id:pct,line_id:pct" — replace TUTTE le allocations
    # esistenti (delete + insert). Stringa vuota = clear allocations.
    allocations: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    from app.models import (
        QuoteAdvanceSchedule, QuoteAdvanceAllocation, AdvanceDueAnchor,
        AdvancePayment, AdvancePaymentAllocation,
        AdvancePaymentDeliverableAllocation,
    )
    s = db.query(QuoteAdvanceSchedule).filter(
        QuoteAdvanceSchedule.id == schedule_id,
        QuoteAdvanceSchedule.tenant_id == current_tenant_id(),
    ).first()
    if not s:
        raise HTTPException(404, "Schedule non trovato")
    if label is not None: s.label = label.strip()
    # v3.5.0-alpha.172.47 — HARD-BLOCK overflow Σ pct (escludi self da calc)
    _check_advance_schedule_total(
        db, s.quote_id, new_pct=pct, new_amount_fixed=amount_fixed,
        exclude_id=schedule_id,
    )
    # v3.5.0-alpha.166 — Mutual exclusion: settare pct azzera amount_fixed e viceversa.
    if pct is not None:
        if pct < 0 or pct > 1.0:
            raise HTTPException(400, "pct deve essere tra 0 e 1.0")
        s.pct = pct
        if pct > 0:
            s.amount_fixed = None
    if amount_fixed is not None:
        if amount_fixed < 0:
            raise HTTPException(400, "amount_fixed deve essere >= 0")
        s.amount_fixed = amount_fixed
        if amount_fixed > 0:
            s.pct = None
    if due_anchor is not None:
        try:
            s.due_anchor = AdvanceDueAnchor(due_anchor)
        except ValueError:
            raise HTTPException(400, f"due_anchor non valido: {due_anchor}")
    if due_offset_days is not None: s.due_offset_days = due_offset_days
    if due_date is not None: s.due_date = due_date
    if milestone_label is not None: s.milestone_label = milestone_label
    if sort_order is not None: s.sort_order = sort_order
    if notes is not None: s.notes = notes

    # v3.5.0-alpha.172.3 — Bug 1 fix: replace allocations + re-materialize AP.
    if allocations is not None:
        # Delete existing allocations
        db.query(QuoteAdvanceAllocation).filter(
            QuoteAdvanceAllocation.schedule_id == s.id
        ).delete(synchronize_session=False)
        # Parse + insert new ones (formato "line_id:pct,line_id:pct")
        if allocations.strip():
            for token in allocations.split(","):
                token = token.strip()
                if not token:
                    continue
                if ":" not in token:
                    raise HTTPException(400, f"allocations formato invalido: '{token}' (atteso line_id:pct)")
                l_id_s, pct_s = token.split(":", 1)
                try:
                    l_id = int(l_id_s.strip())
                    a_pct = float(pct_s.strip())
                except ValueError:
                    raise HTTPException(400, f"allocations parse error: '{token}'")
                if a_pct <= 0 or a_pct > 1.0:
                    raise HTTPException(400, f"allocation pct deve essere tra 0 e 1.0 (line #{l_id})")
                db.add(QuoteAdvanceAllocation(
                    schedule_id=s.id, quote_line_id=l_id, pct=a_pct,
                ))

        # Re-materialize: se AP già esistente (status pending/draft/confirmed)
        # cancella sue allocations vecchie e re-crea da nuova schedule
        # allocations. AP già invoiced/paid/consumed NON ricreate (semantica
        # immutability: la riallocazione richiede manualmente nota credito).
        from app.models import AdvancePaymentStatus
        editable_statuses = {AdvancePaymentStatus.pending, AdvancePaymentStatus.draft,
                              AdvancePaymentStatus.confirmed}
        existing_aps = db.query(AdvancePayment).filter(
            AdvancePayment.quote_advance_schedule_id == s.id,
            AdvancePayment.status.in_(editable_statuses),
        ).all()
        re_materialized = 0
        for ap in existing_aps:
            # Cancella vecchie AP allocations (JCL + Deliverable)
            db.query(AdvancePaymentAllocation).filter(
                AdvancePaymentAllocation.advance_payment_id == ap.id
            ).delete(synchronize_session=False)
            db.query(AdvancePaymentDeliverableAllocation).filter(
                AdvancePaymentDeliverableAllocation.advance_payment_id == ap.id
            ).delete(synchronize_session=False)
            # Re-materialize via helper riusabile
            from app.services.advance_schedule_to_payment import (
                rebuild_ap_allocations_from_schedule,
            )
            rebuild_ap_allocations_from_schedule(db, ap, s)
            re_materialized += 1
        s._re_materialized_count = re_materialized  # debug payload only

    db.commit()
    db.refresh(s)
    result = _serialize_schedule(s)
    result["re_materialized_count"] = getattr(s, "_re_materialized_count", 0)
    return result


@router.delete("/api/advance-schedules/{schedule_id}", dependencies=[RequireEditQuotes])
async def delete_advance_schedule(schedule_id: int, db: Session = Depends(get_db)):
    from app.models import QuoteAdvanceSchedule
    s = db.query(QuoteAdvanceSchedule).filter(
        QuoteAdvanceSchedule.id == schedule_id,
        QuoteAdvanceSchedule.tenant_id == current_tenant_id(),
    ).first()
    if not s:
        raise HTTPException(404, "Schedule non trovato")
    db.delete(s)  # cascade su allocations
    db.commit()
    return {"id": schedule_id, "deleted": True}


@router.put("/api/{quote_id}/category-discount", dependencies=[RequireEditQuotes])
async def update_category_discount(
    quote_id: int,
    category: str = Form(...),
    pct: float = Form(...),
    db: Session = Depends(get_db),
):
    """Imposta lo sconto su una categoria (positivo = riduzione, es. 0.15 = 15%).
    Passare pct=0 per rimuovere lo sconto."""
    q = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
    ).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404, "Quotazione non trovata")
    discounts = dict(q.category_discounts or {})
    if pct == 0:
        discounts.pop(category, None)
    else:
        discounts[category] = pct
    q.category_discounts = discounts
    _recalc_quote(q)
    db.commit()
    return {
        "category_discounts": q.category_discounts or {},
        "category_order": q.category_order or [],
        "subtotal_gross": q.subtotal_gross,
        "subtotal_gross_jcl": getattr(q, "subtotal_gross_jcl", 0.0) or 0.0,
        "subtotal_gross_deliverable": getattr(q, "subtotal_gross_deliverable", 0.0) or 0.0,
        "subtotal": q.subtotal,
        "total_after_discount": q.total_after_discount,
        "total_with_vat": q.total_with_vat,
    }


@router.put("/api/{quote_id}/lines-reorder", dependencies=[RequireEditQuotes])
async def reorder_quote_lines(
    quote_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Riordina le voci. Body JSON: {order: [line_id1, line_id2, ...]}"""
    data = await request.json()
    order = data.get("order", [])
    if not isinstance(order, list):
        raise HTTPException(400, "Campo 'order' deve essere una lista di line_id")
    q = db.query(Quote).options(joinedload(Quote.lines)).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404)
    _assert_quote_mutable(q, action="riordino voci")
    line_map = {l.id: l for l in q.lines}
    for idx, line_id in enumerate(order):
        line = line_map.get(int(line_id))
        if line:
            line.sort_order = (idx + 1) * 10
    db.commit()
    return {"ok": True, "count": len(order)}


@router.put("/api/{quote_id}/category-order", dependencies=[RequireEditQuotes])
async def reorder_quote_categories(
    quote_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """v3.4.34 — Persiste l'ordine delle categorie nelle voci preventivo.
    Body JSON: {order: ["PICTURE", "SOUND", "Altro"]}.
    Le categorie non listate appaiono dopo nell'ordine naturale."""
    data = await request.json()
    order = data.get("order", [])
    if not isinstance(order, list):
        raise HTTPException(400, "Campo 'order' deve essere una lista di nomi categoria")
    q = db.query(Quote).filter(Quote.id == quote_id).first()
    if not q:
        raise HTTPException(404, "Quotazione non trovata")
    cleaned = [str(c) for c in order if c]
    q.category_order = cleaned or None
    db.commit()
    return {"ok": True, "category_order": q.category_order or []}


@router.post("/api/{quote_id}/lines", dependencies=[RequireEditQuotes])
async def add_quote_line(
    quote_id: int,
    description: str = Form(...),
    section: str = Form("A"),
    position: str = Form("A.1"),
    detail: Optional[str] = Form(None),
    quantity: float = Form(1.0),
    unit: str = Form("day"),
    price_level: PriceLevel = Form(PriceLevel.list_price),
    unit_price: float = Form(0.0),
    allowance: float = Form(0.0),
    line_discount_pct: float = Form(0.0),
    hardcosts: float = Form(0.0),
    price_item_id: Optional[int] = Form(None),
    category_override: Optional[str] = Form(None),
    is_optional: bool = Form(False),
    section_label: Optional[str] = Form(None),
    # Segnale esplicito: il frontend deve inviare from_listino=1 quando aggiunge
    # una voce dal picker del listino con prezzo già prefillato. Evita doppia
    # conversione valuta se unit_price è non-zero ma proviene già dalla base.
    from_listino: int = Form(0),
    db: Session = Depends(get_db),
):
    q = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
    ).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404)
    _assert_quote_mutable(q, action="aggiunta voce")
    # v3.5.0-alpha.172.146 — eredita dalla voce di listino sia il prezzo (se 0)
    # sia il DETTAGLIO (se vuoto): prima il campo detail restava sempre vuoto
    # aggiungendo una voce dal listino.
    #
    # True se prezzo proviene dal listino (già in valuta base, NON convertire).
    # Segnale esplicito dal frontend (from_listino=1) OPPURE euristica legacy
    # (price_item_id presente + prezzo non digitato a mano, cioè unit_price==0).
    price_from_listino = bool(from_listino) or (price_item_id is not None and unit_price == 0)
    if price_item_id:
        item = db.query(PriceItem).filter(PriceItem.id == price_item_id).first()
        if item:
            if unit_price == 0:
                unit_price = _resolve_item_unit_price(item, price_level)
                # prezzo risolto dal listino → già in base, non convertire
            if not (detail or "").strip() and (item.description or "").strip():
                detail = item.description.strip()
    unit_price = _line_price_to_base(db, q, entered_price=unit_price, from_price_item=price_from_listino)
    sort_order = max((l.sort_order for l in q.lines), default=0) + 10
    cat_override_clean = (category_override or "").strip() or None
    section_label_clean = (section_label or "").strip() or None
    line = QuoteLine(
        quote_id=quote_id, description=description, section=section,
        position=position, detail=detail, quantity=quantity, unit=unit,
        price_level=price_level, unit_price=unit_price,
        allowance=allowance, line_discount_pct=line_discount_pct,
        total=0.0, hardcosts=hardcosts,
        price_item_id=price_item_id, sort_order=sort_order,
        category_override=cat_override_clean,
        is_optional=bool(is_optional),
        section_label=section_label_clean,
    )
    db.add(line); db.flush()
    db.refresh(q)
    _recalc_quote(q)

    # v3.4.36 (R1.2): se la quote ha già un job approvato (post-conversione),
    # auto-crea JobCostLine corrispondente per mantenere job e quote allineati.
    # Idempotente: skip se esiste già JobCostLine con quote_line_id=line.id.
    job_cost_line_created = False
    if q.job and q.job.status not in (JobStatus.completed, JobStatus.invoiced, JobStatus.cancelled):
        existing = db.query(JobCostLine).filter(
            JobCostLine.quote_line_id == line.id
        ).first()
        if not existing:
            db.add(JobCostLine(
                job_id=q.job.id,
                quote_line_id=line.id,
                price_item_id=line.price_item_id,
                description=line.description,
                quantity_quoted=line.quantity,
                quantity_actual=0.0,
                unit=line.unit,
                unit_price=line.unit_price,
                total_quoted=line.total or 0.0,
                total_accrued=0.0,
                total_expected=line.total or 0.0,
                is_billable=True,
                is_extra=False,
            ))
            job_cost_line_created = True

    db.commit()
    return {
        "id": line.id, "total": line.total, "quote_total": q.total_with_vat,
        "subtotal_gross": q.subtotal_gross,
        "subtotal_gross_jcl": getattr(q, "subtotal_gross_jcl", 0.0) or 0.0,
        "subtotal_gross_deliverable": getattr(q, "subtotal_gross_deliverable", 0.0) or 0.0,
        "subtotal": q.subtotal,
        "total_after_discount": q.total_after_discount,
        "total_with_vat": q.total_with_vat,
        "subtotal_optional": round(
            sum((l.total or 0.0) for l in q.lines if l.is_optional), 2
        ),
        "job_cost_line_created": job_cost_line_created,
    }


@router.post("/api/{quote_id}/load-from-template", dependencies=[RequireEditQuotes])
async def load_from_template(
    quote_id: int,
    template_id: int = Form(...),
    price_level: PriceLevel = Form(PriceLevel.list_price),
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.68.6 — Bulk-add quote_lines da DeliveryTemplate.suggested_items.

    Per ogni item del template:
      - skip se price_item mancante (no auto-create)
      - skip se già esiste riga con stesso price_item_id (no duplicati)
      - aggiunge QuoteLine con quantity=qty_hint, unit/price ereditati dal listino
      - section dal template (A/B/C), sort_order incrementale

    Idempotente: ri-eseguibile, aggiunge solo le righe mancanti."""
    q = db.query(Quote).options(
        joinedload(Quote.lines)
    ).filter(
        Quote.id == quote_id,
        Quote.tenant_id == current_tenant_id(),
    ).first()
    if not q:
        raise HTTPException(404, "Quote non trovata")
    if q.status in (QuoteStatus.approved, QuoteStatus.rejected):
        raise HTTPException(409, f"Quote in stato {q.status.value}, non modificabile")

    t = db.query(DeliveryTemplate).filter(
        DeliveryTemplate.id == template_id,
        DeliveryTemplate.tenant_id == current_tenant_id(),
    ).first()
    if not t:
        raise HTTPException(404, "Template non trovato")
    items = t.suggested_items or []
    if not items:
        raise HTTPException(400, "Il template non ha suggested_items configurate")

    existing_pi = {l.price_item_id for l in q.lines if l.price_item_id}
    sort_order = max((l.sort_order for l in q.lines), default=0)
    added = 0
    skipped_dup = 0
    skipped_missing = 0
    section_counters: dict[str, int] = {}
    for it in items:
        pid = it.get("price_item_id")
        if not pid:
            skipped_missing += 1
            continue
        if pid in existing_pi:
            skipped_dup += 1
            continue
        item = db.query(PriceItem).filter(
            PriceItem.id == int(pid),
            PriceItem.tenant_id == current_tenant_id(),
            PriceItem.is_active == True,  # noqa: E712
        ).first()
        if not item:
            skipped_missing += 1
            continue
        price = _resolve_item_unit_price(item, price_level)
        section = (it.get("section") or "A").strip().upper()[:1]
        section_counters[section] = section_counters.get(section, 0) + 1
        position = f"{section}.{section_counters[section]}"
        sort_order += 10
        qty = float(it.get("qty_hint") or 1)
        line = QuoteLine(
            quote_id=quote_id,
            description=item.name,
            section=section,
            position=position,
            detail=(it.get("notes") or None),
            quantity=qty,
            unit=item.unit,
            price_level=price_level,
            unit_price=price,
            allowance=0.0,
            line_discount_pct=0.0,
            total=0.0,
            hardcosts=0.0,
            price_item_id=item.id,
            sort_order=sort_order,
            is_optional=False,
        )
        db.add(line)
        added += 1
        existing_pi.add(item.id)

    if added == 0:
        return {
            "ok": True,
            "added": 0,
            "skipped_duplicate": skipped_dup,
            "skipped_missing": skipped_missing,
            "message": "Nessuna riga aggiunta (già presenti o price_item mancanti)",
        }
    db.flush()
    db.refresh(q)
    _recalc_quote(q)
    db.commit()
    return {
        "ok": True,
        "added": added,
        "skipped_duplicate": skipped_dup,
        "skipped_missing": skipped_missing,
        "quote_total": q.total_with_vat,
        "subtotal": q.subtotal,
        "template_code": t.code,
        "template_name": t.name,
    }


@router.get("/api/template-buckets/{template_id}", dependencies=[RequireEditQuotes])
async def template_buckets(template_id: int, db: Session = Depends(get_db)):
    """F2 picker — voci-bucket distinte derivate dai DeliveryItem del template
    (decisione 10). Sorgente per il picker a spunte nella quote."""
    t = db.query(DeliveryTemplate).filter(
        DeliveryTemplate.id == template_id,
        DeliveryTemplate.tenant_id == current_tenant_id(),
    ).first()
    if not t:
        raise HTTPException(404, "Template non trovato")
    buckets = template_bucket_options(db, current_tenant_id(), template_id)
    return {
        "ok": True,
        "template_id": t.id,
        "template_code": t.code,
        "template_name": t.name,
        "buckets": buckets,
    }


@router.post("/api/{quote_id}/load-from-template-items", dependencies=[RequireEditQuotes])
async def load_from_template_items(
    quote_id: int,
    template_id: int = Form(...),
    price_item_ids: List[int] = Form(...),
    price_level: PriceLevel = Form(PriceLevel.list_price),
    section: str = Form("A"),
    with_detail: bool = Form(True),
    db: Session = Depends(get_db),
):
    """F2 — aggiunge righe quote dai bucket selezionati nel picker (decisione 2+10).

    A differenza di ``load-from-template`` (statico da ``suggested_items``), le voci
    sono il sottoinsieme spuntato dei bucket derivati dai DeliveryItem del template.
    Se ``with_detail``, precompila ``detail`` con le note di capitolato aggregate
    per quel bucket (piattaforma upload, naming, ecc.). Idempotente sui price_item.
    """
    q = db.query(Quote).options(joinedload(Quote.lines)).filter(
        Quote.id == quote_id,
        Quote.tenant_id == current_tenant_id(),
    ).first()
    if not q:
        raise HTTPException(404, "Quote non trovata")
    if q.status in (QuoteStatus.approved, QuoteStatus.rejected):
        raise HTTPException(409, f"Quote in stato {q.status.value}, non modificabile")

    # Mappa bucket→detail derivata dal template (validazione: solo bucket di QUESTO template).
    options = {o["price_item_id"]: o for o in template_bucket_options(db, current_tenant_id(), template_id)}
    if not options:
        raise HTTPException(400, "Il template non ha voci-bucket derivabili (DeliveryItem non linkati)")

    # v3.5.0-alpha.172.146 — etichetta automatica del capitolato sulle deliveries.
    # Le consegne di broadcaster diversi (es. Sky vs NBCU) hanno specs diverse
    # (LUFS, livelli, timeline, ordine loghi): la section_label = broadcaster del
    # capitolato evita di confonderle. Fallback al nome template.
    t = db.query(DeliveryTemplate).filter(
        DeliveryTemplate.id == template_id,
        DeliveryTemplate.tenant_id == current_tenant_id(),
    ).first()
    cap_label = (((t.broadcaster if t and t.broadcaster else (t.name if t else None)) or "").strip()) or None

    # Dedup per (voce + capitolato): la stessa voce-bucket dallo STESSO capitolato
    # non viene duplicata; ma la stessa voce da un capitolato DIVERSO viene
    # aggiunta (con la sua etichetta). Prima il dedup era solo su price_item_id →
    # una voce già nel listino/quote non veniva mai aggiunta da un altro capitolato.
    existing_keys = {(l.price_item_id, l.section_label) for l in q.lines if l.price_item_id}
    sort_order = max((l.sort_order for l in q.lines), default=0)
    sect = (section or "A").strip().upper()[:1] or "A"
    sect_count = sum(1 for l in q.lines if l.section == sect)
    added, skipped_dup, skipped_invalid = 0, 0, 0

    for pid in price_item_ids:
        opt = options.get(pid)
        if not opt:
            skipped_invalid += 1
            continue
        if (pid, cap_label) in existing_keys:
            skipped_dup += 1
            continue
        item = db.query(PriceItem).filter(
            PriceItem.id == pid,
            PriceItem.tenant_id == current_tenant_id(),
            PriceItem.is_active == True,  # noqa: E712
        ).first()
        if not item:
            skipped_invalid += 1
            continue
        price = _resolve_item_unit_price(item, price_level)
        sect_count += 1
        sort_order += 10
        line = QuoteLine(
            quote_id=quote_id,
            description=item.name,
            section=sect,
            position=f"{sect}.{sect_count}",
            detail=(opt.get("detail_suggestion") if with_detail else None),
            quantity=1.0,
            unit=item.unit,
            price_level=price_level,
            unit_price=price,
            allowance=0.0,
            line_discount_pct=0.0,
            total=0.0,
            hardcosts=0.0,
            price_item_id=item.id,
            sort_order=sort_order,
            is_optional=False,
            section_label=cap_label,
        )
        db.add(line)
        existing_keys.add((item.id, cap_label))
        added += 1

    if added == 0:
        return {"ok": True, "added": 0, "skipped_duplicate": skipped_dup,
                "skipped_invalid": skipped_invalid,
                "message": "Nessuna riga aggiunta (già presenti o non valide)"}
    db.flush()
    db.refresh(q)
    _recalc_quote(q)
    db.commit()
    return {
        "ok": True, "added": added, "skipped_duplicate": skipped_dup,
        "skipped_invalid": skipped_invalid,
        "quote_total": q.total_with_vat, "subtotal": q.subtotal,
    }


@router.put("/api/{quote_id}/lines/{line_id}", dependencies=[RequireEditQuotes])
async def update_quote_line(
    quote_id: int, line_id: int,
    description: Optional[str] = Form(None),
    detail: Optional[str] = Form(None),
    quantity: Optional[float] = Form(None),
    unit: Optional[str] = Form(None),
    unit_price: Optional[float] = Form(None),
    allowance: Optional[float] = Form(None),
    line_discount_pct: Optional[float] = Form(None),
    category_override: Optional[str] = Form(None),
    is_optional: Optional[bool] = Form(None),
    section_label: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    line = db.query(QuoteLine).filter(QuoteLine.id == line_id, QuoteLine.quote_id == quote_id).first()
    if not line: raise HTTPException(404)
    # v3.5.0-alpha.172.18 — HARD-BLOCK su quote approvate non-phantom (Consuntivi
    # sono modificabili perché sono il cost report vivo del progetto).
    q_pre = db.query(Quote).filter(Quote.id == quote_id).first()
    if q_pre:
        _assert_quote_mutable(q_pre, action="modifica voce")
    if description is not None: line.description = description
    if detail is not None: line.detail = detail
    if quantity is not None: line.quantity = quantity
    if unit is not None: line.unit = unit
    if unit_price is not None:
        line.unit_price = _line_price_to_base(db, q_pre, entered_price=unit_price, from_price_item=False)
    if allowance is not None: line.allowance = allowance
    if line_discount_pct is not None: line.line_discount_pct = line_discount_pct
    if category_override is not None:
        cleaned = category_override.strip()
        # FastAPI Form(None) parsa "" come None, quindi usiamo un sentinel
        # esplicito per cancellare l'override.
        if cleaned in ("__CLEAR__", "__none__"):
            line.category_override = None
        else:
            line.category_override = cleaned or None
    if is_optional is not None:
        line.is_optional = bool(is_optional)
    if section_label is not None:
        cleaned = section_label.strip()
        if cleaned in ("__CLEAR__", "__none__"):
            line.section_label = None
        else:
            line.section_label = cleaned or None
    q = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
    ).filter(Quote.id == quote_id).first()
    _recalc_quote(q)

    # v3.4.36 (R1.3): sync verso JobCostLine collegata, se presente.
    # Blocca se job in stato terminale (completed/invoiced/cancelled).
    job_cost_line_synced = False
    jcl = db.query(JobCostLine).filter(JobCostLine.quote_line_id == line.id).first()
    if jcl:
        if jcl.job and jcl.job.status in (JobStatus.completed, JobStatus.invoiced, JobStatus.cancelled):
            raise HTTPException(
                409,
                f"Modifica bloccata: il job {jcl.job.code} è in stato "
                f"{jcl.job.status.value}. Riapri/duplica il job per modificare."
            )
        # v3.5.0-alpha.172.36 (Sprint 2 BLOCCO 2) — slice-lock guard JCL-driven.
        # Se la JCL ha già una slice attiva (fatturata), bloccare la mutazione:
        # corromperebbe link JCL↔fattura emessa (snapshot resta corretto ma JCL
        # diverge per future ri-trasmissioni). Soluzione formale: TD04 + nuova
        # versione di quote.
        assert_jcl_lock_safe(db, jcl.id, action="modificare la voce di costo collegata a")
        jcl.description = line.description
        jcl.quantity_quoted = line.quantity
        jcl.unit = line.unit
        jcl.unit_price = line.unit_price
        jcl.total_quoted = line.total or 0.0
        # total_expected si aggiorna se non è ancora stato sovrascritto manualmente
        # (heuristica: se total_expected == previous total_quoted, segue).
        # Per sicurezza: lo lasciamo come riferimento iniziale (no sovrascrittura).
        job_cost_line_synced = True

    # v3.5.0-alpha.172.99 — sync JobDeliverable + respawn cross-nature.
    # Step 1: prova respawn (cambio nature OR cambio N row spawn-per-unit).
    # Pre-check safety: solleva 409 se ci sono consegne/booking link/billed.
    # Step 2: se no-respawn, sync inline name/unit/price/qty.
    job_deliverables_synced = 0
    respawn_info = {}
    parent_job = db.query(Job).filter(Job.quote_id == quote_id).first()
    if parent_job:
        respawn_info = _respawn_line_artifacts(db, line, parent_job)
    if not respawn_info.get("respawned"):
        # Sync inline (no nature/rows change)
        from app.services.cost_line_sync import unit_nature_for as _unfor
        deliverables = db.query(JobDeliverable).filter(
            JobDeliverable.quote_line_id == line.id
        ).all()
        for d in deliverables:
            d.name = line.description
            d.unit = line.unit
            d.unit_price = line.unit_price or 0.0
            if (d.quantity_delivered or 0.0) == 0.0:
                # Per `deliverable_qty` spawn-per-unit: qty_planned=1.0 per row,
                # già coerente (count rows == line.quantity gestito dal respawn).
                # Per volume/manual_allow (1 row aggregato): aggiorna a qty quote.
                if _unfor(line.unit) != "deliverable_qty":
                    d.quantity_planned = float(line.quantity or 0.0)
            d.total_quoted = round((d.quantity_planned or 0.0) * (d.unit_price or 0.0), 2)
            job_deliverables_synced += 1
    else:
        job_deliverables_synced = respawn_info.get("new_n_rows", 0)

    db.commit()
    return {
        "id": line.id, "total": line.total,
        "subtotal_gross": q.subtotal_gross,
        "subtotal_gross_jcl": getattr(q, "subtotal_gross_jcl", 0.0) or 0.0,
        "subtotal_gross_deliverable": getattr(q, "subtotal_gross_deliverable", 0.0) or 0.0,
        "subtotal": q.subtotal,
        "total_after_discount": q.total_after_discount,
        "total_with_vat": q.total_with_vat,
        "subtotal_optional": round(
            sum((l.total or 0.0) for l in q.lines if l.is_optional), 2
        ),
        "is_optional": bool(line.is_optional),
        "section_label": line.section_label or None,
        "job_cost_line_synced": job_cost_line_synced,
        "job_deliverables_synced": job_deliverables_synced,
    }


@router.delete("/api/{quote_id}/lines/{line_id}", dependencies=[RequireEditQuotes])
async def delete_quote_line(quote_id: int, line_id: int, db: Session = Depends(get_db)):
    """v3.4.55 — cancellazione QuoteLine: HARD-BLOCK se ha JobCostLine con
    booking attivi (status != cancelled). Fino a v3.4.54 facevamo soft-detach
    (SET NULL job_cost_line_id sui Booking) → produceva booking orfani senza
    lavorazione → cost report vuoto → paradosso segnalato da Matteo.

    Nuova policy: la riga di quote NON può essere cancellata finché esistono
    booking attivi sulla JobCostLine collegata. Modifica resta consentita.
    Per togliere la riga: cancella prima i booking (o marcali cancelled),
    oppure modifica la riga (descrizione, qty, prezzo) senza eliminarla.
    """
    line = db.query(QuoteLine).filter(QuoteLine.id == line_id).first()
    if not line:
        raise HTTPException(404)

    # Trova JobCostLine collegate
    cost_lines = db.query(JobCostLine).filter(
        JobCostLine.quote_line_id == line_id
    ).all()

    # v3.4.55 — HARD-BLOCK su booking attivi
    blocking_bookings = []
    for jcl in cost_lines:
        active_bk = db.query(Booking).filter(
            Booking.job_cost_line_id == jcl.id,
            Booking.status != BookingStatus.cancelled,
        ).all()
        for b in active_bk:
            blocking_bookings.append({
                "booking_id": b.id,
                "start": b.start_datetime.isoformat() if b.start_datetime else None,
                "execution_status": b.execution_status.value if hasattr(b.execution_status, "value") else str(b.execution_status),
            })
    # v3.5.0-alpha.171.8 (Sprint 2 Step 6) — Propagazione su CR per quote
    # APPROVATA: invece di hard-block, sposta JCL+booking a "Quotazione a
    # Consuntivo" del progetto (auto-creandola se manca).
    #
    # v3.5.0-alpha.172.18 — Estensione: propagazione attiva su QUALUNQUE delete
    # su quote approvata non-phantom (anche senza booking attivi) per coerenza
    # con regola "quote approvata immutabile". Inclusione cascade JobDeliverable
    # nella propagazione (prima venivano lasciati orfani sul Job → consegne
    # fantasma in /jobs/{id}).
    parent_quote = db.query(Quote).filter(Quote.id == quote_id).first()
    is_approved_propagation = (
        parent_quote
        and parent_quote.status == QuoteStatus.approved
        and not parent_quote.is_phantom
    )
    if is_approved_propagation:
        # Ha ore maturate (booking attivi): muovi a Consuntivo. Crea se manca.
        from app.models import PhantomStatus
        phantom = db.query(Quote).filter(
            Quote.project_id == parent_quote.project_id,
            Quote.is_phantom == True,  # noqa: E712
            Quote.phantom_status == PhantomStatus.standby,
        ).first()
        if not phantom:
            from datetime import date as _date
            phantom = Quote(
                number=_next_quote_number_progressive(db),
                version=1,
                project_id=parent_quote.project_id,
                client_id=parent_quote.client_id,
                title=f"Consuntivo — {parent_quote.title or 'progetto'}",
                status=QuoteStatus.approved,
                is_phantom=True,
                phantom_status=PhantomStatus.standby,
                issue_date=_date.today(),
                notes=f"Generata da propagazione delete voce in quote {parent_quote.number}.",
                tenant_id=current_tenant_id(),
            )
            db.add(phantom)
            db.flush()
        # Clona QuoteLine nella phantom (qty_quoted=0, ore reali da JCL.actual)
        from app.services.reverse_quote import _next_position, _next_sort_order
        cloned = QuoteLine(
            quote_id=phantom.id,
            price_item_id=line.price_item_id,
            section="A",
            position=_next_position(phantom),
            description=line.description,
            detail=(line.detail or "") + f"\n[ex-quote {parent_quote.number} L#{line.id}, voce eliminata dopo approval]",
            quantity=0.0,
            unit=line.unit,
            unit_price=line.unit_price,
            total=0.0,
            hardcosts=line.hardcosts,
            sort_order=_next_sort_order(phantom),
        )
        db.add(cloned)
        db.flush()
        # Sposta tutte le JCL collegate alla nuova QuoteLine sulla phantom
        for jcl in cost_lines:
            jcl.quote_line_id = cloned.id
        # v3.5.0-alpha.172.18 — Sposta anche i JobDeliverable spawnati dalla
        # QuoteLine originale alla cloned. Senza questo passaggio le consegne
        # restavano linkate alla QL eliminata → fantasma in /jobs/{id} deliveries.
        deliverables_moved = db.query(JobDeliverable).filter(
            JobDeliverable.quote_line_id == line_id
        ).all()
        for d in deliverables_moved:
            d.quote_line_id = cloned.id
        # Cancella la QuoteLine originale (ora orfana, ma JCL+Deliverable stanno nella phantom)
        db.delete(line)
        q_main = db.query(Quote).options(
            joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
        ).filter(Quote.id == quote_id).first()
        q_main.lines = [l for l in q_main.lines if l.id != line_id]
        _recalc_quote(q_main)
        from app.services.reverse_quote import _recalc_quote_totals as _rqt
        _rqt(phantom)
        db.commit()
        return {
            "ok": True,
            "propagated_to_phantom": True,
            "phantom_quote_id": phantom.id,
            "phantom_number": phantom.number,
            "cost_lines_moved": len(cost_lines),
            "deliverables_moved": len(deliverables_moved),
            "blocking_bookings": len(blocking_bookings),
        }
    if blocking_bookings:
        raise HTTPException(
            409,
            f"Impossibile eliminare: questa riga ha {len(blocking_bookings)} "
            f"booking attivi collegati. Cancella o annulla prima i booking, "
            f"oppure modifica la riga senza eliminarla. "
            f"Booking ostativi: {[b['booking_id'] for b in blocking_bookings[:5]]}"
            + (f" e altri {len(blocking_bookings)-5}" if len(blocking_bookings) > 5 else "")
        )

    # Cascade su JobCostLine "pulite" (senza booking attivi): rimosse insieme
    for jcl in cost_lines:
        # Blocca anche se il job è in stato terminale
        if jcl.job and jcl.job.status in (JobStatus.completed, JobStatus.invoiced):
            raise HTTPException(
                409,
                f"Impossibile cancellare: il job {jcl.job.code} è in stato "
                f"{jcl.job.status.value} e ha già consuntivato questa lavorazione."
            )
        # v3.5.0-alpha.172.36 (Sprint 2 BLOCCO 2) — slice-lock guard:
        # cancellare una JCL con slice attive lascerebbe slice orfane
        # (FK senza ondelete CASCADE → integrity error o NULL silente).
        assert_jcl_lock_safe(db, jcl.id, action="eliminare la voce di costo collegata a")
        # Per TimePunch (HR, separato dal cost report): soft-detach OK
        db.query(TimePunch).filter(
            TimePunch.job_cost_line_id == jcl.id
        ).update({"job_cost_line_id": None}, synchronize_session=False)
        db.delete(jcl)

    # v3.5.0-alpha.172.18 — Cascade su JobDeliverable: rimosse insieme se pulite.
    # Block se confermate (confirmed_at) o se già in batch/billed/paid (billing_status).
    deliverables = db.query(JobDeliverable).filter(JobDeliverable.quote_line_id == line_id).all()
    for d in deliverables:
        from app.models import DeliverableBillingStatus
        if d.confirmed_at:
            raise HTTPException(
                409,
                f"Impossibile cancellare: consegna '{d.name}' è già confermata "
                f"({d.confirmed_at.isoformat()}). Annulla la conferma o usa una "
                f"nuova versione di quote."
            )
        if d.billing_status in (
            DeliverableBillingStatus.in_batch,
            DeliverableBillingStatus.billed,
            DeliverableBillingStatus.paid,
        ):
            raise HTTPException(
                409,
                f"Impossibile cancellare: consegna '{d.name}' è già "
                f"{d.billing_status.value} (fatturata o in fattura). Crea nuova "
                f"versione di quote per modifiche."
            )
        db.delete(d)

    db.delete(line)
    q = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
    ).filter(Quote.id == quote_id).first()
    q.lines = [l for l in q.lines if l.id != line_id]
    _recalc_quote(q)
    db.commit()
    return {
        "ok": True,
        "cost_lines_deleted": len(cost_lines),
        "deliverables_deleted": len(deliverables),
    }


# ── v3.5.0-alpha.171.10 (Sprint 2 Step 6 batch) ──
# Batch delete N line in singola chiamata. Per ogni line applica la stessa
# logica di `delete_quote_line` (incluso propagation su Consuntivo per quote
# approved con booking attivi). Raccoglie risultati e ritorna summary.

@router.post("/api/{quote_id}/lines-batch-delete", dependencies=[RequireEditQuotes])
async def batch_delete_quote_lines(
    quote_id: int,
    line_ids: str = Form(..., description="CSV di line IDs da eliminare"),
    db: Session = Depends(get_db),
):
    """Batch delete di più QuoteLine in singola transazione.

    Per ogni line:
    - Se quote NOT approved: hard-block 409 se ha booking attivi → l'INTERA
      batch fallisce (rollback). Altrimenti delete cascadea JCL "pulite".
    - Se quote approved: propaga JCL+booking alla Consuntivo del progetto
      (auto-crea se manca). Stessa Consuntivo riusata per tutte le line del
      batch (ottimizzazione).

    Body: `line_ids=1,5,12,18` (CSV).
    Response: `{ok: true, deleted: N, propagated_to_phantom: bool,
                phantom_quote_id, cost_lines_moved, details: [...]}`
    """
    from app.models import PhantomStatus
    from datetime import date as _date
    try:
        ids = [int(x.strip()) for x in line_ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(400, "line_ids deve essere CSV di interi")
    if not ids:
        raise HTTPException(400, "line_ids vuoto")

    parent_quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not parent_quote:
        raise HTTPException(404, "Quote non trovata")
    is_approved = (
        parent_quote.status == QuoteStatus.approved
        and not parent_quote.is_phantom
    )

    # Phantom target (singolo per la batch, lazy-init alla prima propagation)
    phantom = None
    def _get_or_create_phantom():
        nonlocal phantom
        if phantom is not None:
            return phantom
        existing = db.query(Quote).filter(
            Quote.project_id == parent_quote.project_id,
            Quote.is_phantom == True,  # noqa: E712
            Quote.phantom_status == PhantomStatus.standby,
        ).first()
        if existing:
            phantom = existing
            return phantom
        phantom = Quote(
            number=_next_quote_number_progressive(db),
            version=1,
            project_id=parent_quote.project_id,
            client_id=parent_quote.client_id,
            title=f"Consuntivo — {parent_quote.title or 'progetto'}",
            status=QuoteStatus.approved,
            is_phantom=True,
            phantom_status=PhantomStatus.standby,
            issue_date=_date.today(),
            notes=f"Generata da batch-delete in quote {parent_quote.number}.",
            tenant_id=current_tenant_id(),
        )
        db.add(phantom)
        db.flush()
        return phantom

    from app.services.reverse_quote import _next_position, _next_sort_order, _recalc_quote_totals as _rqt
    total_moved = 0
    details = []
    for lid in ids:
        line = db.query(QuoteLine).filter(
            QuoteLine.id == lid, QuoteLine.quote_id == quote_id,
        ).first()
        if not line:
            details.append({"line_id": lid, "skipped": "not_found"})
            continue
        cost_lines = db.query(JobCostLine).filter(
            JobCostLine.quote_line_id == lid
        ).all()
        # Check booking attivi
        blocking_count = 0
        for jcl in cost_lines:
            blocking_count += db.query(Booking).filter(
                Booking.job_cost_line_id == jcl.id,
                Booking.status != BookingStatus.cancelled,
            ).count()
        if blocking_count > 0 and not is_approved:
            db.rollback()
            raise HTTPException(
                409,
                f"Line #{lid} ha {blocking_count} booking attivi: impossibile "
                f"eliminare in batch su quote non-approved. Annulla i booking "
                f"prima, oppure approva la quote (la propagazione su Consuntivo "
                f"avverrà automatica)."
            )
        # v3.5.0-alpha.172.18 — Su quote approved propaga SEMPRE (anche senza
        # booking attivi) per coerenza con regola "quote approvata immutabile".
        # Include cascade JobDeliverable per evitare consegne fantasma.
        if is_approved:
            phantom = _get_or_create_phantom()
            cloned = QuoteLine(
                quote_id=phantom.id,
                price_item_id=line.price_item_id,
                section="A",
                position=_next_position(phantom),
                description=line.description,
                detail=(line.detail or "") + f"\n[ex-quote {parent_quote.number} L#{line.id}, batch-delete]",
                quantity=0.0,
                unit=line.unit,
                unit_price=line.unit_price,
                total=0.0,
                hardcosts=line.hardcosts,
                sort_order=_next_sort_order(phantom),
            )
            db.add(cloned)
            db.flush()
            for jcl in cost_lines:
                jcl.quote_line_id = cloned.id
            deliverables_to_move = db.query(JobDeliverable).filter(
                JobDeliverable.quote_line_id == lid
            ).all()
            for d in deliverables_to_move:
                d.quote_line_id = cloned.id
            total_moved += len(cost_lines)
            db.delete(line)
            details.append({
                "line_id": lid, "propagated": True,
                "cost_lines_moved": len(cost_lines),
                "deliverables_moved": len(deliverables_to_move),
            })
            continue
        # No blocking bookings + quote not approved: pulizia diretta
        for jcl in cost_lines:
            if jcl.job and jcl.job.status in (JobStatus.completed, JobStatus.invoiced):
                db.rollback()
                raise HTTPException(
                    409, f"JCL del job {jcl.job.code} è in stato {jcl.job.status.value}, non cancellabile."
                )
            # v3.5.0-alpha.172.36 (Sprint 2 BLOCCO 2) — slice-lock guard
            try:
                assert_jcl_lock_safe(db, jcl.id, action="eliminare la voce di costo collegata a")
            except HTTPException as _e:
                db.rollback()
                raise
            db.query(TimePunch).filter(
                TimePunch.job_cost_line_id == jcl.id
            ).update({"job_cost_line_id": None}, synchronize_session=False)
            db.delete(jcl)
        # v3.5.0-alpha.172.18 — Cascade Deliverable (block se confermati/billed)
        from app.models import DeliverableBillingStatus
        deliverables_clean = db.query(JobDeliverable).filter(
            JobDeliverable.quote_line_id == lid
        ).all()
        for d in deliverables_clean:
            if d.confirmed_at:
                db.rollback()
                raise HTTPException(
                    409,
                    f"Consegna '{d.name}' (line #{lid}) è già confermata: "
                    f"impossibile eliminare. Annulla conferma o crea nuova versione."
                )
            if d.billing_status in (
                DeliverableBillingStatus.in_batch,
                DeliverableBillingStatus.billed,
                DeliverableBillingStatus.paid,
            ):
                db.rollback()
                raise HTTPException(
                    409,
                    f"Consegna '{d.name}' (line #{lid}) è {d.billing_status.value}: "
                    f"impossibile eliminare. Crea nuova versione di quote."
                )
            db.delete(d)
        db.delete(line)
        details.append({
            "line_id": lid, "deleted_clean": True,
            "cost_lines_deleted": len(cost_lines),
            "deliverables_deleted": len(deliverables_clean),
        })

    # Recalc parent + phantom
    parent_fresh = db.query(Quote).options(
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category)
    ).filter(Quote.id == quote_id).first()
    _recalc_quote(parent_fresh)
    if phantom:
        _rqt(phantom)
    db.commit()
    return {
        "ok": True,
        "deleted": len([d for d in details if d.get("deleted_clean") or d.get("propagated")]),
        "propagated_to_phantom": phantom is not None,
        "phantom_quote_id": phantom.id if phantom else None,
        "phantom_number": phantom.number if phantom else None,
        "cost_lines_moved": total_moved,
        "details": details,
    }


# ── Soft-delete dell'intera Quote (v3.5.0-alpha.7) ───────────

@router.delete("/api/{quote_id}")
async def delete_quote(
    quote_id: int,
    request: Request,
    force: bool = False,
    db: Session = Depends(get_db),
):
    """Soft-delete di una Quote (sposta nel cestino) o pulizia totale (admin).

    Permessi:
    - `delete_quotes`: soft-delete normale. HARD-BLOCK 409 se ci sono booking
      attivi sul Job collegato (con elenco bloccanti nel response body).
    - `purge_total` (solo admin per default): può passare `?force=true` per
      hard-delete cascade su Quote + Job + JobCostLine + Booking + assignments.

    Response success:
      200 {"ok": true, "mode": "soft" | "purge_total", ...}

    Response HARD-BLOCK:
      409 {"detail": "...", "blocking": {"bookings": [...]}, "can_force": true|false}
    """
    from app.services.rbac import current_user_optional, has_permission
    from app.services.soft_delete import (
        soft_delete_quote, fetch_quote_including_trash, DeleteBlocked,
    )
    from fastapi.responses import JSONResponse

    user = current_user_optional(request)
    if not has_permission(user, "delete_quotes"):
        raise HTTPException(403, "Permesso negato (delete_quotes)")
    if force and not has_permission(user, "purge_total"):
        raise HTTPException(403, "Solo un admin con permesso 'purge_total' può forzare la pulizia totale")

    q = fetch_quote_including_trash(db, quote_id)
    if not q:
        raise HTTPException(404, "Quotazione non trovata")

    try:
        result = soft_delete_quote(db, q, user=user, force=force)
    except DeleteBlocked as e:
        # Mantieni in lista anche se l'utente è admin: il frontend deciderà
        # se mostrare il bottone "Pulizia totale" usando `can_force`.
        return JSONResponse(
            status_code=409,
            content={
                "detail":    e.message,
                "blocking":  {"bookings": e.bookings, "jobs": e.jobs},
                "can_force": has_permission(user, "purge_total"),
            },
        )

    db.commit()
    return result


@router.post("/api/{quote_id}/restore")
async def restore_quote_endpoint(
    quote_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Ripristina una Quote dal cestino. Idempotente."""
    from app.services.rbac import current_user_optional, has_permission
    from app.services.soft_delete import fetch_quote_including_trash, restore_quote

    user = current_user_optional(request)
    if not has_permission(user, "restore_trash"):
        raise HTTPException(403, "Permesso negato (restore_trash)")

    q = fetch_quote_including_trash(db, quote_id)
    if not q:
        raise HTTPException(404, "Quotazione non trovata")
    result = restore_quote(db, q)
    db.commit()
    return result


@router.post("/api/{quote_id}/convert-to-job", dependencies=[RequireEditQuotes], deprecated=True)
async def convert_to_job_legacy(
    quote_id: int,
    job_code: Optional[str] = Form(None),
    start_date: Optional[date] = Form(None),
    end_date: Optional[date] = Form(None),
    db: Session = Depends(get_db),
):
    """DEPRECATED dal v3.4.8 — usare PUT /api/{quote_id}/status con status=approved.

    Mantenuto come wrapper per retrocompatibilità: ignora `job_code`/`start_date`/
    `end_date` e delega alla nuova logica auto-promote (codice auto-generato
    {project.code}-J{N}, title da project.title, no date hardcoded).
    """
    q = (
        db.query(Quote)
        .options(joinedload(Quote.lines), joinedload(Quote.project), joinedload(Quote.job))
        .filter(Quote.id == quote_id)
        .first()
    )
    if not q:
        raise HTTPException(404)
    if q.job:
        raise HTTPException(400, "Job già esistente")
    job = _create_job_from_quote(db, q)
    q.status = QuoteStatus.approved
    db.commit()
    return {"job_id": job.id, "job_code": job.code, "deprecated": True}


@router.get("/api/{quote_id}/pdf")
async def quote_pdf(quote_id: int, db: Session = Depends(get_db)):
    from app.services.quote_pdf import generate_quote_pdf
    q = db.query(Quote).options(
        joinedload(Quote.client),
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category),
    ).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404)
    pdf = generate_quote_pdf(q)
    filename = f"quote_{q.number.replace('/', '-')}.pdf"
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── Export CSV/Excel quotazione ──────────────────────────────

def _quote_export_rows(q: Quote) -> tuple[list[str], list[list], list[list]]:
    """Header colonne + righe per categoria con subtotali + footer totali."""
    headers = ["Categoria", "Posizione", "Descrizione", "Quantità", "Unità",
               "Prezzo unitario €", "Sconto riga %", "Totale riga €"]
    sorted_lines = sorted(q.lines, key=lambda l: l.sort_order)
    groups: dict[str, list] = {}
    order = []
    for l in sorted_lines:
        cat = _line_category(l)
        if cat not in groups:
            groups[cat] = []
            order.append(cat)
        groups[cat].append(l)

    cat_disc = q.category_discounts or {}
    body = []
    for cat in order:
        cat_lines = groups[cat]
        body.append([cat.upper(), "", "", "", "", "", "", ""])
        cat_subtotal = 0.0
        for l in cat_lines:
            disc = (l.line_discount_pct or 0) * 100
            body.append([
                "", l.position or "", l.description or "",
                l.quantity, l.unit or "",
                round(l.unit_price or 0, 2),
                round(disc, 2) if disc else "",
                round(l.total or 0, 2),
            ])
            cat_subtotal += (l.total or 0)
        body.append(["", "", f"Subtotale {cat}", "", "", "", "", round(cat_subtotal, 2)])
        cd = cat_disc.get(cat, 0)
        if cd:
            disc_amount = cat_subtotal * cd
            body.append(["", "", f"Sconto categoria {cd*100:.1f}%", "", "", "", "", -round(disc_amount, 2)])

    pkg_pct = abs((q.package_discount or 0) * 100)
    pkg_amount = (q.subtotal or 0) - (q.total_after_discount or 0)
    vat_amount = (q.total_with_vat or 0) - (q.total_after_discount or 0)
    footer = [
        ["", "", "Totale lordo (no sconti)", "", "", "", "", round(q.subtotal_gross or 0, 2)],
        ["", "", "Subtotale dopo sconti voci/categorie", "", "", "", "", round(q.subtotal or 0, 2)],
    ]
    if pkg_pct > 0.05:
        footer.append(["", "", f"Sconto pacchetto {pkg_pct:.1f}%", "", "", "", "", -round(pkg_amount, 2)])
    footer.extend([
        ["", "", "Totale netto base IVA", "", "", "", "", round(q.total_after_discount or 0, 2)],
        ["", "", f"IVA {q.vat_rate:.0f}%", "", "", "", "", round(vat_amount, 2)],
        ["", "", "TOTALE (incl. IVA)", "", "", "", "", round(q.total_with_vat or 0, 2)],
    ])
    return headers, body, footer


@router.get("/api/{quote_id}/export.csv")
async def quote_export_csv(quote_id: int, db: Session = Depends(get_db)):
    import csv, io
    q = db.query(Quote).options(
        joinedload(Quote.client),
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category),
    ).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404)
    headers, body, footer = _quote_export_rows(q)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow([f"Quotazione {q.number} — {q.title or ''}"])
    if q.client:
        w.writerow([f"Cliente: {q.client.name}"])
    w.writerow([f"Data: {q.issue_date}"])
    w.writerow([])
    w.writerow(headers)
    for r in body: w.writerow(r)
    w.writerow([])
    for r in footer: w.writerow(r)
    body_str = "﻿" + buf.getvalue()
    filename = f"quote_{q.number.replace('/', '-')}.csv"
    return Response(content=body_str, media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/api/{quote_id}/export.xlsx")
async def quote_export_xlsx(quote_id: int, db: Session = Depends(get_db)):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(Quote).options(
        joinedload(Quote.client),
        joinedload(Quote.lines).joinedload(QuoteLine.price_item).joinedload(PriceItem.category),
    ).filter(Quote.id == quote_id).first()
    if not q: raise HTTPException(404)

    headers, body, footer = _quote_export_rows(q)
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotazione"

    indigo = PatternFill("solid", fgColor="6272f5")
    cat_fill = PatternFill("solid", fgColor="EEF1FF")
    sub_fill = PatternFill("solid", fgColor="F8F9FB")
    bold_white = Font(bold=True, color="FFFFFF")
    bold_indigo = Font(bold=True, color="6272f5")
    bold_dark = Font(bold=True, color="1A1A2E")
    thin = Side(border_style="thin", color="DDE1F0")

    # Intestazione
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Quotazione {q.number} — {q.title or ''}"
    ws["A1"].font = Font(bold=True, size=14, color="6272f5")
    ws["A2"] = f"Cliente: {q.client.name if q.client else '—'}"
    ws["A3"] = f"Data emissione: {q.issue_date}    Valida fino al: {q.valid_until or '—'}"
    ws["A2"].font = Font(color="6B7280")
    ws["A3"].font = Font(color="6B7280")

    start_row = 5
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=h)
        c.font = bold_white
        c.fill = indigo
        c.alignment = Alignment(vertical="center", horizontal="center")

    r = start_row + 1
    for row in body:
        for col_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        # Riga categoria (prima colonna piena, altre vuote)
        if row[0] and not row[1]:
            for col_idx in range(1, 9):
                ws.cell(row=r, column=col_idx).fill = cat_fill
                ws.cell(row=r, column=col_idx).font = bold_indigo
        # Riga subtotale (descrizione "Subtotale ..." colonna C, totale colonna H)
        elif (row[2] and (str(row[2]).startswith("Subtotale ") or str(row[2]).startswith("Sconto categoria "))):
            for col_idx in range(1, 9):
                ws.cell(row=r, column=col_idx).fill = sub_fill
            ws.cell(row=r, column=3).font = bold_dark
            ws.cell(row=r, column=8).font = bold_dark
        # Format numerico per colonne prezzi/totali
        for col_idx in (4, 6, 7, 8):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        r += 1

    # Footer totals
    r += 1
    for row in footer:
        for col_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            if col_idx == 3: cell.font = bold_dark
            if col_idx == 8 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
                cell.font = bold_dark
        # Ultima riga (TOTALE incl. IVA): sfondo indigo
        if str(row[2]).startswith("TOTALE"):
            for col_idx in range(1, 9):
                ws.cell(row=r, column=col_idx).fill = indigo
                ws.cell(row=r, column=col_idx).font = bold_white
        r += 1

    # Larghezze (uso get_column_letter perché row=1 è una merged cell)
    widths = [16, 10, 50, 10, 10, 16, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{start_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"quote_{q.number.replace('/', '-')}.xlsx"
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── DUPLICAZIONE / VERSIONING (v3.4.39) ──────────────────────


def _quote_root(db: Session, q: Quote) -> Quote:
    """Risale la catena delle versioni fino alla radice (parent_quote_id IS NULL)."""
    seen = {q.id}
    while q.parent_quote_id is not None:
        parent = db.query(Quote).filter(Quote.id == q.parent_quote_id).first()
        if not parent or parent.id in seen:
            break
        seen.add(parent.id)
        q = parent
    return q


def _quote_chain(db: Session, root: Quote) -> list[Quote]:
    """Ritorna root + tutti i discendenti (BFS sui parent_quote_id)."""
    visited = {root.id}
    chain = [root]
    frontier = [root.id]
    while frontier:
        children = db.query(Quote).filter(Quote.parent_quote_id.in_(frontier)).all()
        frontier = []
        for c in children:
            if c.id not in visited:
                visited.add(c.id)
                chain.append(c)
                frontier.append(c.id)
    return chain


def _next_quote_number_progressive(db: Session, project: Optional[Project] = None, client = None) -> str:
    """v3.5.0-alpha.66.14.8 — wrapper sul numbering service unificato.
    v3.5.0-alpha.115 — Cabling NumberingConfig: legge format custom dal
    pannello /settings#numbering. Variabili supportate per "quote":
    YYYY/YY/MM/DD/YYYYMMDD/NNN/NN/NNNN/PROJECT_CODE/CLIENT_CODE.
    Fallback al pattern default Q-{YYYY}-{NNN} se config assente.

    v3.5.0-alpha.172.97 — Folder-view: tutte le quote nuove nascono con
    suffix `-v1`. Idempotente: se il NumberingConfig pattern include già
    `-vN`, non viene duplicato.
    """
    from app.services.numbering import (
        gen_doc_code, next_year_progressive, with_v1_suffix,
    )
    from app.context import current_tenant_id
    try:
        code, _ = gen_doc_code(
            db, "quote",
            tenant_id=current_tenant_id(),
            project_code=(project.code if project else None),
            client_code=(client.name[:8].upper() if client and getattr(client, "name", None) else None),
        )
        full = with_v1_suffix(code)
        # v3.5.0-alpha.172.145 — Collision detection robusta. Prima si
        # controllava solo la stringa ESATTA (con soft-delete): falliva quando
        # il base progressivo era già usato da una quote ATTIVA in altra
        # versione. Scenario reale: contatore NumberingConfig.current_seq
        # disallineato dal max reale (il versioning -vN e gli import snapshot
        # NON bumpano il contatore) → gen_doc_code riemette un base già usato,
        # e il bin-rename del -v1 cancellato lasciava libera la stringa esatta.
        # Ora si controlla anche il BASE tra le quote attive → fallback allo
        # scan autoritativo (next_year_progressive = max reale + 1).
        from app.models import Quote as _Q
        from app.services.numbering import split_version_suffix
        base_code, _v = split_version_suffix(full)
        exact = (
            db.query(_Q).execution_options(include_deleted=True)
            .filter(_Q.number == full).first()
        )
        base_clash = (
            db.query(_Q).filter(_Q.number.like(base_code + "-v%")).first()
        )
        if exact or base_clash:
            # Fallback: scan del max progressivo reale (gestisce -vN, ignora
            # i binnati ~Bn~ che non matchano il prefisso).
            return with_v1_suffix(next_year_progressive(
                db, Quote, base="Q", code_field="number", include_deleted=True,
            ))
        return full
    except Exception as _e:
        print(f"[quote_numbering] gen_doc_code failed, fallback: {_e}")
        return with_v1_suffix(next_year_progressive(
            db, Quote, base="Q", code_field="number", include_deleted=True,
        ))


def _copy_quote_lines(src_lines: list, dest_quote_id: int, track_parent: bool) -> list[QuoteLine]:
    """Crea copie delle QuoteLine. Se track_parent, valorizza parent_line_id (versioning).
    Ritorna la lista delle nuove righe (non ancora flushate)."""
    new_lines = []
    for sl in sorted(src_lines, key=lambda x: x.sort_order):
        nl = QuoteLine(
            quote_id=dest_quote_id,
            price_item_id=sl.price_item_id,
            section=sl.section,
            position=sl.position,
            description=sl.description,
            detail=sl.detail,
            quantity=sl.quantity,
            unit=sl.unit,
            price_level=sl.price_level,
            unit_price=sl.unit_price,
            allowance=sl.allowance,
            line_discount_pct=sl.line_discount_pct,
            total=sl.total,
            hardcosts=sl.hardcosts,
            sort_order=sl.sort_order,
            category_override=sl.category_override,
            source_hint=sl.source_hint,
            parent_line_id=sl.id if track_parent else None,
        )
        new_lines.append(nl)
    return new_lines


@router.post("/api/{quote_id}/duplicate")
async def duplicate_quote(
    quote_id: int,
    request: Request,
    project_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
):
    """Duplica una quote in modo INDIPENDENTE (no parent_quote_id).
    Use case: scenario alternativo, template per un nuovo progetto.
    Numero auto-progressivo `Q-{anno}-NNN`. Status=draft.

    v3.4.43 — `project_id` opzionale: se valorizzato, la copia viene
    associata a un progetto diverso da quello sorgente (use case template
    per nuovo progetto). Il `client_id` viene riallineato al cliente del
    nuovo progetto."""
    from app.services.rbac import current_user_optional, has_permission
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di duplicare le quotazioni")

    src = (
        db.query(Quote)
        .options(joinedload(Quote.lines))
        .filter(Quote.id == quote_id).first()
    )
    if not src:
        raise HTTPException(404, "Quotazione sorgente non trovata")

    # v3.4.43 — Risolvi project + client target
    target_project_id = project_id if project_id else src.project_id
    target_client_id = src.client_id
    if project_id and project_id != src.project_id:
        target_project = db.query(Project).filter(Project.id == project_id).first()
        if not target_project:
            raise HTTPException(404, f"Progetto target #{project_id} non trovato")
        target_client_id = target_project.client_id

    # v3.5.0-alpha.115 — passa project+client per espandere {PROJECT_CODE}/{CLIENT_CODE}
    _proj_for_num = target_project if (project_id and project_id != src.project_id) else (src.project if src else None)
    _cli_for_num = db.query(Client).filter(Client.id == target_client_id).first() if target_client_id else None
    new_q = Quote(
        number=_next_quote_number_progressive(db, project=_proj_for_num, client=_cli_for_num),
        version=1,
        project_id=target_project_id,
        client_id=target_client_id,
        title=src.title,
        status=QuoteStatus.draft,
        issue_date=date.today(),
        valid_until=src.valid_until,
        production_material=src.production_material,
        length_minutes=src.length_minutes,
        fps=src.fps,
        delivery_format=src.delivery_format,
        shooting_days=src.shooting_days,
        shooting_format=src.shooting_format,
        package_discount=src.package_discount,
        category_discounts=dict(src.category_discounts) if src.category_discounts else None,
        category_order=list(src.category_order) if src.category_order else None,
        vat_rate=src.vat_rate,
        notes=src.notes,
        payment_terms=src.payment_terms,
        # NESSUN parent_quote_id: duplicato indipendente.
    )
    db.add(new_q); db.flush()

    new_lines = _copy_quote_lines(src.lines, new_q.id, track_parent=False)
    db.add_all(new_lines)
    db.flush()

    _recalc_quote(new_q)
    db.commit()
    db.refresh(new_q)
    return {
        "id": new_q.id,
        "number": new_q.number,
        "title": new_q.title,
        "project_id": new_q.project_id,
        "client_id": new_q.client_id,
        "lines_count": len(new_lines),
        "kind": "duplicate",
    }


@router.put("/api/{quote_id}/move-to-project")
async def move_quote_to_project(
    quote_id: int,
    request: Request,
    project_id: int = Form(...),
    db: Session = Depends(get_db),
):
    """v3.4.43 — Sposta una quote esistente a un altro progetto.
    Vincoli:
      - Solo quote in stato `draft` (una quote già inviata/approvata/etc
        non può cambiare progetto: è un cambio di scope vincolante).
      - La quote NON deve avere un Job collegato (incoerenza grave).
      - Il cliente viene riallineato al cliente del progetto target.
    """
    from app.services.rbac import current_user_optional, has_permission
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di spostare le quotazioni")

    q = db.query(Quote).options(joinedload(Quote.job)).filter(Quote.id == quote_id).first()
    if not q:
        raise HTTPException(404, "Quotazione non trovata")
    if q.status != QuoteStatus.draft:
        raise HTTPException(
            400,
            f"Lo spostamento di progetto è ammesso solo su quote in bozza. "
            f"Stato corrente: {q.status.value}."
        )
    if q.job:
        raise HTTPException(
            400,
            f"Quote ha un job collegato ({q.job.code}): impossibile spostare di progetto. "
            "Annulla il job o crea una nuova versione."
        )
    target = db.query(Project).filter(Project.id == project_id).first()
    if not target:
        raise HTTPException(404, f"Progetto target #{project_id} non trovato")

    old_project_id = q.project_id
    q.project_id = project_id
    q.client_id = target.client_id
    db.commit()
    return {
        "id": q.id,
        "number": q.number,
        "old_project_id": old_project_id,
        "new_project_id": project_id,
        "new_client_id": target.client_id,
    }


@router.post("/api/{quote_id}/new-version")
async def new_version_quote(
    quote_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Crea una nuova versione della quote (parent_quote_id valorizzato).
    Numero `{root_number}-v{N}` dove N = max(version) + 1 nella catena.
    Le righe ereditano `parent_line_id` per re-bind preciso al migrate-job."""
    from app.services.rbac import current_user_optional, has_permission
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di creare nuove versioni")

    src = (
        db.query(Quote)
        .options(joinedload(Quote.lines))
        .filter(Quote.id == quote_id).first()
    )
    if not src:
        raise HTTPException(404, "Quotazione sorgente non trovata")

    root = _quote_root(db, src)
    chain = _quote_chain(db, root)
    next_version = max(q.version for q in chain) + 1
    # v3.4.50.1: pulisci eventuale suffisso `-vN` dal root.number per evitare
    # numeri duplicati come `Q-2026-001-v1-v2`. Pattern: rstrip al match -v\d+
    import re
    base_number = re.sub(r"-v\d+$", "", root.number)
    new_number = f"{base_number}-v{next_version}"

    # Conflitto improbabile ma garantiamo unicità (anche su cestino)
    if (db.query(Quote)
          .execution_options(include_deleted=True)
          .filter(Quote.number == new_number).first()):
        raise HTTPException(409, f"Numero quotazione '{new_number}' già esistente (eventualmente nel cestino)")

    new_q = Quote(
        number=new_number,
        version=next_version,
        parent_quote_id=src.id,
        project_id=src.project_id,
        client_id=src.client_id,
        title=src.title,
        status=QuoteStatus.draft,
        issue_date=date.today(),
        valid_until=src.valid_until,
        production_material=src.production_material,
        length_minutes=src.length_minutes,
        fps=src.fps,
        delivery_format=src.delivery_format,
        shooting_days=src.shooting_days,
        shooting_format=src.shooting_format,
        package_discount=src.package_discount,
        category_discounts=dict(src.category_discounts) if src.category_discounts else None,
        category_order=list(src.category_order) if src.category_order else None,
        vat_rate=src.vat_rate,
        notes=src.notes,
        payment_terms=src.payment_terms,
    )
    db.add(new_q); db.flush()

    new_lines = _copy_quote_lines(src.lines, new_q.id, track_parent=True)
    db.add_all(new_lines)
    db.flush()

    _recalc_quote(new_q)
    db.commit()
    db.refresh(new_q)
    return {
        "id": new_q.id,
        "number": new_q.number,
        "version": new_q.version,
        "parent_quote_id": new_q.parent_quote_id,
        "title": new_q.title,
        "lines_count": len(new_lines),
        "kind": "version",
    }


@router.get("/api/{quote_id}/versions")
async def list_quote_versions(quote_id: int, db: Session = Depends(get_db)):
    """Catena versioni completa (root + discendenti, ordinata per version)."""
    q = db.query(Quote).filter(Quote.id == quote_id).first()
    if not q:
        raise HTTPException(404)
    root = _quote_root(db, q)
    chain = _quote_chain(db, root)
    chain_sorted = sorted(chain, key=lambda x: (x.version, x.id))
    return {
        "root_id": root.id,
        "current_id": q.id,
        "versions": [
            {
                "id": c.id, "number": c.number, "version": c.version,
                "status": c.status, "title": c.title,
                "parent_quote_id": c.parent_quote_id,
                "superseded_by_id": c.superseded_by_id,
                "has_job": c.job is not None,
                "job_id": c.job.id if c.job else None,
                "job_code": c.job.code if c.job else None,
                "issue_date": str(c.issue_date),
                "total_after_discount": c.total_after_discount,
            }
            for c in chain_sorted
        ],
    }


# ───── v3.5.0-alpha.171.3 — Sprint 2 Step 3 ─────
# Endpoint workflow Quotazione a Consuntivo (ex Phantom Quote).
#
# 1) promote-phantom: Consuntivo → quote effettiva (is_phantom=False,
#    phantom_status=promoted). Mantiene status approved.
# 2) merge-into/{target}: accorpa Consuntivo in quote target (anche
#    approvata). Crea NUOVA VERSIONE di target con voci Consuntivo
#    aggiunte. Consuntivo marcata merged_into. Target marcata superseded.

@router.post("/api/{quote_id}/promote-phantom")
async def promote_phantom_quote(
    quote_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.171.3 — Promuove una Quotazione a Consuntivo standby a
    quote effettiva. is_phantom passa a False, phantom_status passa a
    `promoted` (per audit). Lo status quote (approved) resta.

    L'effetto pratico: la quote diventa "normale" e perde il badge Consuntivo.
    Il Job/JCL legato non viene toccato (resta lo stesso, ora associato a una
    quote non-phantom).
    """
    from app.services.rbac import current_user_optional, has_permission
    from app.models import PhantomStatus
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di promuovere quote")

    q = (
        db.query(Quote)
        .filter(Quote.id == quote_id, Quote.tenant_id == current_tenant_id())
        .first()
    )
    if not q:
        raise HTTPException(404, "Quotazione non trovata")
    if not q.is_phantom:
        raise HTTPException(400, "Solo Quotazioni a Consuntivo possono essere promosse")
    if q.phantom_status != PhantomStatus.standby:
        raise HTTPException(
            400,
            f"Quotazione a Consuntivo già in stato {q.phantom_status.value if q.phantom_status else 'unknown'}. "
            f"Solo standby può essere promossa."
        )

    q.is_phantom = False
    q.phantom_status = PhantomStatus.promoted
    db.commit()
    db.refresh(q)
    return {
        "id": q.id,
        "number": q.number,
        "status": q.status.value,
        "is_phantom": q.is_phantom,
        "phantom_status": q.phantom_status.value if q.phantom_status else None,
        "promoted": True,
    }


@router.post("/api/{quote_id}/merge-into/{target_id}")
async def merge_phantom_into(
    quote_id: int,
    target_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """v3.5.0-alpha.171.3 — Accorpa Quotazione a Consuntivo in altra quote.

    Effetti:
    1. Crea NUOVA VERSIONE della quote target (parent_quote_id=target_id).
       La versione clona TUTTE le righe della target.
    2. Copia le righe della Consuntivo come nuove QuoteLine sulla nuova
       versione (no parent_line_id: provengono dalla Consuntivo, non dalla
       target).
    3. Ricalcola totali della nuova versione.
    4. Marca target: status=superseded, superseded_by_id=new_version.id
       (se non già superseded).
    5. Marca Consuntivo: phantom_status=merged_into, merged_into_quote_id=new_version.id.

    Vincoli:
    - source.is_phantom=True E phantom_status=standby
    - target esiste, NON è phantom, stesso project_id
    - target.status può essere qualsiasi (anche approved — il versioning permette
      modifica con audit-trail).
    """
    from app.services.rbac import current_user_optional, has_permission
    from app.models import PhantomStatus
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di accorpare quote")

    src = (
        db.query(Quote)
        .options(joinedload(Quote.lines))
        .filter(Quote.id == quote_id, Quote.tenant_id == current_tenant_id())
        .first()
    )
    if not src:
        raise HTTPException(404, "Quotazione Consuntivo non trovata")
    if not src.is_phantom or src.phantom_status != PhantomStatus.standby:
        raise HTTPException(400, "Source deve essere Quotazione a Consuntivo standby")

    target = (
        db.query(Quote)
        .options(joinedload(Quote.lines))
        .filter(Quote.id == target_id, Quote.tenant_id == current_tenant_id())
        .first()
    )
    if not target:
        raise HTTPException(404, "Quotazione target non trovata")
    if target.is_phantom:
        raise HTTPException(400, "Target non può essere una Quotazione a Consuntivo")
    if target.project_id != src.project_id:
        raise HTTPException(400, "Source e target devono appartenere allo stesso progetto")
    if src.id == target.id:
        raise HTTPException(400, "Source e target non possono coincidere")

    # Costruisci nuova versione di target (clone + lines + lines Consuntivo)
    root = _quote_root(db, target)
    chain = _quote_chain(db, root)
    next_version = max(q.version for q in chain) + 1
    import re
    base_number = re.sub(r"-v\d+$", "", root.number)
    new_number = f"{base_number}-v{next_version}"

    if (
        db.query(Quote)
        .execution_options(include_deleted=True)
        .filter(Quote.number == new_number)
        .first()
    ):
        raise HTTPException(409, f"Numero '{new_number}' già esistente (anche cestino)")

    new_q = Quote(
        number=new_number,
        version=next_version,
        parent_quote_id=target.id,
        project_id=target.project_id,
        client_id=target.client_id,
        title=target.title,
        status=QuoteStatus.draft,
        issue_date=date.today(),
        valid_until=target.valid_until,
        production_material=target.production_material,
        length_minutes=target.length_minutes,
        fps=target.fps,
        delivery_format=target.delivery_format,
        shooting_days=target.shooting_days,
        shooting_format=target.shooting_format,
        package_discount=target.package_discount,
        category_discounts=dict(target.category_discounts) if target.category_discounts else None,
        category_order=list(target.category_order) if target.category_order else None,
        vat_rate=target.vat_rate,
        notes=(target.notes or "") + f"\n[α.171.3 merge] Voci accorpate da Quotazione a Consuntivo {src.number}",
        payment_terms=target.payment_terms,
        tenant_id=current_tenant_id(),
    )
    db.add(new_q); db.flush()

    # Copia voci dal target (track parent per rebind futuro)
    new_lines_target = _copy_quote_lines(target.lines, new_q.id, track_parent=True)
    db.add_all(new_lines_target)
    db.flush()

    # Copia voci dalla Consuntivo (NO parent_line_id: provengono da phantom)
    new_lines_phantom = _copy_quote_lines(src.lines, new_q.id, track_parent=False)
    # Marca le righe da phantom con note "[da Consuntivo]" se non già
    for nl in new_lines_phantom:
        if not nl.detail or "[da Consuntivo]" not in nl.detail:
            nl.detail = (nl.detail or "") + f"\n[da Consuntivo {src.number}]"
    db.add_all(new_lines_phantom)
    db.flush()

    _recalc_quote(new_q)

    # Marca target come superseded
    target.superseded_by_id = new_q.id
    if target.status != QuoteStatus.superseded:
        target.status = QuoteStatus.superseded

    # Marca Consuntivo come merged_into
    src.phantom_status = PhantomStatus.merged_into
    src.merged_into_quote_id = new_q.id

    db.commit()
    db.refresh(new_q); db.refresh(src); db.refresh(target)
    return {
        "new_version_id": new_q.id,
        "new_version_number": new_q.number,
        "new_version_status": new_q.status.value,
        "target_id": target.id,
        "target_number": target.number,
        "target_status": target.status.value,
        "source_id": src.id,
        "source_number": src.number,
        "source_phantom_status": src.phantom_status.value,
        "lines_from_target": len(new_lines_target),
        "lines_from_phantom": len(new_lines_phantom),
    }


def _build_migration_preview(db: Session, new_q: Quote) -> dict:
    """Analisi pre-migrazione tra V_old (= new_q.parent_quote) e V_new (= new_q).
    Identifica righe nuove, modificate, orfane. Evidenzia righe orfane con
    quantity_actual > 0 o booking done (lavoro registrato).
    """
    if not new_q.parent_quote_id:
        raise HTTPException(400, "Quote senza parent: niente da migrare")
    old_q = db.query(Quote).options(
        joinedload(Quote.lines), joinedload(Quote.job)
    ).filter(Quote.id == new_q.parent_quote_id).first()
    if not old_q:
        raise HTTPException(400, "Parent quote non trovata")

    job = old_q.job
    # Costline mappato per quote_line_id sorgente
    cost_by_old_line = {}
    if job:
        for jcl in job.cost_lines:
            if jcl.quote_line_id:
                cost_by_old_line[jcl.quote_line_id] = jcl

    new_lines_by_parent = {l.parent_line_id: l for l in new_q.lines if l.parent_line_id}
    orphans = []        # righe presenti in V_old ma non più in V_new
    inherited = []      # righe V_new con parent → re-bind ok
    fresh = []          # righe V_new senza parent → nuove pure
    overruns = []       # righe V_new con quantity_quoted < quantity_actual già registrato

    for ol in old_q.lines:
        if ol.id not in new_lines_by_parent:
            jcl = cost_by_old_line.get(ol.id)
            has_actual = bool(jcl and (jcl.quantity_actual or 0) > 0)
            orphans.append({
                "old_line_id": ol.id,
                "description": ol.description,
                "quantity": ol.quantity,
                "unit": ol.unit,
                "total": ol.total,
                "has_jobcostline": jcl is not None,
                "jobcostline_id": jcl.id if jcl else None,
                "quantity_actual": (jcl.quantity_actual if jcl else 0) or 0,
                "has_actual_work": has_actual,
            })

    for nl in new_q.lines:
        if nl.parent_line_id:
            inherited.append({
                "new_line_id": nl.id,
                "old_line_id": nl.parent_line_id,
                "description": nl.description,
            })
            jcl = cost_by_old_line.get(nl.parent_line_id)
            if jcl and (jcl.quantity_actual or 0) > (nl.quantity or 0):
                overruns.append({
                    "new_line_id": nl.id,
                    "description": nl.description,
                    "quantity_quoted_new": nl.quantity,
                    "quantity_actual": jcl.quantity_actual,
                    "delta": (jcl.quantity_actual or 0) - (nl.quantity or 0),
                })
        else:
            fresh.append({
                "new_line_id": nl.id,
                "description": nl.description,
                "quantity": nl.quantity,
            })

    return {
        "old_quote_id": old_q.id,
        "old_quote_number": old_q.number,
        "new_quote_id": new_q.id,
        "new_quote_number": new_q.number,
        "has_job": job is not None,
        "job_id": job.id if job else None,
        "job_code": job.code if job else None,
        "inherited": inherited,
        "fresh": fresh,
        "orphans": orphans,
        "overruns": overruns,
        "has_blockers": False,  # per ora nessun blocker hard, solo avvisi
    }


@router.get("/api/{quote_id}/migrate-preview")
async def migrate_preview(quote_id: int, db: Session = Depends(get_db)):
    """Anteprima della migrazione job da V_old a V_new (questo quote_id)."""
    new_q = (
        db.query(Quote)
        .options(joinedload(Quote.lines))
        .filter(Quote.id == quote_id).first()
    )
    if not new_q:
        raise HTTPException(404)
    return _build_migration_preview(db, new_q)


@router.post("/api/{quote_id}/migrate-job")
async def migrate_job(
    quote_id: int,
    request: Request,
    orphan_strategy: str = Form("keep_as_extra"),
    db: Session = Depends(get_db),
):
    """Applica la migrazione del Job da V_old a V_new (= quote_id).

    `orphan_strategy`:
      - `keep_as_extra` (default): le JobCostLine orfane restano sul Job ma
        vengono marcate `is_extra=True` e perdono `quote_line_id` (SET NULL).
        Il Job rimane legato a V_new; le righe orfane producono extra in
        consuntivo per riconciliazione finance.
      - `floating_job`: il Job viene scollegato (quote_id=NULL) e diventa un
        "Floating Job" gestito in `/financial`. La V_new resta non legata; il
        ciclo di vita del Job va riconciliato manualmente (riassegnazione a un
        nuovo progetto/quote o chiusura).

    Effetti:
      - V_new.status = approved
      - V_old.status = superseded; V_old.superseded_by_id = V_new.id
      - Job.quote_id = V_new.id (a meno di floating_job)
      - JobCostLine: re-bind via QuoteLine.parent_line_id; quote_line_id rimappato
        sull'id della riga V_new corrispondente.
    """
    from app.services.rbac import current_user_optional, has_permission
    user = current_user_optional(request)
    if not has_permission(user, "edit_quotes"):
        raise HTTPException(403, "Non hai il permesso di migrare le quotazioni")

    if orphan_strategy not in ("keep_as_extra", "floating_job"):
        raise HTTPException(400, f"orphan_strategy non valido: {orphan_strategy}")

    new_q = (
        db.query(Quote)
        .options(joinedload(Quote.lines), joinedload(Quote.parent_quote))
        .filter(Quote.id == quote_id).first()
    )
    if not new_q:
        raise HTTPException(404)
    if not new_q.parent_quote_id:
        raise HTTPException(400, "Quote senza parent: nessuna migrazione possibile")
    if new_q.status not in (QuoteStatus.draft, QuoteStatus.sent):
        raise HTTPException(400, f"V_new in stato {new_q.status.value}: non migrabile")

    old_q = (
        db.query(Quote)
        .options(joinedload(Quote.lines), joinedload(Quote.job))
        .filter(Quote.id == new_q.parent_quote_id).first()
    )
    if not old_q:
        raise HTTPException(400, "Parent quote non trovata")

    job = old_q.job
    # Map old_line_id → new_line per re-bind
    new_line_by_parent = {l.parent_line_id: l for l in new_q.lines if l.parent_line_id}

    job_action = None
    cost_lines_rebound = 0
    cost_lines_orphaned = 0
    cost_lines_created = 0
    deliverables_rebound = 0
    deliverables_orphaned = 0
    deliverables_created = 0

    if job:
        # v3.5.0-alpha.172.18 — branching JCL vs Deliverable per nature.
        # Le voci time-based (hr/day) sincronizzano JobCostLine; le voci
        # non-time (pc/TB/lump/...) sincronizzano JobDeliverable. La logica
        # parent_line_id resta la chiave di re-bind in entrambi i casi.
        from app.services.cost_line_sync import unit_nature_for as _unfor
        from app.models import (
            JobDeliverable as _JD, DeliverableUnitNature as _DUN,
            DeliverableBillingStatus as _DBS,
        )
        TIME_UNITS = ("hr", "day")
        SPAWN_PER_UNIT_NATURES = ("deliverable_qty",)

        # Re-bind dei JobCostLine via parent_line_id (time-based)
        for jcl in list(job.cost_lines):
            if jcl.is_extra:
                continue  # extra puri non toccati
            if jcl.quote_line_id and jcl.quote_line_id in new_line_by_parent:
                # v3.5.0-alpha.172.36 (Sprint 2 BLOCCO 2) — slice-lock guard:
                # versioning quote che ribinda una JCL già fatturata farebbe
                # divergere prezzo/descrizione vs snapshot fattura. Va emessa
                # NC TD04 prima, oppure il versioning lascia "as-is" le JCL
                # bloccate (qui scegliamo HARD-BLOCK 409 dell'intera operazione
                # di migrate, l'utente vede l'elenco e agisce).
                assert_jcl_lock_safe(
                    db, jcl.id,
                    action="ribindare la voce di costo in nuova versione di quote per",
                )
                # Re-bind: punta alla riga V_new corrispondente
                new_line = new_line_by_parent[jcl.quote_line_id]
                jcl.quote_line_id = new_line.id
                # Aggiorna i campi "pianificati" da V_new (descrizione, quantity_quoted, ecc.)
                jcl.description = new_line.description
                jcl.price_item_id = new_line.price_item_id
                jcl.quantity_quoted = new_line.quantity
                jcl.unit = new_line.unit
                jcl.unit_price = new_line.unit_price
                jcl.total_quoted = new_line.total
                # quantity_actual NON tocco (è effettivo registrato)
                cost_lines_rebound += 1
            else:
                # Orfano: la riga V_old non esiste più in V_new
                if orphan_strategy == "keep_as_extra":
                    jcl.quote_line_id = None
                    jcl.is_extra = True
                cost_lines_orphaned += 1

        # v3.5.0-alpha.172.18 — Re-bind dei JobDeliverable via parent_line_id.
        # Stessa logica di JCL: re-bind + sync campi pianificati. Per spawn-per-unit
        # (deliverable_qty), NON ri-sincronizziamo qty_planned (resta 1.0 per row).
        existing_deliverables = db.query(_JD).filter(_JD.job_id == job.id).all()
        for d in existing_deliverables:
            if d.quote_line_id and d.quote_line_id in new_line_by_parent:
                new_line = new_line_by_parent[d.quote_line_id]
                d.quote_line_id = new_line.id
                d.name = new_line.description
                d.price_item_id = new_line.price_item_id
                d.unit = new_line.unit
                d.unit_price = new_line.unit_price or 0.0
                # quantity_delivered NON tocco. qty_planned sync solo se nessun delivered
                # e nature non spawn-per-unit (per qty discreti ogni row resta a 1.0).
                if (d.quantity_delivered or 0.0) == 0.0 and \
                   _unfor(new_line.unit) not in SPAWN_PER_UNIT_NATURES:
                    d.quantity_planned = float(new_line.quantity or 0.0)
                d.total_quoted = round((d.quantity_planned or 0.0) * (d.unit_price or 0.0), 2)
                deliverables_rebound += 1
            elif d.quote_line_id:
                # Orfano: riga V_old non più in V_new. NON cancello (potrebbero esserci
                # asset linkati o conferme parziali). Soft-detach: quote_line_id = NULL.
                if orphan_strategy == "keep_as_extra":
                    d.quote_line_id = None
                deliverables_orphaned += 1

        # Crea JobCostLine + JobDeliverable per righe NUOVE (V_new ma non V_old).
        existing_new_line_ids_jcl = {jcl.quote_line_id for jcl in job.cost_lines if jcl.quote_line_id}
        existing_new_line_ids_dlv = {d.quote_line_id for d in existing_deliverables if d.quote_line_id}
        for nl in new_q.lines:
            unit_l = (nl.unit or "").strip().lower()
            if unit_l in TIME_UNITS:
                if nl.id not in existing_new_line_ids_jcl:
                    db.add(JobCostLine(
                        job_id=job.id,
                        quote_line_id=nl.id,
                        price_item_id=nl.price_item_id,
                        description=nl.description,
                        quantity_quoted=nl.quantity,
                        unit=nl.unit,
                        unit_price=nl.unit_price,
                        total_quoted=nl.total,
                        total_expected=nl.total,
                    ))
                    cost_lines_created += 1
            else:
                if nl.id not in existing_new_line_ids_dlv:
                    nature_code = _unfor(nl.unit)
                    nature = _DUN(nature_code)
                    qty_total = float(nl.quantity or 0.0)
                    up = float(nl.unit_price or 0.0)
                    if nature_code in SPAWN_PER_UNIT_NATURES:
                        n_rows = max(1, int(round(qty_total)))
                        per_row_qty = 1.0
                    else:
                        n_rows = 1
                        per_row_qty = qty_total if qty_total > 0 else 1.0
                    # v3.5.0-alpha.172.93 (Bundle K2) — auto-classify digital/physical
                    pi_rebind = db.query(PriceItem).filter(PriceItem.id == nl.price_item_id).first() if nl.price_item_id else None
                    phys_nature = _infer_deliverable_nature(pi_rebind)
                    for _idx in range(n_rows):
                        db.add(_JD(
                            tenant_id=new_q.tenant_id,
                            job_id=job.id,
                            quote_line_id=nl.id,
                            price_item_id=nl.price_item_id,
                            name=nl.description,
                            nature=phys_nature,
                            unit=nl.unit,
                            unit_price=up,
                            unit_nature=nature,
                            quantity_planned=per_row_qty,
                            quantity_delivered=0.0,
                            total_quoted=round(per_row_qty * up, 2),
                            total_accrued=0.0,
                            total_cost_accrued=0.0,
                            billing_status=_DBS.not_billed,
                        ))
                        deliverables_created += 1

        if orphan_strategy == "floating_job":
            job.quote_id = None
            job_action = "floated"
        else:
            job.quote_id = new_q.id
            job_action = "rebound"
        # Aggiorna budget_quoted al nuovo totale
        job.budget_quoted = new_q.total_after_discount

    # Aggiorna stati delle quote
    new_q.status = QuoteStatus.approved
    old_q.status = QuoteStatus.superseded
    old_q.superseded_by_id = new_q.id

    db.commit()

    return {
        "ok": True,
        "old_quote_id": old_q.id,
        "old_quote_number": old_q.number,
        "old_quote_status": old_q.status,
        "new_quote_id": new_q.id,
        "new_quote_number": new_q.number,
        "new_quote_status": new_q.status,
        "job_action": job_action,  # "rebound" | "floated" | None
        "job_id": job.id if job else None,
        "cost_lines_rebound": cost_lines_rebound,
        "cost_lines_orphaned": cost_lines_orphaned,
        "cost_lines_created": cost_lines_created,
        "deliverables_rebound": deliverables_rebound,
        "deliverables_orphaned": deliverables_orphaned,
        "deliverables_created": deliverables_created,
        "orphan_strategy": orphan_strategy,
    }
