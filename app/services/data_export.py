"""Admin export service (v3.5.0-alpha.34).

Genera un archivio ZIP con:
- mediaflow.db                  (sempre)
- metadata.json                 (sempre — versione app, schema, opzioni usate)
- README.md                     (sempre — istruzioni di restore)
- claude-memory/*.md            (sempre — memorie Claude della macchina)
- excel/listino.xlsx            (sempre — listino multi-sheet)
- excel/quotazioni.xlsx         (sempre — una sheet per quote attive)
- env/.env                      (opt-in, default OFF)
- uploads/                      (opt-in, default OFF — può essere pesante)
- trash/<dump>                  (opt-in, default OFF — record soft-deleted)

Se `password` è fornita, l'archivio è AES-encrypted (pyzipper, ZIP standard
apribile da 7zip/WinZip con la password). Senza password = ZIP normale.
"""
from __future__ import annotations

import json
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from sqlalchemy.orm import Session

# pyzipper è richiesto solo se l'utente attiva la cifratura. Lazy import
# nella funzione che lo usa per non rompere il boot se la dep manca.

from app.config import settings
from app.models import (
    PriceItem, PriceCategory, Department, Quote, QuoteLine, Project, Client,
)


# ── Helper path mangling per memorie Claude ──────────────────

def _claude_memory_path() -> Optional[Path]:
    """Risolve la cartella memorie Claude per il progetto corrente.

    Pattern osservato: `~/.claude/projects/<mangled>/memory/` dove
    `<mangled>` è il path assoluto della cartella progetto con tutti i
    separatori (`:`, `\\`, `/`) e gli underscore (`_`) sostituiti da `-`.

    Esempio Win: C:\\Users\\frico\\OneDrive\\Documents\\Claude\\Projects\\mediaflow_fase1bis
    →           C--Users-frico-OneDrive-Documents-Claude-Projects-mediaflow-fase1bis
    """
    project_root = Path(__file__).resolve().parent.parent.parent
    abs_str = str(project_root)
    mangled = abs_str.replace(":", "-").replace("\\", "-").replace("/", "-").replace("_", "-")
    candidate = Path.home() / ".claude" / "projects" / mangled / "memory"
    return candidate if candidate.exists() else None


def _mangled_root_for_path(project_path: Path) -> str:
    """Versione standalone del mangling (usata anche da data_import per
    ricalcolare il path su una macchina diversa)."""
    s = str(project_path.resolve())
    return s.replace(":", "-").replace("\\", "-").replace("/", "-").replace("_", "-")


# ── Excel exporters ───────────────────────────────────────────

def _build_listino_xlsx(db: Session, dest: Path) -> None:
    """Excel multi-sheet del listino: PriceItems, Categories, Departments."""
    wb = Workbook()

    ws = wb.active
    ws.title = "PriceItems"
    ws.append([
        "id", "name", "description", "category", "department",
        "unit_pre", "unit",
        "price_list", "price_average", "price_low", "hardcosts",
        "keywords", "is_active",
    ])
    cats = {c.id: c.name for c in db.query(PriceCategory).all()}
    deps = {d.id: d.name for d in db.query(Department).all()}
    for it in db.query(PriceItem).order_by(PriceItem.id).all():
        ws.append([
            it.id, it.name, it.description or "",
            cats.get(it.category_id, ""),
            deps.get(it.department_id, "") if it.department_id else "",
            it.unit_pre, it.unit,
            it.price_list, it.price_average, it.price_low, it.hardcosts,
            ", ".join(it.keywords or []) if it.keywords else "",
            "Sì" if it.is_active else "No",
        ])

    ws2 = wb.create_sheet("Categories")
    ws2.append(["id", "name", "order_index"])
    for c in db.query(PriceCategory).order_by(PriceCategory.id).all():
        ws2.append([c.id, c.name, getattr(c, "order_index", 0) or 0])

    ws3 = wb.create_sheet("Departments")
    ws3.append(["id", "name", "color"])
    for d in db.query(Department).order_by(Department.id).all():
        ws3.append([d.id, d.name, getattr(d, "color", "") or ""])

    wb.save(str(dest))


def _build_quotazioni_xlsx(db: Session, dest: Path) -> None:
    """Excel multi-sheet delle quote attive: una sheet 'Quotes' con header e
    una sheet 'Lines' con tutte le righe (linkate via quote_id)."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Quotes"
    ws.append([
        "id", "number", "title", "project_code", "project_title", "client_name",
        "issue_date", "valid_until", "status",
        "subtotal", "package_discount", "vat_rate", "total",
        "payment_terms", "notes",
    ])
    qs = db.query(Quote).order_by(Quote.id).all()
    proj = {p.id: p for p in db.query(Project).all()}
    cli = {c.id: c for c in db.query(Client).all()}
    for q in qs:
        p = proj.get(q.project_id) if q.project_id else None
        c = cli.get(p.client_id) if (p and p.client_id) else None
        ws.append([
            q.id, q.number or "", q.title or "",
            p.code if p else "", p.title if p else "",
            c.name if c else "",
            q.issue_date.isoformat() if q.issue_date else "",
            q.valid_until.isoformat() if q.valid_until else "",
            q.status.value if hasattr(q.status, "value") else str(q.status or ""),
            getattr(q, "subtotal", 0) or 0,
            getattr(q, "package_discount", 0) or 0,
            getattr(q, "vat_rate", 0) or 0,
            getattr(q, "total", 0) or 0,
            getattr(q, "payment_terms", "") or "",
            (q.notes or "")[:500],
        ])

    ws2 = wb.create_sheet("Lines")
    ws2.append([
        "quote_id", "quote_number", "section", "category_override",
        "price_item_id", "description", "quantity", "unit", "unit_price",
        "subtotal", "is_optional", "section_label",
    ])
    for q in qs:
        for ln in (q.lines or []):
            ws2.append([
                q.id, q.number or "",
                getattr(ln, "section", "") or "",
                getattr(ln, "category_override", "") or "",
                ln.price_item_id, ln.description or "",
                ln.quantity, ln.unit, ln.unit_price,
                getattr(ln, "subtotal", 0) or 0,
                "Sì" if getattr(ln, "is_optional", False) else "No",
                getattr(ln, "section_label", "") or "",
            ])

    wb.save(str(dest))


# ── Trash dump ────────────────────────────────────────────────

def _dump_trash_to_dir(db: Session, dest_dir: Path) -> None:
    """Scrive `dest_dir/trash.json` con i record soft-deleted delle entità
    che hanno `is_deleted` o `trashed_at`. Best-effort, non rompe se
    l'attributo non esiste su un modello."""
    from app.models import models as M
    out: dict = {}
    candidates = [
        ("clients", Client),
        ("projects", Project),
        ("quotes", Quote),
        ("quote_lines", QuoteLine),
        ("price_items", PriceItem),
    ]
    for label, Model in candidates:
        try:
            # execution_options(include_deleted=True) se è il pattern del progetto
            q = db.query(Model)
            try:
                q = q.execution_options(include_deleted=True)
            except Exception:
                pass
            rows = q.all()
            trashed = [
                {
                    "id": r.id,
                    "_attrs": {
                        k: (v.isoformat() if hasattr(v, "isoformat") else v)
                        for k, v in vars(r).items()
                        if not k.startswith("_") and not hasattr(v, "__call__")
                    },
                }
                for r in rows
                if (
                    getattr(r, "is_deleted", False) or getattr(r, "trashed_at", None)
                )
            ]
            if trashed:
                out[label] = trashed
        except Exception as e:
            out[label] = {"error": str(e)}
    dest_dir.mkdir(parents=True, exist_ok=True)
    (dest_dir / "trash.json").write_text(
        json.dumps(out, indent=2, default=str), encoding="utf-8"
    )


def _dump_pricelist_snapshots_to_dir(db: Session, dest_dir: Path) -> None:
    """Esporta tutti i PricelistSnapshot manuali (no auto, no preset) come .json.

    Solo i manual: gli `auto` sono backup transitori e i `preset` sono già
    committati nel repo.
    """
    from app.models import PricelistSnapshot, PricelistSnapshotKind
    dest_dir.mkdir(parents=True, exist_ok=True)
    snaps = (
        db.query(PricelistSnapshot)
        .filter(
            PricelistSnapshot.kind == PricelistSnapshotKind.manual,
            PricelistSnapshot.deleted_at.is_(None),
        )
        .order_by(PricelistSnapshot.created_at.desc())
        .all()
    )
    index = []
    for s in snaps:
        safe_stem = "".join(c if c.isalnum() or c in "-_" else "_" for c in s.name)[:80] or f"snapshot-{s.id}"
        ts = s.created_at.strftime("%Y%m%d-%H%M%S") if s.created_at else "snapshot"
        fname = f"{safe_stem}-{ts}.json"
        (dest_dir / fname).write_text(
            json.dumps(s.payload_json, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        index.append({
            "id": s.id,
            "name": s.name,
            "description": s.description,
            "kind": s.kind.value,
            "item_count": s.item_count,
            "category_count": s.category_count,
            "department_count": s.department_count,
            "schema_version": s.schema_version,
            "created_at": s.created_at.isoformat() + "Z" if s.created_at else None,
            "filename": fname,
        })
    (dest_dir / "_index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── Builder principale ────────────────────────────────────────

def build_export_zip(
    db: Session,
    *,
    include_env: bool = False,
    include_uploads: bool = False,
    include_trash: bool = False,
    include_memory: bool = True,
    include_listino_snapshots: bool = True,
    password: Optional[str] = None,
    app_version: str = "?",
) -> tuple[bytes, str]:
    """Costruisce l'export ZIP. Ritorna (bytes, suggested_filename).

    Se `password` è fornita, usa pyzipper con AES-256. Senza, ZIP normale.
    Il caller è responsabile di servire i bytes via FastAPI Response.
    """
    project_root = Path(__file__).resolve().parent.parent.parent
    db_path = project_root / "mediaflow.db"

    with tempfile.TemporaryDirectory() as tmp:
        tmp_root = Path(tmp)
        staging = tmp_root / "staging"
        staging.mkdir()

        # 1) DB
        if db_path.exists():
            shutil.copy2(db_path, staging / "mediaflow.db")
        else:
            (staging / "mediaflow.db.MISSING").write_text(
                "Il file mediaflow.db non è stato trovato al momento dell'export.\n",
                encoding="utf-8",
            )

        # 2) metadata.json
        meta = {
            "app": "Claqo",
            "app_version": app_version,
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "options": {
                "include_env": include_env,
                "include_uploads": include_uploads,
                "include_trash": include_trash,
                "include_memory": include_memory,
                "include_listino_snapshots": include_listino_snapshots,
                "encrypted": bool(password),
            },
            "source_machine": {
                "project_root": str(project_root),
                "mangled_memory_root": _mangled_root_for_path(project_root),
            },
            "schema": {
                # placeholder — il caller può iniettare info reali se servono
                "format": "sqlite-direct",
            },
        }
        (staging / "metadata.json").write_text(
            json.dumps(meta, indent=2), encoding="utf-8"
        )

        # 3) README
        (staging / "README.md").write_text(
            _README_TEMPLATE.format(version=app_version,
                                    encrypted="sì" if password else "no"),
            encoding="utf-8",
        )

        # 4) Excel
        excel_dir = staging / "excel"
        excel_dir.mkdir()
        try:
            _build_listino_xlsx(db, excel_dir / "listino.xlsx")
        except Exception as e:
            (excel_dir / "listino.ERROR.txt").write_text(str(e), encoding="utf-8")
        try:
            _build_quotazioni_xlsx(db, excel_dir / "quotazioni.xlsx")
        except Exception as e:
            (excel_dir / "quotazioni.ERROR.txt").write_text(str(e), encoding="utf-8")

        # 5) Memorie Claude
        if include_memory:
            mem_src = _claude_memory_path()
            if mem_src and mem_src.exists():
                shutil.copytree(mem_src, staging / "claude-memory")
            else:
                (staging / "claude-memory.MISSING").write_text(
                    "Cartella memorie Claude non trovata su questa macchina.\n",
                    encoding="utf-8",
                )

        # 6) .env opt-in
        if include_env:
            env_path = project_root / ".env"
            if env_path.exists():
                env_dir = staging / "env"
                env_dir.mkdir()
                shutil.copy2(env_path, env_dir / ".env")

        # 7) Uploads opt-in
        if include_uploads:
            up_src = project_root / settings.upload_dir.name \
                if not settings.upload_dir.is_absolute() else settings.upload_dir
            if up_src.exists():
                shutil.copytree(up_src, staging / "uploads", dirs_exist_ok=True)

        # 8) Trash opt-in
        if include_trash:
            try:
                _dump_trash_to_dir(db, staging / "trash")
            except Exception as e:
                (staging / "trash.ERROR.txt").write_text(str(e), encoding="utf-8")

        # 8.5) Listino snapshots opt-in (v3.5.0-alpha.66.6)
        if include_listino_snapshots:
            try:
                _dump_pricelist_snapshots_to_dir(db, staging / "listino-snapshots")
            except Exception as e:
                (staging / "listino-snapshots.ERROR.txt").write_text(
                    str(e), encoding="utf-8"
                )

        # 9) ZIP finale
        zip_path = tmp_root / "export.zip"
        if password:
            try:
                import pyzipper
            except ImportError as e:
                raise RuntimeError(
                    "pyzipper non installato — esegui `pip install -r requirements.txt` "
                    "per abilitare la cifratura ZIP. Senza password l'export funziona "
                    "comunque con zipfile standard."
                ) from e
            with pyzipper.AESZipFile(
                zip_path, "w", compression=pyzipper.ZIP_DEFLATED,
                encryption=pyzipper.WZ_AES,
            ) as zf:
                zf.setpassword(password.encode("utf-8"))
                for f in staging.rglob("*"):
                    if f.is_file():
                        zf.write(f, f.relative_to(staging))
        else:
            import zipfile
            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for f in staging.rglob("*"):
                    if f.is_file():
                        zf.write(f, f.relative_to(staging))

        data = zip_path.read_bytes()

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = ".enc.zip" if password else ".zip"
    fname = f"mediaflow-export-{app_version}-{ts}{suffix}"
    return data, fname


_README_TEMPLATE = """# MediaFlow — Export

Versione app: **{version}**
Cifrato: **{encrypted}**

## Contenuto

- `mediaflow.db` — database SQLite (autocontenuto)
- `metadata.json` — dettagli export, opzioni usate, info macchina sorgente
- `excel/` — listino e quotazioni in formato Excel (human-readable)
- `claude-memory/` — memorie Claude della macchina sorgente (se incluse)
- `env/` — file `.env` con secrets (se incluso)
- `uploads/` — asset library (se inclusi)
- `trash/` — record soft-deleted (se inclusi)
- `listino-snapshots/` — backup del listino salvati dall'utente (se inclusi),
  uno per file `.json` + `_index.json` di sintesi

## Restore

1. Ferma il server MediaFlow
2. Vai in `/settings` → tab "Dati" → "Import"
3. Carica questo file ZIP (e password se cifrato)
4. Conferma l'overwrite

Il sistema farà un backup automatico del DB attuale prima di sostituirlo.

## ATTENZIONE

Il restore SOSTITUISCE tutti i dati attuali. Esegui un export di backup
prima di importare uno snapshot di un'altra macchina.
"""
