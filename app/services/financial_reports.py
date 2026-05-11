"""v3.5.0-alpha.78 — Reportistica finanziaria.

Servizio con:
- period_summary(year, granularity) — wrapper su cashflow/forecast
- year_over_year(year_a, year_b, granularity) — comparazione %delta
- ytd_projection(year) — proiezione full-year lineare da YTD
- export_csv / export_xlsx generation
"""
from __future__ import annotations
import csv
import io
from datetime import date, datetime
from typing import Optional
from sqlalchemy.orm import Session


# Importati lazy nelle funzioni per evitare cicli

SUM_KEYS = [
    "invoiced", "paid", "outstanding",
    "supplier_billed", "supplier_paid", "supplier_due",
    "net_cashflow",
    "forecast_soft", "forecast_committed", "forecast_weighted",
    "pipeline_total", "quotes_approved", "quotes_sent", "quotes_rejected",
    "projected_cash",
]


def _cashflow_months(db: Session, year: int, project_id=None, client_id=None) -> list:
    """Riusa logica cashflow_year (sync core)."""
    from app.routers.finance import cashflow_year_sync
    res = cashflow_year_sync(year, project_id, client_id, db)
    return res.get("months", [])


def aggregate_quarters(months: list) -> list:
    out = []
    for q in range(4):
        slice_ = months[q*3:q*3+3]
        row = {"period": f"Q{q+1}", "month_start": q*3+1}
        for k in SUM_KEYS:
            row[k] = round(sum((m.get(k) or 0) for m in slice_), 2)
        out.append(row)
    return out


def aggregate_year(months: list) -> dict:
    row = {"period": "Anno"}
    for k in SUM_KEYS:
        row[k] = round(sum((m.get(k) or 0) for m in months), 2)
    return row


def year_over_year(db: Session, year_a: int, year_b: int, granularity: str = "quarter",
                   project_id=None, client_id=None) -> dict:
    """Comparazione anno-su-anno. year_a (corrente) vs year_b (ref).
    Granularity: month|quarter|year."""
    ma = _cashflow_months(db, year_a, project_id=project_id, client_id=client_id)
    mb = _cashflow_months(db, year_b, project_id=project_id, client_id=client_id)
    if granularity == "year":
        ra = [aggregate_year(ma)]; rb = [aggregate_year(mb)]
    elif granularity == "quarter":
        ra = aggregate_quarters(ma); rb = aggregate_quarters(mb)
    else:
        ra = ma; rb = mb
    rows = []
    for i, a in enumerate(ra):
        b = rb[i] if i < len(rb) else {k: 0 for k in SUM_KEYS}
        row = {"period": a.get("period") or a.get("month") or i+1}
        for k in SUM_KEYS:
            va = a.get(k) or 0
            vb = b.get(k) or 0
            delta = va - vb
            pct = (delta / vb * 100) if vb else None
            row[k + "_a"] = round(va, 2)
            row[k + "_b"] = round(vb, 2)
            row[k + "_delta"] = round(delta, 2)
            row[k + "_pct"] = round(pct, 1) if pct is not None else None
        rows.append(row)
    return {
        "year_a": year_a, "year_b": year_b,
        "granularity": granularity,
        "rows": rows,
    }


def ytd_projection(db: Session, year: int, project_id=None, client_id=None) -> dict:
    """Proiezione full-year basata sui dati YTD (mesi completi).
    Linear extrapolation: avg mensile × 12 — non considera stagionalità.
    Combina con forecast pipeline per "scenario realistic"."""
    today = date.today()
    if today.year != year:
        # Se anno passato: tutti i 12 mesi noti; se futuro: solo forecast.
        ytd_months = 12 if today.year > year else 0
    else:
        ytd_months = today.month - 1  # mesi completi (escluso corrente)
    months = _cashflow_months(db, year, project_id=project_id, client_id=client_id)
    ytd = {k: 0.0 for k in SUM_KEYS}
    forecast_remaining = {k: 0.0 for k in SUM_KEYS}
    for i, m in enumerate(months):
        if i < ytd_months:
            for k in SUM_KEYS:
                ytd[k] += (m.get(k) or 0)
        else:
            for k in ("forecast_weighted", "pipeline_total"):
                forecast_remaining[k] += (m.get(k) or 0)
    # Linear projection: (YTD / mesi_completi) × 12
    linear_full_year = {}
    if ytd_months > 0:
        for k in SUM_KEYS:
            linear_full_year[k] = round(ytd[k] / ytd_months * 12, 2)
    else:
        linear_full_year = {k: 0.0 for k in SUM_KEYS}
    # Realistic = YTD actual + forecast pesato rimanente (per revenue/paid)
    realistic = dict(ytd)
    realistic["paid"] = round(ytd["paid"] + forecast_remaining["forecast_weighted"], 2)
    realistic["forecast_weighted"] = round(forecast_remaining["forecast_weighted"], 2)
    return {
        "year": year,
        "ytd_months": ytd_months,
        "ytd": {k: round(v, 2) for k, v in ytd.items()},
        "linear_full_year": linear_full_year,
        "realistic_full_year": realistic,
        "pipeline_remaining": round(forecast_remaining["pipeline_total"], 2),
        "forecast_weighted_remaining": round(forecast_remaining["forecast_weighted"], 2),
    }


def export_csv(db: Session, year: int, granularity: str = "month",
               project_id=None, client_id=None) -> bytes:
    months = _cashflow_months(db, year, project_id=project_id, client_id=client_id)
    if granularity == "year":
        rows = [aggregate_year(months)]
    elif granularity == "quarter":
        rows = aggregate_quarters(months)
    else:
        rows = months
    buf = io.StringIO()
    buf.write("﻿")  # BOM for Excel
    w = csv.writer(buf, delimiter=";")
    w.writerow([f"MediaFlow Financial Report — Anno {year} — Granularità {granularity}"])
    w.writerow([f"Generato: {datetime.now():%Y-%m-%d %H:%M}"])
    w.writerow([])
    header = ["Periodo"] + [
        "Fatturato","Incassato","Aperto","Fatt. passive","Outflow","Scaduto",
        "Cassa netta","Forecast soft","Forecast committed","Forecast pesato",
        "Pipeline","Approved","Sent","Rejected","Cassa proiettata",
    ]
    w.writerow(header)
    for r in rows:
        period = r.get("period") or r.get("month_start") or r.get("month") or "?"
        if isinstance(period, int):
            MESI = ["Gen","Feb","Mar","Apr","Mag","Giu","Lug","Ago","Set","Ott","Nov","Dic"]
            period = MESI[period-1] if 1 <= period <= 12 else str(period)
        w.writerow([period] + [round(r.get(k) or 0, 2) for k in SUM_KEYS])
    return buf.getvalue().encode("utf-8")


def export_xlsx(db: Session, year: int, granularity: str = "month",
                project_id=None, client_id=None) -> bytes:
    """Export Excel multi-sheet: Summary + YoY + YTD Projection."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    months = _cashflow_months(db, year, project_id=project_id, client_id=client_id)
    if granularity == "year": rows = [aggregate_year(months)]
    elif granularity == "quarter": rows = aggregate_quarters(months)
    else: rows = months
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True, color="FFFFFF")
    indigo = PatternFill(start_color="6272F5", end_color="6272F5", fill_type="solid")
    ws.append([f"MediaFlow Financial Report — {year} — {granularity}"])
    ws["A1"].font = Font(bold=True, size=14, color="6272F5")
    ws.append([f"Generato {datetime.now():%Y-%m-%d %H:%M}"])
    ws.append([])
    header = ["Periodo"] + [
        "Fatturato","Incassato","Aperto","Fatt. passive","Outflow","Scaduto",
        "Cassa netta","Forecast soft","Forecast committed","Forecast pesato",
        "Pipeline","Approved","Sent","Rejected","Cassa proiettata",
    ]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = bold; cell.fill = indigo
        cell.alignment = Alignment(horizontal="center")
    MESI = ["Gen","Feb","Mar","Apr","Mag","Giu","Lug","Ago","Set","Ott","Nov","Dic"]
    for r in rows:
        period = r.get("period") or (MESI[(r.get("month") or 1) - 1] if r.get("month") else "?")
        ws.append([period] + [round(r.get(k) or 0, 2) for k in SUM_KEYS])
    # Width
    widths = [14] + [14] * len(SUM_KEYS)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    # Sheet YTD projection
    proj = ytd_projection(db, year, project_id=project_id, client_id=client_id)
    ws2 = wb.create_sheet("YTD Projection")
    ws2.append([f"YTD Projection — {year}"])
    ws2["A1"].font = Font(bold=True, size=13, color="6272F5")
    ws2.append([f"Mesi completi: {proj['ytd_months']}"])
    ws2.append([])
    ws2.append(["Metrica", "YTD", "Linear full-year", "Realistic full-year"])
    for cell in ws2[ws2.max_row]:
        cell.font = bold; cell.fill = indigo
    for k in SUM_KEYS:
        ws2.append([
            k,
            proj["ytd"].get(k, 0),
            proj["linear_full_year"].get(k, 0),
            proj["realistic_full_year"].get(k, 0),
        ])
    for i, w in enumerate([22, 18, 22, 22], start=1):
        ws2.column_dimensions[ws2.cell(row=4, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
