"""Bundle L Stack 1 — Batch parser per 17 capitolati corpus.

Pipeline:
1. Estrae testo da PDF/DOCX/TXT/XLSX
2. Chunk a ~6k token sliding window
3. Per ogni chunk: chiama AI Claude per estrarre items strutturati
4. Classifica T1/T2/T3 via keyword heuristic + AI
5. Mappa item a JSON Schema variant_v1
6. Output: docs/superpowers/specs/capitolati-parsed/<vendor>.variants.json

Usage:
    .venv/Scripts/python.exe scripts/parse_capitolati.py \\
        --corpus docs/capitolati_esempio \\
        --out docs/superpowers/specs/capitolati-parsed \\
        --schema-version v1 \\
        [--ai-provider claude] [--ai-model claude-sonnet-4-6] \\
        [--dry-run]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Optional


# ── Classificazione T1/T2/T3 via keyword heuristic ──
_T1_PATTERNS = [
    r"\b(master|imf|dcp|prores|mxf|tiff|dpx|mov|mp4)\b",
    r"\b(trailer|teaser|spot|behind[- ]the[- ]scenes|making[- ]of)\b",
    r"\b(textless|tail|head)\b",
    r"\b(audio|atmos|stems|m&e|mix|pcm|wav)\b",
    r"\b(subtitle|sub|scc|vtt|ttml|sdh|cc)\b",
    r"\b(lto|hdd|cru|tape|archive)\b",
    r"\b(stills|artwork|key art|poster)\b",
    r"\b(kdm|key delivery message)\b",
]
_T2_PATTERNS = [
    r"\b(cdl|color decision|lut|look[- ]up)\b",
    r"\b(spotting|dialogue list|dialog list)\b",
    r"\b(music cue sheet|midem|cue sheet)\b",
    r"\b(metadata template|metadata sheet)\b",
    r"\b(report|technical report)\b",
]
_T3_PATTERNS = [
    r"\b(nda|contratto|contract|legal)\b",
    r"\b(materials required|form|consent|release)\b",
    r"\b(certificazione|certification|certificate)\b",
]


def classify_item_tier(text: str) -> str:
    """Classifica un item testuale in T1 (technical) / T2 (documentation) / T3 (compilation).

    Heuristic keyword-based (case-insensitive). In caso di ambiguity privilegia T1.
    """
    t = text.lower()
    for pat in _T1_PATTERNS:
        if re.search(pat, t):
            return "t1_technical"
    for pat in _T2_PATTERNS:
        if re.search(pat, t):
            return "t2_documentation"
    for pat in _T3_PATTERNS:
        if re.search(pat, t):
            return "t3_compilation"
    return "t1_technical"  # default conservativo


def extract_text_from_file(path: str) -> str:
    """Estrae testo plain da file capitolato. Supporta TXT, PDF, DOCX.

    Ritorna stringa vuota se file mancante / formato non supportato.
    """
    p = Path(path)
    if not p.exists():
        return ""
    suffix = p.suffix.lower()
    try:
        if suffix == ".txt":
            return p.read_text(encoding="utf-8", errors="replace")
        if suffix == ".pdf":
            try:
                from pypdf import PdfReader
            except ImportError:
                return ""
            reader = PdfReader(str(p))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        if suffix == ".docx":
            try:
                from docx import Document  # python-docx
            except ImportError:
                return ""
            doc = Document(str(p))
            return "\n".join(par.text for par in doc.paragraphs)
        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError:
                return ""
            wb = load_workbook(str(p), data_only=True, read_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"=== {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    parts.append(" | ".join(str(c) if c is not None else "" for c in row))
            return "\n".join(parts)
    except Exception as e:
        print(f"[extract_text] {path}: {type(e).__name__}: {e}", file=sys.stderr)
        return ""
    return ""


def chunk_text(text: str, max_chars: int = 24000) -> list[str]:
    """Sliding window con overlap 1000 char. Default max 24k char (~6k token)."""
    if not text:
        return []
    if len(text) <= max_chars:
        return [text]
    chunks = []
    overlap = 1000
    i = 0
    while i < len(text):
        end = min(i + max_chars, len(text))
        chunks.append(text[i:end])
        if end >= len(text):
            break
        i = end - overlap
    return chunks


def vendor_from_filename(name: str) -> str:
    """Estrae vendor slug da nome file capitolato."""
    base = Path(name).stem.lower()
    base = re.sub(r"[^a-z0-9]+", "-", base)
    return base.strip("-")


def main():
    ap = argparse.ArgumentParser(description="Batch parser capitolati Bundle L Stack 1")
    ap.add_argument("--corpus", required=True, help="Directory con file capitolato")
    ap.add_argument("--out", required=True, help="Directory output JSON parsed")
    ap.add_argument("--schema-version", default="v1")
    ap.add_argument("--ai-provider", default="claude")
    ap.add_argument("--ai-model", default="claude-sonnet-4-6")
    ap.add_argument("--dry-run", action="store_true",
                    help="Estrai testo + chunk, ma NON chiama AI. Output stub JSON.")
    args = ap.parse_args()

    corpus_dir = Path(args.corpus)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not corpus_dir.exists():
        print(f"[error] corpus dir non trovata: {corpus_dir}", file=sys.stderr)
        sys.exit(1)

    files = [f for f in corpus_dir.iterdir() if f.is_file() and f.suffix.lower() in (".txt", ".pdf", ".docx", ".xlsx", ".doc")]
    print(f"[parse_capitolati] trovati {len(files)} file in {corpus_dir}")

    report_lines = ["# Capitolati parsed — Bundle L Stack 1\n", f"Corpus: `{corpus_dir}` · Schema: `{args.schema_version}`\n"]

    for f in files:
        vendor = vendor_from_filename(f.name)
        text = extract_text_from_file(str(f))
        chunks = chunk_text(text)
        print(f"  [{vendor}] {len(text):>7} char → {len(chunks)} chunk")

        variants = []
        if args.dry_run:
            # Output stub: 1 variant placeholder per chunk per smoke test
            for i, chunk in enumerate(chunks):
                snippet = chunk[:120].replace("\n", " ")
                variants.append({
                    "code": f"{vendor}-stub-{i+1}",
                    "name": f"[DRY-RUN] {vendor} chunk {i+1}",
                    "category": classify_item_tier(snippet),
                    "spec_json": {},
                    "source_capitolato": f.name,
                    "source_section": f"chunk {i+1}/{len(chunks)}",
                    "_snippet": snippet,
                })
        else:
            # Chiamata AI reale (Stack 1: implementazione minimale, prompt
            # strutturato. Capability runtime piena → Stack 5.)
            from app.services.ai_provider import get_provider
            provider = get_provider(args.ai_provider, args.ai_model)
            for i, chunk in enumerate(chunks):
                resp = provider.extract_variants_from_chunk(chunk, schema_version=args.schema_version)
                # resp: list[dict] secondo JSON Schema variant_v1
                for v in (resp or []):
                    v.setdefault("source_capitolato", f.name)
                    v.setdefault("source_section", f"chunk {i+1}/{len(chunks)}")
                    v["category"] = classify_item_tier(v.get("name", ""))
                    variants.append(v)

        out_file = out_dir / f"{vendor}.variants.json"
        out_file.write_text(json.dumps(variants, ensure_ascii=False, indent=2), encoding="utf-8")
        report_lines.append(f"- **{vendor}** ({f.name}): {len(variants)} variants extracted → `{out_file.name}`")

    (out_dir / "REPORT.md").write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    print(f"[parse_capitolati] done. Report: {out_dir / 'REPORT.md'}")


if __name__ == "__main__":
    main()
